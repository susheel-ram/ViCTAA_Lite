import sys
import types
import os
from flask import Flask, request, render_template, redirect, url_for, session, flash, make_response, send_file, send_from_directory
import mysql.connector
from mysql.connector import Error
import bcrypt
from flask_wtf.csrf import CSRFProtect, CSRFError
import platform
import re
from functools import wraps
import uuid
from flask_mail import Mail, Message
from jinja2 import Environment, FileSystemLoader
import socket
import requests
from datetime import datetime, timedelta
import random
import sqlite3
import pdfkit
from io import BytesIO
import psutil
from psutil import net_if_addrs
import json
import time
from pathlib import Path
import platform
import subprocess
from scapy.all import sniff, TCP
from urllib.parse import quote_plus
import pytz
import threading 
from urllib.parse import quote_plus
from concurrent.futures import ThreadPoolExecutor, as_completed
import concurrent.futures
from xhtml2pdf import pisa
import tkinter as tk
from tkinter import messagebox
import ftplib
from reportlab.graphics.barcode import code128
import webbrowser
from flask_socketio import SocketIO, emit
from threading import Lock
import win32com.client
from flask_sqlalchemy import SQLAlchemy  # SQLAlchemy import
from apscheduler.schedulers.background import BackgroundScheduler  # BackgroundScheduler import
from apscheduler.events import EVENT_JOB_ERROR, EVENT_JOB_EXECUTED
from flask import request, jsonify
import pandas as pd
import io
from openpyxl import Workbook
import openpyxl
from openpyxl.drawing.image import Image
import zipfile
from apscheduler.jobstores.sqlalchemy import SQLAlchemyJobStore
from flask import current_app
from openpyxl.styles import PatternFill
import win32com.client
import csv
import matplotlib
matplotlib.use('Agg')  # Use Agg backend for non-GUI environments
from pypdf import PdfReader
from reportlab.pdfgen import canvas
from pypdf import PdfWriter
from reportlab.lib.pagesizes import letter
from random import randint
# from tiny_llama_predictor import run_threat_analysis

startupinfo = subprocess.STARTUPINFO()
startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW


sys.modules['reportlab.graphics.barcode.code93'] = types.ModuleType("code93")

# Platform-specific imports
if sys.platform == "win32":
    import wmi
    import win32evtlog
    import win32api
    import pythoncom


 
app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Required for session management
csrf = CSRFProtect(app)  # Initialize CSRF protection
csrf.init_app(app)


# app.config['MAIL_SERVER'] = 'smtp.gmail.com'  # Example for Gmail SMTP
# app.config['MAIL_PORT'] = 465  # Or 587 for TLS
# app.config['MAIL_USE_SSL'] = True
# app.config['MAIL_USERNAME'] = 'bhaskarallu99@gmail.com'
# app.config['MAIL_PASSWORD'] = 'idptcgkqspprneax'
# app.config['MAIL_DEFAULT_SENDER'] = 'bhaskarallu99@gmail.com'
# mail = Mail(app)

app.config['MAIL_SERVER'] = 'smtp.gmail.com'  # Example for Gmail SMTP
app.config['MAIL_PORT'] = 465  # Or 587 for TLS
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = 'vdatasciences@gmail.com'
app.config['MAIL_PASSWORD'] = 'jfojxnggxkedfiab'
app.config['MAIL_DEFAULT_SENDER'] = 'vdatasciences@gmail.com'
mail = Mail(app) 

from dotenv import load_dotenv
load_dotenv()
# MySQL connection details
db_config = {
    'host': os.getenv('DB_HOST'),
    'port': int(os.getenv('DB_PORT')),
    'database': os.getenv('DB_NAME'),
    'user': os.getenv('DB_USER'),
    'password': os.getenv('DB_PASSWORD')
}


# Regex for validating phone number and email
phone_regex = re.compile(r'^\+?1?\d{9,15}$')
email_regex = re.compile(r'^[\w\.-]+@[\w\.-]+\.\w+$')
 
def validate_input(username, phone, email, password):
    if not username or not phone or not email or not password:
        return False
    if not phone_regex.match(phone) or not email_regex.match(email):
        return False
    return True

# def get_mac_address():
#     # Gets the MAC address of the current machine
#     mac = ':'.join(['{:02x}'.format((uuid.getnode() >> i) & 0xff)
#                     for i in range(0, 8*6, 8)][::-1])
#     return mac

def send_email_otp(recipient_email, otp):
    try:
        msg = Message(
            subject="Your Login OTP Code",
            recipients=[recipient_email],
            body=f"Your one-time password (OTP) for login is: {otp}. It expires in 5 minutes."
        )
        mail.send(msg)
        print(f"OTP sent to {recipient_email}")
    except Exception as e:
        print(f"Failed to send OTP email: {e}")

# def get_mac_address():
#     interfaces = psutil.net_if_addrs()
#     for interface_name, interface_addresses in interfaces.items():
#         for address in interface_addresses:
#             if hasattr(address, 'family') and str(address.family) == 'AddressFamily.AF_LINK':
#                 mac = address.address
#                 # Filter out empty or random virtual interfaces
#                 if mac and len(mac.split(':')) == 6 and mac != "00:00:00:00:00:00":
#                     return mac
#     return None

def get_mac_address():
    try:
        system = platform.system().lower()

        if system == 'windows':
            output = subprocess.check_output("getmac", shell=True, startupinfo=startupinfo).decode()
            # ✅ Fix regex with non-capturing group
            macs = re.findall(r'(?:[0-9A-Fa-f]{2}[-:]){5}[0-9A-Fa-f]{2}', output)
            for mac in macs:
                if not mac.startswith(("00-50-56", "00-09-0F", "02-42", "00-1C-42")):
                    return mac.replace("-", ":").lower()

        elif system in ('linux', 'darwin'):
            try:
                output = subprocess.check_output("ip link", shell=True, startupinfo=startupinfo).decode()
            except:
                output = subprocess.check_output("ifconfig", shell=True, startupinfo=startupinfo).decode()

            macs = re.findall(r'(?:[0-9a-f]{2}[:]){5}[0-9a-f]{2}', output, re.IGNORECASE)
            for mac in macs:
                if not mac.startswith(("00:50:56", "00:09:0f", "02:42", "00:1c:42")):
                    return mac.lower()

    except Exception as e:
        print(f"[ERROR] Could not get MAC address: {e}")

    return None

# Decorators for login and admin verification
def get_public_ip():
    try:
        response = requests.get('https://ipinfo.io/json', timeout=5)
        response.raise_for_status()
        ip_info = response.json()
        return ip_info.get('ip', 'Unknown IP')
    except requests.exceptions.Timeout:
        print("Error: Request timed out while retrieving public IP.")
        return "Timeout Error"
    except requests.exceptions.ConnectionError:
        print("Error: Unable to connect to the internet.")
        return "No Internet Connection"
    except requests.exceptions.RequestException as e:
        print(f"Error retrieving public IP: {e}")
        return "Unknown Error"
    
def scan_port(ip, port, open_ports):
    """Scan a single port and append to the open_ports list if open."""
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(0.1)  # Reduced timeout
            if s.connect_ex((ip, port)) == 0:  # Port is open
                open_ports.append(port)
    except Exception as e:
        print(f"Error scanning port {port}: {e}")

def scan_open_ports(ip, ports_to_scan):
    """Scan the given IP for open ports within a specified range."""
    open_ports = []
    threads = []

    for port in ports_to_scan:
        thread = threading.Thread(target=scan_port, args=(ip, port, open_ports))
        threads.append(thread)
        thread.start()

    for thread in threads:
        thread.join()

    return open_ports


def get_port_details(conn):
    if conn.status != 'LISTEN':
        return None
    try:
        process_name = psutil.Process(conn.pid).name() if conn.pid else "Unknown"
    except (psutil.NoSuchProcess, psutil.AccessDenied):
        process_name = "Unknown"
    return {
        'port': conn.laddr.port,
        'state': conn.status,
        'process': process_name
    }

def sys_open_ports():
    open_ports_local = []
    connections = psutil.net_connections(kind='inet')

    with ThreadPoolExecutor() as executor:
        future_to_conn = {executor.submit(get_port_details, conn): conn for conn in connections}
        for future in as_completed(future_to_conn):
            result = future.result()
            if result and result not in open_ports_local:
                open_ports_local.append(result)

    return open_ports_local

def is_system_using_vpn():
    vpn_keywords = ['tun', 'vpn', 'ethernet', 'openvpn', 'wireguard']  # Common VPN interface keywords
    for interface, addrs in net_if_addrs().items():
        for addr in addrs:
            if any(keyword in interface.lower() for keyword in vpn_keywords):
                return True
    return False

def is_vpn(ip):
    try:
        url = f"https://vpnapi.io/api/{ip}?key=971738d051fc4a87bbf6e961d4e13bad"
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()
        return data.get('security', {}).get('vpn', False)
    except requests.exceptions.RequestException as e:
        print(f"Error during VPN check: {e}")
        return False  # Return False if there's an error (no VPN detected)
    except ValueError as e:
        print(f"Error processing response: {e}")
        return False  # Return False if there's an issue with response data

def check_vpn_status():
    public_ip = get_public_ip()
    system_vpn_status = is_system_using_vpn()
    public_ip_vpn_status = False

    if public_ip:
        public_ip_vpn_status = is_vpn(public_ip)

    return {
        "public_ip": public_ip,
        "system_vpn_status": system_vpn_status,
        "public_ip_vpn_status": public_ip_vpn_status,
    }

# Run the check
check_vpn_status()

VPN_HISTORY_FILE = Path.home() / "vpn_history.json"

def log_vpn_usage():
    current_time = datetime.now()
    vpn_status = check_vpn_status()  # Use the existing check_vpn_status function

    # Load existing history or initialize a new list
    if VPN_HISTORY_FILE.exists():
        with open(VPN_HISTORY_FILE, "r") as file:
            vpn_history = json.load(file)
    else:
        vpn_history = []

    # Add the current record
    vpn_history.append({
        "timestamp": current_time.isoformat(),
        "vpn_status": vpn_status
    })

    # Save updated history
    with open(VPN_HISTORY_FILE, "w") as file:
        json.dump(vpn_history, file, indent=4)

# Function to get VPN history for the last 24 hours
def get_vpn_history_last_24_hours():
    if not VPN_HISTORY_FILE.exists():
        return []  # No history to return

    with open(VPN_HISTORY_FILE, "r") as file:
        vpn_history = json.load(file)

    # Filter records within the last 24 hours
    cutoff_time = datetime.now() - timedelta(hours=24)
    recent_history = [
        record for record in vpn_history
        if datetime.fromisoformat(record["timestamp"]) >= cutoff_time
    ]

    return recent_history

def run_command(command):
    """Helper function to execute a shell command and return output."""
    try:
        result = subprocess.run(command, shell=True, capture_output=True, text=True, startupinfo=startupinfo)
        return result.stdout.strip()
    except Exception as e:
        return str(e)

def get_patch_updates():
    installed_updates = []
    missing_patches = []
    os_type = platform.system()

    if os_type == "Windows":
        # Fetch installed updates
        cmd_installed = 'wmic qfe get HotFixID, InstalledOn, Description'
        installed_result = run_command(cmd_installed).splitlines()

        for line in installed_result[1:]:  # Skip header row
            parts = line.split()
            if len(parts) >= 3:
                update_id = parts[0] if parts[0].startswith("KB") else parts[1]
                installed_on = parts[-1] if len(parts) > 2 else "Unknown"
                details = " ".join(parts[2:-1]) if len(parts) > 3 else "Security Update"

                # Ensure that the update ID, date, and details are correctly mapped
                installed_updates.append({
                    "id": update_id,
                    "installed_on": installed_on,
                    "details": f"Installed: {details}",
                    "status": "Installed"
                })

        # Fetch missing patches using PowerShell
        cmd_pending = 'powershell -Command "Get-WindowsUpdate | Out-String"'
        pending_result = run_command(cmd_pending).split("\n")

        # Extract missing patches using regex
        for line in pending_result:
            match = re.search(r"(KB\d+)\s+(\d+\S*)\s+(.+)", line)
            if match:
                kb_id, size, title = match.groups()
                missing_patches.append({
                    "id": kb_id,
                    "installed_on": "Not Installed",
                    "details": f"{title} - {size}",
                    "status": "Missing Patch"
                })

    elif os_type == "Linux":
        # Fetch installed updates for Linux
        cmd_installed = 'dpkg-query -W --showformat="${Package} ${Version}\n"'  # Debian-based
        installed_result = run_command(cmd_installed).splitlines()

        for line in installed_result:
            parts = line.split()
            if len(parts) >= 2:
                installed_updates.append({
                    "id": parts[0],
                    "installed_on": "N/A",
                    "details": f"Version {parts[1]}",
                    "status": "Installed"
                })

        # Fetch missing updates for Linux
        # cmd_pending = 'apt list --upgradable 2>/dev/null | grep -E "^(.+/[a-z]+ .+)"'
        cmd_pending = "apt list --upgradable | awk -F'/' 'NR>1 {print $1,$2}'"
        pending_result = run_command(cmd_pending).split("\n")

        for line in pending_result:
            parts = line.split()
            if len(parts) >= 2:
                missing_patches.append({
                    "id": parts[0].split('/')[0],
                    "installed_on": "Not Installed",
                    "details": f"Version {parts[1]} available",
                    "status": "Missing Patch"
                })

    elif os_type == "Darwin":  # macOS
        # Fetch installed updates for macOS
        # cmd_installed = "softwareupdate --history | tail -n +2"
        cmd_installed = "softwareupdate --history | awk 'NR>1 {print $1, $2}'"
        installed_result = run_command(cmd_installed).splitlines()

        for line in installed_result:
            parts = line.split("\t")
            if len(parts) >= 2:
                installed_updates.append({
                    "id": parts[0],
                    "installed_on": parts[1] if len(parts) > 1 else "Unknown",
                    "details": "macOS Update",
                    "status": "Installed"
                })

        # Fetch missing updates for macOS
        # cmd_pending = "softwareupdate -l | grep '*'"
        cmd_pending = "softwareupdate -l | awk '/\\*/ {print substr($0, index($0,$2))}'"
        pending_result = run_command(cmd_pending).split("\n")

        for line in pending_result:
            match = re.search(r"\* (.+) \[(.*)\]", line)
            if match:
                update_name, version = match.groups()
                missing_patches.append({
                    "id": update_name,
                    "installed_on": "Not Installed",
                    "details": f"Version {version}",
                    "status": "Missing Patch"
                })

    # Print installed updates and missing patches
    print("Installed Updates:")
    for update in installed_updates:
        print(f"Update ID: {update['id']}, Installed On: {update['installed_on']}, Details: {update['details']}, Status: {update['status']}")

    print("\nMissing Patches (Pending Updates):")
    for patch in missing_patches:
        print(f"Update ID: {patch['id']}, Details: {patch['details']}, Status: {patch['status']}")

    return installed_updates + missing_patches

def get_battery_health():
    battery = psutil.sensors_battery()
    if battery:
        battery_info = {
            "percentage": battery.percent,
            "charging": battery.power_plugged,
            "time_remaining": battery.secsleft if battery.secsleft != psutil.POWER_TIME_UNLIMITED else "N/A"
        }
    else:
        battery_info = {"percentage": "N/A", "charging": False, "time_remaining": "N/A"}

    print("Battery Health Retrieved:", battery_info)  # Debugging print
    return battery_info

def get_blocked_sites(user_id):
    """Fetch user-specific and global blocked sites from the database."""
    try:
        connection = mysql.connector.connect(**db_config)
        cursor = connection.cursor(dictionary=True)

        # Fetch global blocked sites
        cursor.execute("SELECT block_url FROM blocked_url, NULL AS user_id WHERE user_id IS NULL")
        global_sites = [row["block_url"] for row in cursor.fetchall()]

        # Fetch user-specific blocked sites
        cursor.execute("SELECT block_url, user_id FROM blocked_url WHERE user_id = %s", (user_id,))
        user_sites = [row["block_url"] for row in cursor.fetchall()]

        connection.close()
        return global_sites + user_sites
    except Exception as e:
        print(f"Error fetching blocked sites: {e}")
        return []

def update_hosts_file(user_id):
    """Updates the hosts file dynamically for the currently logged-in user **only if necessary**."""
    
    if os.name == "nt":  # Windows
        hosts_path = r"C:\Windows\System32\drivers\etc\hosts"
    else:  # Linux/macOS
        hosts_path = "/etc/hosts"

    redirect_ip = "127.0.0.1"
    blocked_sites = get_blocked_sites(user_id)

    try:
        print(f"🔍 Checking hosts file for User ID: {user_id}...")

        # Read existing hosts file content
        with open(hosts_path, "r") as file:
            lines = file.readlines()

        # Extract current blocked sites from the hosts file
        current_blocked_sites = {
            line.strip().split()[1] for line in lines
            if line.startswith(redirect_ip) and len(line.split()) > 1
        }

        # Generate expected blocked sites
        expected_blocked_sites = {
            f"www.{site}" for site in blocked_sites
        } | {
            site for site in blocked_sites
        }

        # ✅ Check if update is needed
        if current_blocked_sites == expected_blocked_sites:
            print("✅ Hosts file is already up-to-date. Skipping update.")
            return  # Exit function if no update is required

        print(f"🚀 Updating hosts file for User ID: {user_id} with {len(blocked_sites)} sites.")

        # Remove all previous blocked site entries
        new_lines = [line for line in lines if not any(site in line for site in current_blocked_sites)]

        # Add new blocked sites
        for site in blocked_sites:
            site = site.replace("http://", "").replace("https://", "").replace("www.", "").replace("https://www.", "").replace(".com/", ".com")
            new_lines.append(f"{redirect_ip} www.{site}\n")
            new_lines.append(f"{redirect_ip} {site}\n")

        # Write updated content back to hosts file
        with open(hosts_path, "w") as file:
            file.writelines(new_lines)

        print(f"✅ Hosts file updated successfully for User ID: {user_id}")

    except Exception as e:
        print(f"❌ Error updating hosts file: {e}")

def update_hosts_continuously(user_id):
    """Continuously updates the hosts file for the currently logged-in user every 2 minutes."""
    while True:
        try:
            if user_id:  # Ensure user ID is valid
                update_hosts_file(user_id)
                print("UPDATEING--------------------------------------------")
            else:
                print("⚠️ No user logged in. Skipping hosts file update.")

        except Exception as e:
            print(f"❌ Error updating hosts file: {e}")

        time.sleep(120)  # Update every 2 minutes
 
@app.route('/manage_sites')
def block_urls():
    try:
        organization = session.get('organization')
        user_id = session.get('user_id')
        if not organization:
            flash("Session expired or invalid. Please log in again.")
            return redirect(url_for('login_form'))
        
        connection = mysql.connector.connect(**db_config)
        cursor = connection.cursor(dictionary=True)

        # Fetch current admin's license expiry date (same as admin_dashboard)
        cursor.execute('SELECT payment_due_date FROM users WHERE id = %s', (user_id,))
        admin_license = cursor.fetchone()
        license_expiry_date = admin_license['payment_due_date'] if admin_license else None

        # Join blocked_url with users to get both user_id and username
        query = """
            SELECT b.id, b.block_url, b.user_id, u.username
            FROM blocked_url b
            LEFT JOIN users u ON b.user_id = u.id
            WHERE u.organization = %s
        """
        cursor.execute(query, (organization,))
        urls = cursor.fetchall()

        cursor.execute("SELECT id, username FROM users WHERE organization = %s", (organization,))  # Fetch users for dropdown
        users = cursor.fetchall()

        return render_template('blocked_urls.html', urls=urls, users=users, license_expiry_date=license_expiry_date)

    except mysql.connector.Error as e:
        flash(f"Database error: {e}", "error")
        return render_template('blocked_urls.html', urls=[], users=[])
    finally:
        if 'connection' in locals() and connection.is_connected():
            cursor.close()
            connection.close()
 
# @app.route('/add', methods=['POST'])
# @csrf.exempt
# def add_url():
#     url = request.form['url'].strip()
#     all_users = request.form.get('all_users')  # Checkbox for all users
#     user_ids = request.form.getlist('user_ids[]')  # List of selected user IDs
#     timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

#     # Fetch user's IP and location (if available)
#     location_data = get_location()
#     location = f"{location_data['city']}, {location_data['country']}" if location_data else "Unknown"

#     if not url:
#         flash("Please enter a valid website URL.", "error")
#         return redirect(url_for('block_urls'))

#     try:
#         connection = mysql.connector.connect(**db_config)
#         cursor = connection.cursor()

#         if all_users:  # If "All Users" is checked
#             cursor.execute("INSERT INTO blocked_url (block_url, user_id) VALUES (%s, NULL)", (url,))
#             cursor.execute(
#                 "INSERT INTO website_audit_log (user_id, block_url, action, timestamp, location) VALUES (%s, %s, %s, %s, %s)",
#                 (None, url, "Blocked", timestamp, location)
#             )
#         else:
#             for user_id in user_ids:
#                 cursor.execute("INSERT INTO blocked_url (block_url, user_id) VALUES (%s, %s)", (url, user_id))
#                 cursor.execute(
#                     "INSERT INTO website_audit_log (user_id, block_url, action, timestamp, location) VALUES (%s, %s, %s, %s, %s)",
#                     (user_id, url, "Blocked", timestamp, location)
#                 )

#         connection.commit()
#         flash('URL added successfully! The site will be blocked when the user logs in or in 2 minutes.', 'success')

#     except mysql.connector.IntegrityError:
#         flash('URL already exists!', 'error')
#     except mysql.connector.Error as e:
#         flash(f'Database error: {e}', 'error')
#     finally:
#         if 'connection' in locals() and connection.is_connected():
#             cursor.close()
#             connection.close()

#     return redirect(url_for('block_urls'))

@app.route('/add', methods=['POST'])
@csrf.exempt
def add_url():
    url = request.form['url'].strip()
    all_users = request.form.get('all_users')  # Checkbox for all users
    user_ids = request.form.getlist('user_ids[]')  # List of selected user IDs
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Fetch user's IP and location (if available)
    location_data = get_location()
    location = f"{location_data['city']}, {location_data['country']}" if location_data else "Unknown"

    if not url:
        flash("Please enter a valid website URL.", "error")
        return redirect(url_for('block_urls'))

    try:
        # Get the admin's organization from the session
        organization = session.get('organization')
        if not organization:
            flash("Session expired or invalid. Please log in again.")
            return redirect(url_for('login_form'))

        connection = mysql.connector.connect(**db_config)
        cursor = connection.cursor(dictionary=True)  # Ensure the cursor is set to return dictionaries

        # If "All Users" is checked, insert the URL for all users in the admin's organization
        if all_users:
            cursor.execute("SELECT id FROM users WHERE organization = %s", (organization,))
            valid_user_ids = [row['id'] for row in cursor.fetchall()]  # Ensure it's returning a dictionary

            for user_id in valid_user_ids:
                cursor.execute("INSERT INTO blocked_url (block_url, user_id) VALUES (%s, %s)", (url, user_id))
                cursor.execute(
                    "INSERT INTO website_audit_log (user_id, block_url, action, timestamp, location) VALUES (%s, %s, %s, %s, %s)",
                    (user_id, url, "Blocked", timestamp, location)
                )

        else:
            # Insert the URL for selected users
            for user_id in user_ids:
                cursor.execute("INSERT INTO blocked_url (block_url, user_id) VALUES (%s, %s)", (url, user_id))
                cursor.execute(
                    "INSERT INTO website_audit_log (user_id, block_url, action, timestamp, location) VALUES (%s, %s, %s, %s, %s)",
                    (user_id, url, "Blocked", timestamp, location)
                )

        connection.commit()
        flash('URL added successfully! The site will be blocked when the user logs in or in 2 minutes.', 'success')

    except mysql.connector.IntegrityError:
        flash('URL already exists!', 'error')
    except mysql.connector.Error as e:
        flash(f'Database error: {e}', 'error')
    finally:
        if 'connection' in locals() and connection.is_connected():
            cursor.close()
            connection.close()

    return redirect(url_for('block_urls'))
 
def remove_site_from_hosts(site):
    """Removes the given site from the system's hosts file."""
    if os.name == "nt":  # Windows
        hosts_path = r"C:\Windows\System32\drivers\etc\hosts"
    else:  # Linux/macOS
        hosts_path = "/etc/hosts"
 
    redirect_ip = "127.0.0.1"
 
    try:
        # Ensure the file can be modified
        if not os.access(hosts_path, os.W_OK):
            print("Error: Administrator privileges required to modify hosts file.")
            return
 
        # Read the existing hosts file
        with open(hosts_path, "r") as file:
            lines = file.readlines()
 
        # Normalize site (remove http, https, and www)
        site = site.replace("http://", "").replace("https://", "").replace("www.", "")
 
        # Remove only the specific site's entry
        new_lines = [
            line for line in lines
            if not (site in line and redirect_ip in line)
        ]
 
        # Rewrite the hosts file with the updated content
        with open(hosts_path, "w") as file:
            file.writelines(new_lines)
 
        print(f"✅ {site} has been unblocked and removed from the hosts file.")
 
    except PermissionError:
        print("❌ Permission denied. Run as administrator/root.")
    except Exception as e:
        print(f"❌ Error: {str(e)}")
 
@app.route('/delete/<int:id>')
def delete_url(id):
    try:
        connection = mysql.connector.connect(**db_config)
        cursor = connection.cursor(dictionary=True)

        cursor.execute("SELECT block_url, user_id FROM blocked_url WHERE id=%s", (id,))
        url_to_remove = cursor.fetchone()

        if url_to_remove:
            print("Before DELETE execution")
            cursor.execute("DELETE FROM blocked_url WHERE id=%s", (id,))
            connection.commit()
            print("After DELETE execution, before INSERT execution")

            remove_site_from_hosts(url_to_remove["block_url"])

            # Fetch user's location
            location_data = get_location()
            location = f"{location_data['city']}, {location_data['country']}" if location_data else "Unknown"

            # Ensure timestamp is in correct format
            timestamp = datetime.now()

            # Insert into audit log
            cursor.execute(
                "INSERT INTO website_audit_log (user_id, block_url, action, timestamp, location) VALUES (%s, %s, %s, %s, %s)",
                (url_to_remove["user_id"], url_to_remove["block_url"], "Unblocked", timestamp, location)
            )
            connection.commit()
            print("After INSERT execution")

            flash(f"URL '{url_to_remove['block_url']}' removed successfully!", "success")
        else:
            flash("URL not found in database!", "error")

    except mysql.connector.Error as e:
        print(f"Database error: {e}")  # ✅ Debugging step
        flash(f"Database error: {e}", "error")
    finally:
        if "connection" in locals() and connection.is_connected():
            cursor.close()
            connection.close()

    return redirect(url_for("block_urls"))


def get_blocked_sites(user_id):
    """Fetch user-specific blocked sites from the database."""
    try:
        connection = mysql.connector.connect(**db_config)
        cursor = connection.cursor(dictionary=True)

        # Fetch user-specific blocked sites only (all users are treated as user-specific)
        cursor.execute("SELECT block_url FROM blocked_url WHERE user_id = %s", (user_id,))
        user_sites = [row["block_url"] for row in cursor.fetchall()]

        connection.close()

        print(f"🔍 Found {len(user_sites)} user-specific sites for User ID: {user_id}")

        return user_sites
    except Exception as e:
        print(f"❌ Error fetching blocked sites: {e}")
        return []

# def get_blocked_sites(user_id):
#     """Fetch user-specific and global blocked sites from the database."""
#     try:
#         connection = mysql.connector.connect(**db_config)
#         cursor = connection.cursor(dictionary=True)

#         # Fetch global blocked sites (applies to all users)
#         cursor.execute("SELECT block_url FROM blocked_url WHERE user_id IS NULL")
#         global_sites = [row["block_url"] for row in cursor.fetchall()]

#         # Fetch user-specific blocked sites
#         cursor.execute("SELECT block_url FROM blocked_url WHERE user_id = %s", (user_id,))
#         user_sites = [row["block_url"] for row in cursor.fetchall()]

#         connection.close()

#         blocked_sites = global_sites + user_sites
#         print(f"🔍 Found {len(global_sites)} global sites and {len(user_sites)} user-specific sites for User ID: {user_id}")

#         return blocked_sites
#     except Exception as e:
#         print(f"❌ Error fetching blocked sites: {e}")
#         return []

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            flash("You must be logged in to access this page.")
            return redirect(url_for('login_form'))
        return f(*args, **kwargs)
    return decorated_function
 
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session or session.get('role') != 'admin':
            flash("You do not have permission to access this page.")
            return redirect(url_for('login_form'))
        return f(*args, **kwargs)
    return decorated_function
 
# Function to get system information
def get_logged_in_user():
    if platform.system() == "Windows":
        # Windows: Use WMI for logged-in user
        import wmi
        c = wmi.WMI()
        for user in c.Win32_ComputerSystem():
            return user.UserName
    else:
        # Linux/macOS: Use psutil to get current user
        return os.getlogin()

def get_cpu_info():
    cpu_info = []
    if platform.system() == "Windows":
        # Windows: Use WMI to get CPU info
        import wmi
        c = wmi.WMI()
        for cpu in c.Win32_Processor():
            cpu_info.append({
                "Name": cpu.Name,
                "Cores": cpu.NumberOfCores,
                "Threads": cpu.ThreadCount,
                "MaxClockSpeed": cpu.MaxClockSpeed
            })
    else:
        # Linux/macOS: Use psutil or subprocess for CPU info
        cpu_info.append({
            "Name": platform.processor(),
            "Cores": psutil.cpu_count(logical=False),
            "Threads": psutil.cpu_count(logical=True),
            "MaxClockSpeed": subprocess.getoutput("lscpu | grep 'CPU MHz'").split(":")[1].strip() if platform.system() != "Darwin" else "N/A"
        })
    print(cpu_info)
    return cpu_info

def get_memory_info():
    memory_info = []
    if platform.system() == "Windows":
        # Windows: Use WMI to get memory info
        import wmi
        c = wmi.WMI()
        for mem in c.Win32_PhysicalMemory():
            memory_info.append({
                "Capacity": int(mem.Capacity) // (1024**3),  # Convert to GB
                "Speed": mem.Speed,
                "Manufacturer": mem.Manufacturer
            })
    else:
        # Linux/macOS: Use psutil for memory info
        virtual_memory = psutil.virtual_memory()
        memory_info.append({
            "Capacity": virtual_memory.total // (1024**3),  # Convert to GB
            "Speed": "N/A",  # Speed is not easily available on Linux/macOS via psutil
            "Manufacturer": "N/A"
        })
    return memory_info

def get_disk_info():
    disk_info = []
    if platform.system() == "Windows":
        # Windows: Use WMI to get disk info
        import wmi
        c = wmi.WMI()
        for disk in c.Win32_DiskDrive():
            disk_info.append({
                "Model": disk.Model,
                "Size": int(disk.Size) // (1024**3),  # Convert to GB
                "InterfaceType": disk.InterfaceType
            })
    else:
        # Linux/macOS: Use psutil for disk info
        for disk in psutil.disk_partitions():
            disk_info.append({
                "Model": disk.device,
                "Size": psutil.disk_usage(disk.mountpoint).total // (1024**3),  # Convert to GB
                "InterfaceType": "N/A"  # InterfaceType is not easily available on Linux/macOS
            })
    return disk_info

def get_network_info():
    network_info = []
    hostname = socket.gethostname()
    ip_address = socket.gethostbyname(hostname)
    network_info.append({
        "Hostname": hostname,
        "IPAddress": ip_address
    })
    return network_info

# Windows: Using WMI and win32evtlog
if platform.system() == "Windows":
    import wmi
    import win32evtlog

def get_usb_activity():
    usb_activity = []
    
    if platform.system() == "Windows":
        # Windows-specific implementation using win32evtlog to read USB events
        server = 'localhost'
        log_type = 'System'  # USB-related events are typically logged in the 'System' log

        try:
            handle = win32evtlog.OpenEventLog(server, log_type)
            print("Event log opened successfully.")
        except Exception as e:
            print(f"Error opening event log: {e}")
            return usb_activity  # Return empty if event log can't be opened

        flags = win32evtlog.EVENTLOG_BACKWARDS_READ | win32evtlog.EVENTLOG_SEQUENTIAL_READ

        while True:
            events = win32evtlog.ReadEventLog(handle, flags, 0)
            if not events:
                break  # Exit the loop if no events are left
            
            for event in events:
                print(f"Event ID: {event.EventID}, EventCategory: {event.EventCategory}, StringInserts: {event.StringInserts}, EventTime: {event.TimeGenerated}")
                
                if event.EventID in [20001, 20003, 20004]:  # USB connect/disconnect events
                    usb_entry = {
                        "EventID": event.EventID,
                        "EventCategory": event.EventCategory,
                        "EventTime": event.TimeGenerated.Format(),
                        "StringInserts": event.StringInserts
                    }

                    if event.EventID == 20003:
                        usb_entry["DeviceID"] = event.StringInserts[0] if len(event.StringInserts) > 0 else "Unknown"
                        usb_entry["Action"] = "USB Device Connected"
                    elif event.EventID == 20001:
                        usb_entry["Action"] = "USB Device Removed"
                    elif event.EventID == 20004:
                        usb_entry["Action"] = "USB Device Error"

                    usb_activity.append(usb_entry)

        win32evtlog.CloseEventLog(handle)
    
    elif platform.system() in ["Linux", "Darwin"]:  # macOS and Linux
        # Linux/macOS: Parse system logs (e.g., /var/log/syslog) for USB events
        log_path = "/var/log/syslog" if platform.system() == "Linux" else "/var/log/system.log"
        
        try:
            with open(log_path, "r") as log_file:
                logs = log_file.readlines()

            for line in logs:
                if "usb" in line.lower():  # Filtering USB-related logs
                    timestamp_str = line.split(" ")[0] + " " + line.split(" ")[1]  # Date format: Jan 23
                    timestamp = datetime.strptime(timestamp_str, '%b %d')
                    event_details = line.strip()

                    usb_entry = {
                        "EventTime": timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                        "Action": "USB Event",  # Simplified, adjust based on log content
                        "Details": event_details
                    }

                    usb_activity.append(usb_entry)

        except Exception as e:
            print(f"Error reading log file: {e}")
    
    return usb_activity

# Function to get USB history in the last 'days' days
def get_usb_history(usb_activity, days=10):
    current_time = datetime.now()
    cutoff_time = current_time - timedelta(days=days)

    filtered_activity = []
    for entry in usb_activity:
        # Updated format to match the timestamp
        event_time = datetime.strptime(entry['EventTime'], '%a %b %d %H:%M:%S %Y')
        if event_time >= cutoff_time:
            filtered_activity.append(entry)
    
    return filtered_activity

def get_usb_ports_count():
    if platform.system() == "Windows":
        # For Windows, use WMI as you provided
        import wmi
        c = wmi.WMI()
        usb_devices = c.query("SELECT * FROM Win32_USBHub")
        return len(usb_devices)
    
    elif platform.system() == "Linux" or platform.system() == "Darwin":  # Linux and macOS
        # For Linux (Ubuntu) or macOS, use lsusb or check the sysfs for USB devices
        usb_devices = os.popen("lsusb").readlines()
        return len(usb_devices)

    else:
        return 0  # Default if unsupported OS

def get_usb_ports_status():
    print("Checking USB ports status...")  # Debug: Starting the check
    
    if platform.system() == "Windows":
        print("Platform: Windows")  # Debug: Windows platform detected
        # For Windows, use WMI to check USB hubs
        c = wmi.WMI()
        usb_devices = c.query("SELECT * FROM Win32_USBHub")
        usb_ports_status = []
        
        if usb_devices:
            for device in usb_devices:
                # Check if the device is connected (active)
                if device.Status == "OK":  # Device status is OK
                    usb_ports_status.append({"port": device.DeviceID, "status": "opened"})
                else:
                    usb_ports_status.append({"port": device.DeviceID, "status": "closed"})
            print(f"USB devices found on Windows: {len(usb_ports_status)}")  # Debug: Number of USB devices found
        else:
            usb_ports_status = [{"port": "N/A", "status": "closed"}]
            print("No USB devices found.")

        return usb_ports_status
    
    elif platform.system() == "Linux":
        print("Platform: Linux")  # Debug
        usb_devices = os.popen("lsusb").read().strip().split("\n")
        usb_ports_status = [{"port": dev, "status": "opened"} for dev in usb_devices if dev]

        if not usb_ports_status:
            usb_ports_status = [{"port": "N/A", "status": "closed"}]
            print("No USB devices found.")

        return usb_ports_status

    elif platform.system() == "Darwin":  # macOS
        print("Platform: macOS")  # Debug
        usb_devices = os.popen("ioreg -p IOUSB -w0").read().strip().split("\n")
        usb_ports_status = [{"port": dev, "status": "opened"} for dev in usb_devices if dev]

        if not usb_ports_status:
            usb_ports_status = [{"port": "N/A", "status": "closed"}]
            print("No USB devices found.")

        return usb_ports_status

    else:
        print("Platform: Unknown")  # Debug
        return [{"port": "N/A", "status": "closed"}]

    # elif platform.system() == "Linux" or platform.system() == "Darwin":  # Linux and macOS
    #     print("Platform: Linux/macOS")  # Debug: Linux/macOS platform detected
    #     # For Linux (Ubuntu) or macOS, use lsusb to check USB devices
    #     usb_devices = os.popen("lsusb").readlines()
    #     usb_ports_status = []
        
    #     if usb_devices:
    #         for device in usb_devices:
    #             # Assume the device is "opened" if it's listed
    #             usb_ports_status.append({"port": device.strip(), "status": "opened"})
    #         print(f"USB devices found on Linux/macOS: {len(usb_ports_status)}")  # Debug: Number of USB devices found
    #     else:
    #         usb_ports_status = [{"port": "N/A", "status": "closed"}]
    #         print("No USB devices found.")

    #     return usb_ports_status
    
    # else:
    #     print("Platform: Unknown")  # Debug: Unknown platform detected
    #     return [{"port": "N/A", "status": "closed"}]


# def get_installed_software():
#     software_list = []

#     if platform.system() == "Windows":
#         try:
#             import wmi
#             import winreg

#             # Using WMI to retrieve installed software
#             c = wmi.WMI()
#             for product in c.Win32_Product():
#                 if product.Vendor and product.Vendor.lower() not in ["microsoft", "system"]:
#                     software_list.append({
#                         'name': product.Name,
#                         'version': product.Version or "Unknown",
#                         'vendor': product.Vendor
#                     })

#             # Retrieve installed applications via registry, filtering known pre-installed paths
#             uninstall_keys = [
#                 r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall",
#                 r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall"
#             ]

#             for key_path in uninstall_keys:
#                 try:
#                     reg_key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, key_path)
#                     i = 0
#                     while True:
#                         try:
#                             subkey_name = winreg.EnumKey(reg_key, i)
#                             subkey = winreg.OpenKey(reg_key, subkey_name)
#                             name, _ = winreg.QueryValueEx(subkey, "DisplayName")
#                             version, _ = winreg.QueryValueEx(subkey, "DisplayVersion")
#                             vendor = winreg.QueryValueEx(subkey, "Publisher")[0] if "Publisher" in [winreg.EnumValue(subkey, j)[0] for j in range(winreg.QueryInfoKey(subkey)[1])] else "Unknown"

#                             # Skip known system software and Microsoft apps
#                             if vendor and vendor.lower() not in ["microsoft", "system"]:
#                                 software_list.append({
#                                     'name': name,
#                                     'version': version or "Unknown",
#                                     'vendor': vendor
#                                 })
#                         except FileNotFoundError:
#                             pass  # Some entries may not have complete information
#                         except OSError:
#                             break
#                         i += 1
#                 except OSError:
#                     pass  # If the registry path is not accessible

#         except ImportError:
#             print("WMI or winreg module is not available. Ensure they're installed for Windows.")

def get_installed_software():
    software_list = []
    seen = set()

    if platform.system() == "Windows":
        try:
            import wmi
            import winreg

            c = wmi.WMI()
            for product in c.Win32_Product():
                if product.Vendor and product.Vendor.lower() not in ["microsoft", "system"]:
                    key = (product.Name.strip().lower(), product.Version or "Unknown", product.Vendor.strip().lower())
                    if key not in seen:
                        software_list.append({
                            'name': product.Name,
                            'version': product.Version or "Unknown",
                            'vendor': product.Vendor,
                            'last_used': "Unknown"
                        })
                        seen.add(key)

            uninstall_keys = [
                r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall",
                r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall"
            ]

            for key_path in uninstall_keys:
                try:
                    reg_key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, key_path)
                    i = 0
                    while True:
                        try:
                            subkey_name = winreg.EnumKey(reg_key, i)
                            subkey = winreg.OpenKey(reg_key, subkey_name)
                            name, _ = winreg.QueryValueEx(subkey, "DisplayName")
                            version, _ = winreg.QueryValueEx(subkey, "DisplayVersion")
                            vendor = winreg.QueryValueEx(subkey, "Publisher")[0] if "Publisher" in [winreg.EnumValue(subkey, j)[0] for j in range(winreg.QueryInfoKey(subkey)[1])] else "Unknown"

                            key = (name.strip().lower(), version or "Unknown", vendor.strip().lower())
                            if vendor and vendor.lower() not in ["microsoft", "system"] and key not in seen:
                                software_list.append({
                                    'name': name,
                                    'version': version or "Unknown",
                                    'vendor': vendor,
                                    'last_used': "Unknown"
                                })
                                seen.add(key)
                        except FileNotFoundError:
                            pass
                        except OSError:
                            break
                        i += 1
                except OSError:
                    pass

        except ImportError:
            print("WMI or winreg module is not available. Ensure they're installed for Windows.")

    elif platform.system() == "Linux" or platform.system() == "Darwin":  # Linux and macOS
        try:
            # Check for installed packages on Linux (Ubuntu/Debian-based) using dpkg, excluding system packages
            if platform.system() == "Linux":
                result = os.popen("dpkg-query -W -f='${Package} ${Version}\n'").readlines()
                for line in result:
                    parts = line.split()
                    if len(parts) >= 2 and not parts[0].startswith("lib"):  # Filter out common system libraries
                        software_list.append({
                            'name': parts[0],
                            'version': parts[1],
                            'vendor': "N/A"  # Vendor info isn't typically available on Linux
                        })

            # Handle Snap applications
            snap_apps = os.popen("snap list").readlines()
            for line in snap_apps[1:]:
                parts = line.split()
                if len(parts) >= 2 and parts[0] != "core":  # Exclude 'core' snap, a system package
                    software_list.append({
                        'name': parts[0],
                        'version': parts[1],
                        'vendor': "Snap Store"
                    })

            # Handle Flatpak applications
            flatpak_apps = os.popen("flatpak list --app").readlines()
            for line in flatpak_apps:
                if line.strip():
                    parts = line.split()
                    software_list.append({
                        'name': parts[0],
                        'version': "Unknown",
                        'vendor': "Flatpak"
                    })

            # macOS: Check Homebrew, skip system packages like those in `/usr/`
            if platform.system() == "Darwin":
                result = os.popen("brew list --versions").readlines()
                for line in result:
                    parts = line.split()
                    if len(parts) >= 2 and not parts[0].startswith("/"):  # Exclude system directories
                        software_list.append({
                            'name': parts[0],
                            'version': parts[1],
                            'vendor': "Homebrew"
                        })

        except Exception as e:
            print(f"Error retrieving software: {e}")

    return software_list

 
def generate_report(cpu_info, memory_info, disk_info, network_info, software_list):
    env = Environment(loader=FileSystemLoader('templates'))
    template = env.get_template('report.html')
    report = template.render(cpu_info=cpu_info, memory_info=memory_info,
                             disk_info=disk_info, network_info=network_info,
                             software_list=software_list)
    with open('report.html', 'w') as f:
        f.write(report)
       
# New Scanning Functions
 
# 2. Vulnerability Lookup for Installed Software on Laptops or Desktops
def fetch_vulnerability_data(software, nvd_api_url, printed_messages, vulnerabilities):
    software_name = software.get('name')
    software_version = software.get('version')

    if software_version == 'None':
        message = f"Skipping {software_name} (vNone) due to missing version information."
        if message not in printed_messages:
            print(message)
            printed_messages.add(message)
        return None

    forbidden_software_list = [
        "Python 3.12.0",  # Known software causing issues
        "None",           # Skip software with missing version info
    ]

    if software_name and any(forbidden in software_name for forbidden in forbidden_software_list):
        message = f"Skipping {software_name} due to known issues."
        if message not in printed_messages:
            print(message)
            printed_messages.add(message)
        return None

    if software_name is None:
        message = "Skipping entry with missing software name."
        if message not in printed_messages:
            print(message)
            printed_messages.add(message)
        return None

    query = f"{software_name} {software_version}"
    query = quote_plus(query)

    retries = 1
    for _ in range(retries):
        try:
            response = requests.get(nvd_api_url, params={'keyword': query})
            response.raise_for_status()

            if response.status_code == 200:
                result = response.json()
                if 'result' in result and result['result'].get('CVE_Items'):
                    vulnerabilities[software_name] = [
                        {
                            "cve_id": item['cve']['CVE_data_meta']['ID'],
                            "description": item['cve']['description']['description_data'][0]['value'],
                            "severity": item.get('impact', {}).get('baseMetricV2', {}).get('severity', 'Unknown'),
                        }
                        for item in result['result']['CVE_Items']
                    ]
                    return software_name  # Return the software name to indicate success
        except requests.exceptions.HTTPError as http_err:
            if response.status_code == 404:
                message = f"No CVE data found for {software_name} (v{software_version})."
                if message not in printed_messages:
                    print(message)
                    printed_messages.add(message)
                return None
            if response.status_code == 403:
                message = f"Access Forbidden for {software_name} (v{software_version}), skipping..."
                if message not in printed_messages:
                    print(message)
                    printed_messages.add(message)
                return None
            message = f"HTTP error occurred while fetching data for {software_name} (v{software_version}): {http_err}"
            if message not in printed_messages:
                print(message)
                printed_messages.add(message)
            time.sleep(2)
        except requests.exceptions.RequestException as req_err:
            message = f"Request exception occurred: {req_err}"
            if message not in printed_messages:
                print(message)
                printed_messages.add(message)
            time.sleep(2)
        except Exception as err:
            message = f"Other error occurred: {err}"
            if message not in printed_messages:
                print(message)
                printed_messages.add(message)
            break

    return None

def vulnerability_lookup_for_desktop(software_list):
    nvd_api_url = "https://services.nvd.nist.gov/rest/json/cves/2.0"
    vulnerabilities = {}
    printed_messages = set()  # To keep track of printed messages

    with ThreadPoolExecutor() as executor:
        futures = []
        for software in software_list:
            futures.append(executor.submit(fetch_vulnerability_data, software, nvd_api_url, printed_messages, vulnerabilities))

        # Wait for all threads to complete and process their results
        for future in as_completed(futures):
            result = future.result()
            if result:
                print(f"Processed {result}")

    return vulnerabilities

# Define restricted file types
RESTRICTED_EXTENSIONS = ['.mov', '.mp4', '.mp3']

# Function to find restricted files in multiple directories
def find_restricted_files(directories):
    restricted_files = []
    for directory in directories:
        for root, dirs, files in os.walk(directory):
            for file in files:
                # Convert file extension to lowercase for case-insensitive comparison
                if any(file.lower().endswith(ext) for ext in RESTRICTED_EXTENSIONS):
                    restricted_files.append(os.path.join(root, file))
    return restricted_files

# List of common cache file extensions
CACHE_EXTENSIONS = ['.cache', '.tmp', '.dat', '.log', '.bak', '.index', '.sqlite', '.cookie']

# Function to search for cache files in specific directories
def find_cache_files(directories):
    cache_files = []
    
    # Platform-specific program directories
    if platform.system() == 'Windows':
        program_dirs = ['C:\\Program Files', 'C:\\Program Files (x86)']
    elif platform.system() == 'Linux' or platform.system() == 'Darwin':  # macOS is Darwin in platform.system()
        program_dirs = ['/usr/local', '/opt']  # Common locations for Linux/macOS
    else:
        program_dirs = []

    directories.extend(program_dirs)

    # Walk through all directories and look for cache files
    for directory in directories:
        for root, dirs, files in os.walk(directory):
            for file in files:
                if any(file.lower().endswith(ext) for ext in CACHE_EXTENSIONS):
                    file_path = os.path.join(root, file)
                    cache_files.append(file_path)
                
    return cache_files

# 3. Basic Threat Detection (Example: TCP SYN packets)
# Define known malicious processes for system-based threats
KNOWN_MALICIOUS_PROCESSES = ['malware.exe', 'virus.exe']

# Vulnerability detection function (software-based threat)
def detect_software_threats(software_list):
    threats = []
    nvd_api_url = "https://services.nvd.nist.gov/rest/json/cves/2.0"

    # A set to keep track of printed messages to avoid printing them more than once
    printed_messages = set()

    # Function to process each software item
    def process_software(software):
        software_name = software['name']
        software_version = software['version']
        query = f"{software_name} {software_version}"

        # URL encode the query to handle special characters
        query = quote_plus(query)

        # Retry mechanism with a single retry
        retry_attempted = False
        software_threats = []

        try:
            while True:
                try:
                    response = requests.get(nvd_api_url, params={'keyword': query})
                    response.raise_for_status()  # Will raise HTTPError for 4xx/5xx errors

                    if response.status_code == 200:
                        result = response.json()
                        if 'result' in result and result['result'].get('CVE_Items'):
                            # Detect vulnerabilities in installed software
                            for item in result['result']['CVE_Items']:
                                cve_id = item['cve']['CVE_data_meta']['ID']
                                description = item['cve']['description']['description_data'][0]['value']
                                severity = item.get('impact', {}).get('baseMetricV2', {}).get('severity', 'Unknown')
                                software_threats.append(f"Vulnerability detected in {software_name} (v{software_version}): {description} - CVE ID: {cve_id} - Severity: {severity}")
                        else:
                            message = f"No CVE data found for {software_name} (v{software_version})"
                            if message not in printed_messages:
                                print(message)
                                printed_messages.add(message)
                        break  # Exit loop after successful processing
                except requests.exceptions.HTTPError as http_err:
                    if response.status_code == 404:
                        message = f"No CVE data found for {software_name} (v{software_version})"
                        if message not in printed_messages:
                            print(message)
                            printed_messages.add(message)
                        break  # No CVE data; no need to retry
                    elif response.status_code == 403:
                        message = f"Access Forbidden for {software_name} (v{software_version}), skipping..."
                        if message not in printed_messages:
                            print(message)
                            printed_messages.add(message)
                        break  # Access forbidden; no need to retry
                    message = f"HTTP error occurred for {software_name} (v{software_version}): {http_err}"
                    if message not in printed_messages:
                        print(message)
                        printed_messages.add(message)
                except requests.exceptions.RequestException as req_err:
                    message = f"Request exception occurred for {software_name} (v{software_version}): {req_err}"
                    if message not in printed_messages:
                        print(message)
                        printed_messages.add(message)
                except Exception as err:
                    message = f"Unexpected error occurred for {software_name} (v{software_version}): {err}"
                    if message not in printed_messages:
                        print(message)
                        printed_messages.add(message)
                    break  # Exit loop on unexpected errors

                # Handle retry logic
                if not retry_attempted:
                    retry_attempted = True
                    retry_message = f"Retrying for {software_name} (v{software_version})..."
                    if retry_message not in printed_messages:
                        print(retry_message)
                        printed_messages.add(retry_message)
                    time.sleep(2)  # Wait before retrying
                else:
                    break  # Exit loop after a single retry
        except Exception as outer_err:
            outer_message = f"Final error occurred for {software_name} (v{software_version}): {outer_err}"
            if outer_message not in printed_messages:
                print(outer_message)
                printed_messages.add(outer_message)

        return software_threats

    # Use ThreadPoolExecutor to process the software list in parallel
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = [executor.submit(process_software, software) for software in software_list]
        
        for future in concurrent.futures.as_completed(futures):
            threats.extend(future.result())  # Collect results as they are completed

    return threats


# System-based threat detection function (no changes needed here)
def detect_system_threats():
    threats = []

    # Check for high CPU usage (threshold set to 90%)
    cpu_usage = psutil.cpu_percent(interval=1)
    if cpu_usage > 90:
        threats.append(f"High CPU usage detected: {cpu_usage}%")
    
    # Check for high memory usage (threshold set to 80%)
    memory_usage = psutil.virtual_memory().percent
    if memory_usage > 80:
        threats.append(f"High memory usage detected: {memory_usage}%")
    
    # Check for suspicious processes (e.g., known malicious processes)
    for proc in psutil.process_iter(['pid', 'name']):
        if proc.info['name'].lower() in KNOWN_MALICIOUS_PROCESSES:
            threats.append(f"Malicious process detected: {proc.info['name']} (PID {proc.info['pid']})")
    
    # If threats are detected, return the list of threats
    if threats:
        return threats
    else:
        # If no threats are detected, return a healthy system message
        return []

# Function to check for restricted files (software-based threat)
def scan_for_malicious_code(file_path):
    # Simulate scanning logic - you can replace this with actual scanning techniques or libraries
    # For example, check for embedded scripts or suspicious patterns
    malicious_patterns = ['<script>', 'eval(', 'base64', '.exe', '.vbs']  # List of suspicious patterns
    try:
        with open(file_path, 'rb') as file:
            content = file.read()
            # Check if any malicious pattern is found in the file
            if any(pattern.encode() in content for pattern in malicious_patterns):
                return True
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")
    return False

def find_restricted_files(directories, restricted_extensions=['.mov', '.mp4', '.mp3']):
    restricted_files = []

    for directory in directories:
        for root, dirs, files in os.walk(directory):
            for file in files:
                # Check for restricted file extensions
                if any(file.endswith(ext) for ext in restricted_extensions):
                    file_path = os.path.join(root, file)
                    # Scan the file for malicious code
                    if scan_for_malicious_code(file_path):
                        restricted_files.append(f"Malicious content detected in restricted file: {file_path}")
                    else:
                        restricted_files.append(file_path)
    
    # Return both restricted files (safe ones) and threats (malicious ones)
    return restricted_files

# Asynchronous command execution (doesn't wait for the process to complete)
def run_command_async(command):
    process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, startupinfo=startupinfo)
    return process

def run_command(command):
    result = subprocess.run(command, shell=True, capture_output=True, text=True, startupinfo=startupinfo)
    return result.stdout.strip()

def check_system_updates():
    threats = []

    try:
        system = platform.system()
        print(f"Detected platform: {system}")

        if system == "Windows":
            check_policy_command = 'powershell -Command "Get-ExecutionPolicy -Scope CurrentUser"'
            current_policy = run_command(check_policy_command)
            if current_policy != "RemoteSigned":
                set_policy_command = (
                    'powershell -Command "Set-ExecutionPolicy RemoteSigned -Scope CurrentUser -Force"'
                )
                threading.Thread(target=run_command_async, args=(set_policy_command,)).start()
                print("Set PowerShell execution policy to RemoteSigned.")

            check_module_command = 'powershell -Command "Get-Module -ListAvailable PSWindowsUpdate"'
            module_installed = run_command(check_module_command)
            if not module_installed:
                install_command = (
                    'powershell -Command '
                    '"Install-Module -Name PSWindowsUpdate -Force -Scope CurrentUser; '
                    'Import-Module -Name PSWindowsUpdate"'
                )
                threading.Thread(target=run_command_async, args=(install_command,)).start()
                print("Ensured PSWindowsUpdate module is installed and imported.")

            process = run_command_async('powershell -Command "Get-WindowsUpdate | Out-String"')
            stdout, stderr = process.communicate()
            result = stdout.decode('utf-8').strip()  # Decode the bytes to string
            print(f"Windows Update Output: {result}")

            if "CommandNotFoundException" in result or "The term" in result:
                print("Windows Update command is not recognized.")
            elif "No updates" not in result and result:
                print("Detected pending updates on Windows.")
                updates = result.splitlines()
                if updates:
                    last_update = updates[-1].strip()
                    print(f"Last Update Line: {last_update}")
                    threats.append(f"Pending system updates detected: {last_update}")
            else:
                print("No updates detected on Windows.")

        elif system == "Linux":
            process = run_command_async("apt list --upgradable 2>/dev/null")
            stdout, stderr = process.communicate()
            result = stdout.decode('utf-8').strip()  # Decode the bytes to string
            print(f"Linux Update Output: {result}")

            if "upgradable" in result:
                print("Detected pending updates on Linux.")
                updates = result.splitlines()
                if updates:
                    last_update = updates[-1].strip()
                    print(f"Last Update Line: {last_update}")
                    threats.append(f"Pending system updates detected: {last_update}")
            else:
                print("No updates detected on Linux.")

        elif system == "Darwin":
            process = run_command_async("softwareupdate -l")
            stdout, stderr = process.communicate()
            result = stdout.decode('utf-8').strip()  # Decode the bytes to string
            print(f"macOS Update Output: {result}")

            if "No new software" not in result and result:
                print("Detected pending updates on macOS.")
                updates = result.splitlines()
                if updates:
                    last_update = updates[-1].strip()
                    print(f"Last Update Line: {last_update}")
                    threats.append(f"Pending system updates detected: {last_update}")
            else:
                print("No updates detected on macOS.")

    except Exception as e:
        print(f"Error occurred: {str(e)}")
        threats.append(f"Failed to check updates: {str(e)}")

    return threats

# Example usage:
updates = check_system_updates()
print(updates)


# Combined threat detection function
def detect_all_threats(cpu_info, memory_info, software_list, directories):
    threats = []
    
    # Detect system-based threats
    system_threats = detect_system_threats()
    threats.extend(system_threats)
    
    # Detect software-based threats (vulnerabilities)
    software_threats = detect_software_threats(software_list)
    threats.extend(software_threats)

    update_threats = check_system_updates()
    threats.extend(update_threats)

    return threats

# def get_application_paths():
#     """Retrieve application paths based on the operating system, excluding non-executable system files like desktop.ini."""
#     os_name = platform.system()
#     application_paths = []

#     # Define unwanted files to ignore, such as desktop.ini
#     unwanted_files = ["desktop.ini"]

#     if os_name == "Windows":
#         # Example directories for Windows applications
#         program_files = os.getenv("ProgramFiles")
#         program_files_x86 = os.getenv("ProgramFiles(x86)")
#         app_data = os.getenv("APPDATA")

#         # Adding typical application directories
#         if program_files:
#             application_paths.extend([os.path.join(program_files, app) for app in os.listdir(program_files) if app.lower() not in unwanted_files and os.path.isfile(os.path.join(program_files, app))])
#         if program_files_x86:
#             application_paths.extend([os.path.join(program_files_x86, app) for app in os.listdir(program_files_x86) if app.lower() not in unwanted_files and os.path.isfile(os.path.join(program_files_x86, app))])
#         if app_data:
#             application_paths.extend([os.path.join(app_data, app) for app in os.listdir(app_data) if app.lower() not in unwanted_files and os.path.isfile(os.path.join(app_data, app))])

#     elif os_name in ["Linux", "Darwin"]:  # Darwin is macOS
#         # Example directories for Linux/macOS applications
#         application_dirs = ["/usr/bin", "/usr/local/bin", os.path.expanduser("~/.local/share")]
#         for directory in application_dirs:
#             if os.path.exists(directory):
#                 application_paths.extend([os.path.join(directory, app) for app in os.listdir(directory) if app.lower() not in unwanted_files and os.path.isfile(os.path.join(directory, app))])

#     return application_paths

def get_application_paths():
    """Retrieve application paths based on the operating system, excluding non-executable system files like desktop.ini."""
    os_name = platform.system()
    application_paths = set()  # Use a set to store unique application paths

    # Define unwanted files to ignore, such as desktop.ini
    unwanted_files = ["desktop.ini"]

    if os_name == "Windows":
        # Example directories for Windows applications
        program_files = os.getenv("ProgramFiles")
        program_files_x86 = os.getenv("ProgramFiles(x86)")
        app_data = os.getenv("APPDATA")
        app_data_local = os.getenv("LOCALAPPDATA")  # Added LOCALAPPDATA

        # Adding typical application directories
        if program_files:
            for app in os.listdir(program_files):
                app_path = os.path.join(program_files, app)
                if app.lower() not in unwanted_files and os.path.isfile(app_path):
                    application_paths.add(clean_app_name(app.lower()))  # Ensure unique apps

        if program_files_x86:
            for app in os.listdir(program_files_x86):
                app_path = os.path.join(program_files_x86, app)
                if app.lower() not in unwanted_files and os.path.isfile(app_path):
                    application_paths.add(clean_app_name(app.lower()))  # Ensure unique apps

        if app_data:
            for app in os.listdir(app_data):
                app_path = os.path.join(app_data, app)
                if app.lower() not in unwanted_files and os.path.isfile(app_path):
                    application_paths.add(clean_app_name(app.lower()))  # Ensure unique apps

        if app_data_local:
            for app in os.listdir(app_data_local):
                app_path = os.path.join(app_data_local, app)  # Now including LOCALAPPDATA
                if app.lower() not in unwanted_files and os.path.isfile(app_path):
                    application_paths.add(clean_app_name(app.lower()))  # Ensure unique apps

    elif os_name in ["Linux", "Darwin"]:  # Darwin is macOS
        # Example directories for Linux/macOS applications
        application_dirs = ["/usr/bin", "/usr/local/bin", os.path.expanduser("~/.local/share")]
        for directory in application_dirs:
            if os.path.exists(directory):
                for app in os.listdir(directory):
                    app_path = os.path.join(directory, app)
                    if app.lower() not in unwanted_files and os.path.isfile(app_path):
                        application_paths.add(clean_app_name(app.lower()))  # Ensure unique apps

    return application_paths

def check_file_access_time(file_path):
    """Check and log the last access time of a known file for debugging."""
    try:
        last_access_time = datetime.fromtimestamp(os.stat(file_path).st_atime)  # Access time (st_atime)
        print(f"Debug: File {file_path}, Last Access Time: {last_access_time}")
        return last_access_time
    except Exception as e:
        print(f"Error accessing file {file_path}: {e}")
        return None

def check_unused_apps(cutoff_time):
    unused_apps = []

    for proc in psutil.process_iter(['pid', 'name', 'create_time']):
        try:
            if proc.info['name'] == "System Idle Process":  # Exclude System Idle Process
                continue
            
            last_used_time = datetime.fromtimestamp(proc.info['create_time'])

            if last_used_time < cutoff_time:
                unused_apps.append(proc.info['name'])

        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue

    return list(set(unused_apps))  # Remove duplicates


# List of known system processes to exclude
SYSTEM_PROCESSES = [
    "sihost.exe", "svchost.exe", "explorer.exe", "winlogon.exe", 
    "dwm.exe", "conhost.exe", "ctfmon.exe", "fontdrvhost.exe", 
    "WmiPrvSE.exe", "RuntimeBroker.exe", "SearchProtocolHost.exe",
    "SystemSettingsBroker.exe", "audiodg.exe", "taskhostw.exe",
    "ShellExperienceHost.exe", "backgroundTaskHost.exe", "csrss.exe",
    "dllhost.exe", "TextInputHost.exe", "WidgetService.exe"
]

# def get_installed_apps():
#     """Retrieve installed application paths."""
#     os_name = platform.system()
#     application_paths = []

#     if os_name == "Windows":
#         program_files = os.getenv("ProgramFiles")
#         program_files_x86 = os.getenv("ProgramFiles(x86)")

#         # Fetch installed applications from Program Files directories
#         if program_files:
#             application_paths.extend(os.listdir(program_files))
#         if program_files_x86:
#             application_paths.extend(os.listdir(program_files_x86))

#     return set(application_paths)  # Ensure unique application names

import winreg
def clean_app_name(name):
    return name.strip().lower()

def is_component_app(name):
    """Filter out component entries like test suites, path tools, docs, etc."""
    keywords = [
        "documentation", "add to path", "test suite", "utility scripts",
        "executables", "core interpreter", "redistributables"
    ]
    name_lower = name.lower()
    return any(keyword in name_lower for keyword in keywords)

def get_registry_installed_apps():
    installed_apps = set()
    seen_descriptions = set()
    
    registry_paths = [
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall"),
        (winreg.HKEY_CURRENT_USER, r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall"),
        (winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall")
    ]

    for hive, path in registry_paths:
        try:
            with winreg.OpenKey(hive, path) as key:
                for i in range(winreg.QueryInfoKey(key)[0]):
                    try:
                        subkey_name = winreg.EnumKey(key, i)
                        with winreg.OpenKey(key, subkey_name) as subkey:
                            try:
                                name = winreg.QueryValueEx(subkey, "DisplayName")[0]
                            except FileNotFoundError:
                                continue

                            try:
                                version = winreg.QueryValueEx(subkey, "DisplayVersion")[0]
                            except FileNotFoundError:
                                version = "Unknown"

                            try:
                                vendor = winreg.QueryValueEx(subkey, "Publisher")[0]
                            except FileNotFoundError:
                                vendor = "Unknown"

                            full_description = f"{name} (Version: {version}, Vendor: {vendor})"
                            cleaned = clean_app_name(full_description)

                            if cleaned and cleaned not in seen_descriptions and not is_component_app(name):
                                installed_apps.add(cleaned)
                                seen_descriptions.add(cleaned)

                    except (FileNotFoundError, PermissionError, OSError, TypeError):
                        continue
        except FileNotFoundError:
            continue

    return installed_apps

def get_installed_apps():
    """Retrieve installed application names (file system + registry)."""
    application_paths = set()
    seen_descriptions = set()
    os_name = platform.system()

    if os_name == "Windows":
        for env_var in ["ProgramFiles", "ProgramFiles(x86)", "APPDATA", "LOCALAPPDATA"]:
            base_path = os.getenv(env_var)
            if base_path:
                try:
                    for app in os.listdir(base_path):
                        app_path = os.path.join(base_path, app)
                        if os.path.isdir(app_path):
                            cleaned = clean_app_name(app)
                            if cleaned and cleaned not in seen_descriptions:
                                application_paths.add(cleaned)
                                seen_descriptions.add(cleaned)
                except Exception:
                    continue

        registry_apps = get_registry_installed_apps()
        for app in registry_apps:
            if app not in seen_descriptions:
                application_paths.add(app)
                seen_descriptions.add(app)

    return application_paths

SYSTEM_PROCESSES = [
    "System", "Registry", "services.exe", "smss.exe", "wininit.exe",
    "lsass.exe", "WUDFHost.exe", "IntelCpHDCPSvc.exe", "SecurityHealthService.exe",
    "FMService64.exe", "AutoModeDetect.exe", "MsMpEng.exe", "MemCompression",
    "igfxCUIServiceN.exe", "spoolsv.exe", "MfeAVSvc.exe", "jhi_service.exe",
    "MMSSHOST.exe", "wlanext.exe", "LNBITSSvc.exe", "mfemms.exe",
    "Lenovo.Modern.ImController.exe", "LenovoUtilityService.exe",
    "OneApp.IGCC.WinService.exe", "IntelAudioService.exe", "OfficeClickToRun.exe",
    "LMS.exe", "PEFService.exe", "RstMwService.exe", "SessionService.exe",
    "UDClientService.exe", "WMIRegistrationService.exe", "esif_uf.exe", "mfevtps.exe",
    "csrss.exe","svchost.exe","fontdrvhost.exe","msedgewebview2.exe","CrossDeviceService.exe",
    "RuntimeBroker.exe","LenovoVantageService.exe","Widgets.exe","DAX3API.exe",
    "MpDefenderCoreService.exe","LenovoVantage-(SmartDisplayAddin).exe","msedgewebview2.exe",
    "ctfmon.exe","conhost.exe","ModuleCoreService.exe","RtkAudUService64.exe",
    "uihost.exe","backgroundTaskHost.exe","ProtectedModuleHost.exe","Lenovo.Modern.ImController.PluginHost.Device.exe",
    "NisSrv.exe","unsecapp.exe","WmiPrvSE.exe","WmiApSrv.exe","AggregatorHost.exe","StartMenuExperienceHost.exe",
    "AppProvisioningPlugin.exe","McCSPServiceHost.exe","LenovoVantage-(DeviceSettingsSystemAddin).exe",
    "winlogon.exe","pet.exe","FnHotkeyUtility.exe","servicehost.exe","SearchIndexer.exe",
    "mcapexe.exe","PresentationFontCache.exe","TextInputHost.exe","UserOOBEBroker.exe","LenovoVantage-(GenericMessagingAddin).exe",
    "FnHotkeyCapsLKNumLK.exe","LockApp.exe","FileCoAuth.exe","ApplicationFrameHost.exe","SearchProtocolHost.exe",
    "SystemSettingsBroker.exe","taskhostw.exe","LenovoVantage-(VantageCoreAddin).exe","SearchHost.exe","ShellExperienceHost.exe",
    "WidgetService.exe","LocationNotificationWindows.exe","SecurityHealthSystray.exe","Locator.exe","PhoneExperienceHost.exe",
    "audiodg.exe","sihost.exe","explorer.exe","dwm.exe","dllhost.exe","smartscreen.exe","McUICnt.exe","igfxEMN.exe",
    "mcshield.exe","powershell.exe"
]

RESTRICTED_APPS = ["whatsapp.exe", "telegram.exe", "tiktok.exe", "discord.exe","pubg.exe","valorant.exe","minecraft.exe","tinder.exe",
                   "netflix.exe","amazon.exe","prime.exe","hotstar.exe","bumble.exe","hinge.exe","snapchat.exe","spotify.exe","slims.exe"]  # Define restricted apps

def clean_app_name(app_name):
    """Remove .exe from app names."""
    return app_name.rsplit('.exe', 1)[0]  # Removes .exe only if it exists

def parse_time_spent(seconds):
    """Convert seconds to HH:MM:SS format."""
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"

def check_recently_used_apps(cutoff_time):
    """Track recently used main processes within the last 24 hours, summing only the latest 5 instances."""
    recently_used_apps = {}

    # Temporary dictionary to store all instances before selecting the latest 5
    process_instances = {}

    for proc in psutil.process_iter(['name', 'create_time']):
        try:
            app_name = proc.info['name']  # Get process name
            if not app_name:  # Skip if app name is empty or None
                continue

            app_name = app_name.lower()  # Convert to lowercase for comparison
            process_start_time = datetime.fromtimestamp(proc.info['create_time'])

            # Skip system processes
            if app_name in [name.lower() for name in SYSTEM_PROCESSES]:
                continue

            # Ignore apps that started before the last 24 hours
            if process_start_time < cutoff_time:
                continue

            # Calculate time spent since process was started
            time_spent = (datetime.now() - process_start_time).total_seconds()

            # Store process instance
            if app_name not in process_instances:
                process_instances[app_name] = []
            process_instances[app_name].append((process_start_time, time_spent))

        except (psutil.NoSuchProcess, psutil.AccessDenied):
            continue

    # Process only the latest 5 instances for each app
    for app_name, instances in process_instances.items():
        # Sort instances by start time (newest first)
        instances.sort(reverse=True, key=lambda x: x[0])

        # Take only the latest 5 instances
        latest_5_instances = instances[:5]

        # Sum up the time spent for these instances
        total_time_spent = sum(instance[1] for instance in latest_5_instances)

        # Store in the recently used apps dictionary
        clean_name = clean_app_name(app_name)  # Remove .exe
        recently_used_apps[clean_name] = {
            "last_used": latest_5_instances[0][0].strftime("%Y-%m-%d %H:%M:%S"),  # Most recent instance
            "time_spent": parse_time_spent(min(total_time_spent, 24 * 3600)),  # Format as HH:MM:SS
            "highlight": clean_name in [clean_app_name(app) for app in RESTRICTED_APPS]  # Highlight if restricted
        }

    return recently_used_apps

def check_unused_apps(installed_apps, recently_used_apps):
    """
    Identify unused apps from the installed apps list by comparing with recently used apps.
    """
    recently_used_app_names = set(recently_used_apps.keys())

    unused_apps = []
    for app in installed_apps:
        app_lower = app.lower()
        if app_lower not in recently_used_app_names:
            unused_apps.append({
                "name": app,
                "restricted": app_lower in [app.lower() for app in RESTRICTED_APPS]  # Check if restricted
            })

    return unused_apps

# Define cutoff times
cutoff_time_24_hours = datetime.now() - timedelta(hours=24)
cutoff_time_7_days = datetime.now() - timedelta(days=7)

# Get all installed applications
installed_apps = get_installed_apps()

# Get recently used apps within the last 7 days
recently_used_apps = check_recently_used_apps(cutoff_time_24_hours)

# Identify unused apps by passing both arguments
unused_apps = check_unused_apps(installed_apps, recently_used_apps)

# Output results
print("Recently Used Apps:", recently_used_apps)
print("Unused Apps:", unused_apps)


# Function to Get Device Type
def get_location():
    """Fetch user's location details based on public IP."""
    try:
        response = requests.get("https://ipinfo.io/json")  # Fetch data from ipinfo.io
        if response.status_code == 200:
            data = response.json()
            return {
                "public_ip": data.get("ip", "Unknown IP"),
                "city": data.get("city", "Unknown City"),
                "region": data.get("region", "Unknown Region"),
                "country": data.get("country", "Unknown Country"),
                "latitude_longitude": data.get("loc", "Unknown Coordinates"),
                "isp": data.get("org", "Unknown ISP")
            }
    except Exception as e:
        print(f"Error fetching location: {e}")
    return {
        "public_ip": "Unknown",
        "city": "Unknown",
        "region": "Unknown",
        "country": "Unknown",
        "latitude_longitude": "Unknown",
        "isp": "Unknown"
    }

# def get_device_type():
#     """Fetch system details including OS, hostname, and MAC address."""
#     return {
#         "os_name": platform.system(),  # Windows, Linux, macOS
#         "os_version": platform.version(),
#         "hostname": socket.gethostname(),
#         "processor": platform.processor(),
#         "architecture": platform.architecture()[0],
#         "mac_address": ':'.join(['{:02x}'.format((uuid.getnode() >> i) & 0xff) for i in range(0, 48, 8)])
#     }

def check_bitlocker_status():
    try:
        output = subprocess.check_output("manage-bde -status", shell=True, startupinfo=startupinfo).decode()
        return "Protection On" in output
    except Exception as e:
        print(f"BitLocker check failed: {e}")
        return False

def is_firewall_enabled():
    try:
        output = subprocess.check_output("netsh advfirewall show allprofiles", shell=True, startupinfo=startupinfo).decode()
        return "State ON" in output
    except Exception as e:
        print(f"Firewall check failed: {e}")
        return False


def get_device_type():
    """Fetch system details including OS, hostname, MAC address, and security posture."""
    try:
        os_name = platform.system()
        os_version = platform.version()
        hostname = socket.gethostname()
        processor = platform.processor()
        architecture = platform.architecture()[0]
        mac_address = ':'.join(['{:02x}'.format((uuid.getnode() >> i) & 0xff) for i in range(0, 48, 8)])

        print(f"[DEBUG] OS: {os_name} {os_version}, Hostname: {hostname}, MAC: {mac_address}")

        # Additional Security Controls
        bitlocker = check_bitlocker_status()
        firewall = is_firewall_enabled()
        defender = any("MsMpEng.exe" in p.name() for p in psutil.process_iter())

        print(f"[DEBUG] BitLocker: {bitlocker}, Firewall: {firewall}, Windows Defender: {defender}")

        return {
            "os_name": os_name,
            "os_version": os_version,
            "hostname": hostname,
            "processor": processor,
            "architecture": architecture,
            "mac_address": mac_address,
            "security_controls": {
                "BitLocker": bitlocker,
                "Firewall": firewall,
                "WindowsDefender": defender
            }
        }
    except Exception as e:
        print(f"[ERROR] get_device_type() failed: {e}")
        return {}

@app.route('/get-login-data')
def get_login_data():

    organization = session.get('organization')
    if not organization:
        return jsonify({"error": "Session expired or invalid."}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("SELECT login_status, COUNT(*) AS count FROM users WHERE organization = %s GROUP BY login_status", (organization,))
    status_counts = cursor.fetchall()

    cursor.execute("SELECT id, username FROM users WHERE login_status = 'logged_in' AND organization = %s", (organization,))
    logged_in_users = cursor.fetchall()

    cursor.execute("SELECT id, username FROM users WHERE login_status = 'logged_out' AND organization = %s", (organization,))
    logged_out_users = cursor.fetchall()

    conn.close()

    if not status_counts:
        return jsonify({"error": "No data found"}), 404

    return jsonify({
        'status_counts': status_counts,
        'logged_in_users': logged_in_users,
        'logged_out_users': logged_out_users
    })

 
# Routes
@app.route('/')
def index():
    # print('fxghjoiyfgjhvkbll')
    return redirect(url_for('login_form'))

def open_browser():
    # This will open the browser automatically
    webbrowser.open("http://127.0.0.1:5001")

def validate_input(username, phone, email, password):
    # Null check
    if not username or not phone or not email or not password:
        return False
    # Username: alphanumeric with _, ., @, +, -
    if not re.match(r'^[\w.@+-]+$', username):
        return False
    # Phone: international pattern (basic)
    if not re.match(r'^\+?\d{9,15}$', phone):
        return False
    # Email: basic format validation
    if not re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
        return False
    # Password: at least 6 characters (you can expand this to include complexity)
    if len(password) < 6:
        return False
    return True

@app.route('/register', methods=['GET'])
def register_form():
    return render_template('register.html')
 
@app.route('/register', methods=['POST'])
def register_user():
    username = request.form['username'].strip()
    phone = request.form['phone'].strip()
    organization = request.form['organization'].strip()
    email = request.form['email'].strip()
    license_key = request.form['license_key'].strip()
    password = request.form['password']
 
    # Validate user input
    if not validate_input(username, phone, email, password) or not license_key:
        flash("Invalid input. Please check your data and try again.")
        return redirect(url_for('register_form'))
 
    # Hash the password
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    mac_address = get_mac_address()  # Get MAC address at registration time
    print(f"[DEBUG] Captured MAC at Registration: {mac_address}")

    try:
        connection = mysql.connector.connect(**db_config)
        if connection.is_connected():
            cursor = connection.cursor(dictionary=True)

            # Check for duplicate username, phone, or email
            duplicate_check_query = """
            SELECT * FROM users 
            WHERE username = %s OR phone = %s OR email = %s
            """
            cursor.execute(duplicate_check_query, (username, phone, email))
            existing_user = cursor.fetchone()

            if existing_user:
                # Adjusted indices based on table structure (id, username, phone, email)
                if existing_user[1] == username:
                    flash("Username already exists. Please choose another one.")
                elif existing_user[2] == phone:
                    flash("Phone number already exists. Please use a different number.")
                elif existing_user[3] == email:
                    flash("Email already exists. Please use a different email.")
                return redirect(url_for('register_form'))
            
             # ✅ Validate license key
            cursor.execute("""
                SELECT * FROM licenses 
                WHERE license_key = %s AND status = 'not_used'
            """, (license_key,))
            license_record = cursor.fetchone()

            if not license_record:
                flash("Invalid or already used license key.")
                return redirect(url_for('register_form'))
            
            remaining_license_months = license_record['validity_months']
            role = license_record['type']


            # Calculate payment_due_date
            validity_months = license_record['validity_months']
            payment_due_date = (datetime.today() + relativedelta(months=validity_months)).date()
            print(f"[DEBUG] payment_due_date calculated: {payment_due_date}")

            # Insert user if no duplicates are found
            query = "INSERT INTO users (username, phone, organization, email, password, mac_address, license_key, payment_due_date, role, remaining_license_months) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
            cursor.execute(query, (username, phone, organization, email, hashed_password, mac_address, license_key, payment_due_date, role, remaining_license_months))

             # Mark license as used
            cursor.execute("""
                UPDATE licenses SET status = 'used' WHERE license_key = %s
            """, (license_key,))

            connection.commit()
            flash("Registration successful! You can now log in.")
            return redirect(url_for('login_form'))
    except mysql.connector.Error as e:
        print(f"[ERROR] Registration failed: {e}")  # Safe to log internally
        flash("Something went wrong during registration. Please try again.")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
 
    return redirect(url_for('register_form'))
 
# @app.route('/login', methods=['GET'])
# def login_form():
#     return render_template('login.html')
 
# @app.route('/login', methods=['POST'])
# def login_user():
#     username = request.form['username']
#     password = request.form['password']
   
#     try:
#         connection = mysql.connector.connect(**db_config)
#         if connection.is_connected():
#             cursor = connection.cursor(dictionary=True)
#             query = "SELECT * FROM users WHERE username = %s"
#             cursor.execute(query, (username,))
#             user = cursor.fetchone()
           
#             if user and bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8')):  # Adjust index as needed

#                 mac_address = get_mac_address()  # Get MAC at login time

#                 if user.get('mac_address') and user['mac_address'] != mac_address:
#                     flash("Unauthorized device. Login denied.")
#                     return redirect(url_for('login_form'))

#                 # Get payment due date
#                 payment_due_date = user.get('payment_due_date')

#                 if payment_due_date:
#                     today = datetime.today().date()
#                     payment_due_date = datetime.strptime(str(payment_due_date), "%Y-%m-%d").date()
#                     days_left = (payment_due_date - today).days

#                     # If subscription has expired, prevent login
#                     if days_left < 0:
#                         flash("Your subscription has expired. Please subscribe to proceed with your login.")
#                         return redirect(url_for('login_form'))

#                     # If subscription is expiring soon (3 days or less), show warning
#                     elif days_left <= 3:
#                         flash("Your subscription is expiring soon. Please renew to use services seamlessly.")
                        
#                 session['username'] = user['username']
#                 session['role'] = user['role']
#                 session['user_id'] = user['id']  # Store user ID in session for later use
#                 session['admin_email'] = user.get('admin_email', None)  # Fetch admin email, if exists
#                 print(session)

#                 location = str(get_location())
#                 device_type = str(get_device_type())
                
#                 print(f"User Login - Location: {location}, Device: {device_type}")

#                 # Update login status to 'logged_in'
#                 update_status_query = "UPDATE users SET login_status = 'logged_in' WHERE id = %s"
#                 cursor.execute(update_status_query, (user['id'],))
#                 connection.commit()
#                 # Step 3: Insert or update login timestamp in the `device_report` table
                
#                 insert_query = """
#                 INSERT INTO device_report (existing_table_id, recent_login_time, location, device)
#                 VALUES (%s, %s, %s, %s)
#                 ON DUPLICATE KEY UPDATE recent_login_time = VALUES(recent_login_time),
#                 location = VALUES(location),
#                 device = VALUES(device)
#                 """
#                 print("⏳ Attempting to insert into device_report...-------------------------------------------------------------")
#                 cursor.execute(insert_query, (user['id'], datetime.now(), str(location), str(device_type)))
#                 connection.commit()

#                 # ✅ Start a background thread to update hosts file
#                 threading.Thread(target=update_hosts_file, args=(user['id'],), daemon=True).start()
#                 print(f"🟢 Background thread started for User ID: {user['id']}")

#                 # 🚀 **2️⃣ Start a background thread for continuous updates**
#                 threading.Thread(target=update_hosts_continuously, args=(user['id'],), daemon=True).start()
#                 print(f"🔄 Continuous hosts update started for User ID: {user['id']}")

#                 flash("Login successful!")

#                  # Fetch previous data (vulnerabilities_count, threats_count, usb_ports_count, software_count)
#                 fetch_query = """
#                 SELECT 
#                     COALESCE(vulnerabilities_count, 0) AS total_vulnerabilities,
#                     COALESCE(threats_count, 0) AS total_threats,
#                     COALESCE(network_ports_count, 0) AS open_ports_count,
#                     COALESCE(software_count, 0) AS total_software_installed
#                 FROM device_report
#                 WHERE existing_table_id = %s AND cpu_information IS NOT NULL
#                 ORDER BY recent_login_time DESC
#                 LIMIT 1
#                 """
#                 cursor.execute(fetch_query, (user['id'],))
#                 previous_data = cursor.fetchone()
#                 print(previous_data)

#                 # If no data is found, fetch the data from the previous scan (next available data)
#                 if not previous_data:
#                     fetch_fallback_query = """
#                     SELECT 
#                         COALESCE(vulnerabilities_count, 0) AS total_vulnerabilities,
#                         COALESCE(threats_count, 0) AS total_threats,
#                         COALESCE(network_ports_count, 0) AS open_ports_count,
#                         COALESCE(software_count, 0) AS total_software_installed
                        
#                     FROM device_report
#                     WHERE existing_table_id = %s AND cpu_information IS NOT NULL
#                     ORDER BY recent_login_time ASC 
#                     LIMIT 1
#                     """
#                     cursor.execute(fetch_fallback_query, (user['id'],))
#                     previous_data = cursor.fetchone()
#                     print(previous_data)

#                 # Store the fetched data in session
#                 if previous_data:
#                     session['total_vulnerabilities'] = previous_data['total_vulnerabilities']
#                     session['total_threats'] = previous_data['total_threats']
#                     session['open_ports_count'] = previous_data['open_ports_count']
#                     session['total_software_installed'] = previous_data['total_software_installed']
#                 else:
#                     # Handle the case where no data is available
#                     session['total_vulnerabilities'] = 0
#                     session['total_threats'] = 0
#                     session['open_ports_count'] = 0
#                     session['total_software_installed'] = 0

#                 # Add query for the most recent scan time
#                 scan_time_query = """
#                 SELECT dr.last_check_timestamp 
#                 FROM device_report dr
#                 WHERE dr.existing_table_id = %s AND dr.cpu_information IS NOT NULL
#                 ORDER BY dr.last_check_timestamp DESC
#                 LIMIT 1
#                 """
#                 cursor.execute(scan_time_query, (user['id'],))
#                 scan_time_result = cursor.fetchone()    
#                 session['scan_time_result'] = (
#                     scan_time_result['last_check_timestamp'] 
#                     if scan_time_result else "No scan data available."
#                 )
#                 if scan_time_result:
#                     print(f"Previous scan time fetched: {scan_time_result['last_check_timestamp']}")
#                     session['previous_scan_time'] = scan_time_result['last_check_timestamp']
#                 else:
#                     print("No previous scan time found for user.")
#                     session['previous_scan_time'] = None

#                 print(previous_data)

#                 user_id = session.get('user_id')  # Get the user ID from session
#                 if user_id:
#                     start_periodic_check(user_id)

#                 return redirect(url_for('admin_dashboard' if user['role'] == 'admin' else 'welcome'))
#             else:
#                 flash("Invalid username or password.")
#     except Error as e:
#         flash(f"Error occurred: {e}")
#     finally:
#         if connection.is_connected():
#             cursor.close()
#             connection.close()
#     return redirect(url_for('login_form'))

@app.route('/login', methods=['GET'])
def login_form():
    return render_template('login.html')

# @app.route('/login', methods=['POST'])
# def login_user():
#     username = request.form['username']
#     password = request.form['password']
#     otp_input = request.form.get('otp')
#     action = request.form.get('action')

#     try:
#         connection = mysql.connector.connect(**db_config)
#         if connection.is_connected():
#             cursor = connection.cursor(dictionary=True)
#             query = "SELECT * FROM users WHERE username = %s"
#             cursor.execute(query, (username,))
#             user = cursor.fetchone()

#             if user and bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8')):

#                 # ✅ Handle OTP Resend
#                 if action == 'resend':
#                     pre_2fa_user = session.get('pre_2fa_user')
#                     if not pre_2fa_user:
#                         flash("Session expired. Please login again.")
#                         return redirect(url_for('login_form'))

#                     otp = f"{randint(100000, 999999)}"
#                     otp_expiry = datetime.now() + timedelta(minutes=5)
#                     session['otp'] = otp
#                     session['otp_expiry'] = otp_expiry.strftime('%Y-%m-%d %H:%M:%S')
#                     send_email_otp(pre_2fa_user['email'], otp)

#                     flash("A new OTP has been sent to your registered email.")
#                     return render_template('otp_verify.html', username=username, password=password)

#                 # ✅ Check subscription BEFORE generating OTP
#                 payment_due_date = user.get('payment_due_date')
#                 if not payment_due_date:
#                     flash("No subscription info found. Contact administrator.")
#                     return redirect(url_for('login_form'))

#                 try:
#                     if isinstance(payment_due_date, str):
#                         payment_due_date = payment_due_date.strip("'")
#                         payment_due_date = datetime.strptime(payment_due_date, "%Y-%m-%d").date()
#                     elif isinstance(payment_due_date, datetime):
#                         payment_due_date = payment_due_date.date()
#                 except Exception as e:
#                     flash("Invalid subscription date format.")
#                     print("❌ Error parsing payment_due_date:", e)
#                     return redirect(url_for('login_form'))

#                 today = datetime.today().date()
#                 days_left = (payment_due_date - today).days
#                 print(f"📆 Subscription check: today = {today}, due = {payment_due_date}, days_left = {days_left}")

#                 if days_left < 0:
#                     flash("Your subscription has expired. Please subscribe to proceed with your login.")
#                     return redirect(url_for('login_form'))
#                 elif days_left <= 3:
#                     flash("Your subscription is expiring soon. Please renew to use services seamlessly.")

#                 # ✅ OTP not submitted yet → Generate and Send
#                 if not otp_input:
#                     otp = f"{randint(100000, 999999)}"
#                     otp_expiry = datetime.now() + timedelta(minutes=5)

#                     session['otp'] = otp
#                     session['otp_expiry'] = otp_expiry.strftime('%Y-%m-%d %H:%M:%S')
#                     session['pre_2fa_user'] = {
#                         'id': user['id'],
#                         'username': user['username'],
#                         'role': user['role'],
#                         'admin_email': user.get('admin_email', None),
#                         'email': user['email'],
#                         'mac_address': user.get('mac_address'),
#                         'payment_due_date': str(payment_due_date),
#                     }

#                     send_email_otp(user['email'], otp)
#                     flash("An OTP has been sent to your registered email. Please enter it below.")
#                     return render_template('otp_verify.html', username=username, password=password)

#                 # ✅ OTP Submitted: Validate
#                 saved_otp = session.get('otp')
#                 otp_expiry_str = session.get('otp_expiry')
#                 pre_2fa_user = session.get('pre_2fa_user')

#                 if not saved_otp or not otp_expiry_str or not pre_2fa_user:
#                     flash("Session expired or invalid. Please login again.")
#                     return redirect(url_for('login_form'))

#                 otp_expiry = datetime.strptime(otp_expiry_str, '%Y-%m-%d %H:%M:%S')
#                 if datetime.now() > otp_expiry:
#                     flash("OTP expired. Please login again.")
#                     session.clear()
#                     return redirect(url_for('login_form'))

#                 if otp_input != saved_otp:
#                     flash("Invalid OTP. Please try again.")
#                     return render_template('otp_verify.html', username=username, password=password)

#                 # ✅ OTP is valid, complete login
#                 user = pre_2fa_user
#                 mac_address = get_mac_address()
#                 if user.get('mac_address') and user['mac_address'] != mac_address:
#                     flash("Unauthorized device. Login denied.")
#                     return redirect(url_for('login_form'))

#                 session['username'] = user['username']
#                 session['role'] = user['role']
#                 session['user_id'] = user['id']
#                 session['existing_table_id'] = user['id']
#                 session['admin_email'] = user.get('admin_email', None)

#                 location = str(get_location())
#                 # device_type = str(get_device_type())

#                 device_type = json.dumps(get_device_type())

#                 update_status_query = "UPDATE users SET login_status = 'logged_in' WHERE id = %s"
#                 cursor.execute(update_status_query, (user['id'],))
#                 connection.commit()

#                 insert_query = """
#                 INSERT INTO device_report (existing_table_id, recent_login_time, location, device)
#                 VALUES (%s, %s, %s, %s)
#                 ON DUPLICATE KEY UPDATE recent_login_time = VALUES(recent_login_time),
#                 location = VALUES(location),
#                 device = VALUES(device)
#                 """
#                 cursor.execute(insert_query, (user['id'], datetime.now(), location, device_type))
#                 connection.commit()

#                 threading.Thread(target=update_hosts_file, args=(user['id'],), daemon=True).start()
#                 threading.Thread(target=update_hosts_continuously, args=(user['id'],), daemon=True).start()

#                 flash("Login successful!")
#                 session.pop('otp', None)
#                 session.pop('otp_expiry', None)
#                 session.pop('pre_2fa_user', None)

#                 # Fetch latest scan stats
#                 fetch_query = """
#                 SELECT 
#                     COALESCE(vulnerabilities_count, 0) AS total_vulnerabilities,
#                     COALESCE(threats_count, 0) AS total_threats,
#                     COALESCE(network_ports_count, 0) AS open_ports_count,
#                     COALESCE(software_count, 0) AS total_software_installed
#                 FROM device_report
#                 WHERE existing_table_id = %s AND cpu_information IS NOT NULL
#                 ORDER BY recent_login_time DESC
#                 LIMIT 1
#                 """
#                 cursor.execute(fetch_query, (user['id'],))
#                 previous_data = cursor.fetchone() or {}

#                 session['total_vulnerabilities'] = previous_data.get('total_vulnerabilities', 0)
#                 session['total_threats'] = previous_data.get('total_threats', 0)
#                 session['open_ports_count'] = previous_data.get('open_ports_count', 0)
#                 session['total_software_installed'] = previous_data.get('total_software_installed', 0)

#                 scan_time_query = """
#                 SELECT last_check_timestamp FROM device_report
#                 WHERE existing_table_id = %s AND cpu_information IS NOT NULL
#                 ORDER BY last_check_timestamp DESC LIMIT 1
#                 """
#                 cursor.execute(scan_time_query, (user['id'],))
#                 scan_time_result = cursor.fetchone()

#                 session['scan_time_result'] = (
#                     scan_time_result['last_check_timestamp']
#                     if scan_time_result else "No scan data available."
#                 )
#                 session['previous_scan_time'] = (
#                     scan_time_result['last_check_timestamp']
#                     if scan_time_result else None
#                 )

#                 return redirect(url_for('admin_dashboard' if user['role'] == 'admin' else 'welcome'))
#             else:
#                 flash("Invalid username or password.")
#     except Exception as e:
#         flash(f"Error occurred: {e}")
#         print("Exception during login:", e)
#     finally:
#         if connection.is_connected():
#             cursor.close()
#             connection.close()

#     return redirect(url_for('login_form'))

def validate_login_input(username, password):
    if not username or not password:
        return False
    if not re.match(r'^[\w.@+-]+$', username):  # only safe characters
        return False
    if len(password) < 6:
        return False
    return True
 
@app.route('/login', methods=['POST'])  #previous recent logic
def login_user():
    username = request.form['username']
    password = request.form['password']
    license_key_input = request.form.get('license_key', '').strip()
    otp_input = request.form.get('otp')
    action = request.form.get('action')

    if not validate_login_input(username, password):
        flash("Invalid input format. Please check your username or password.")
        return redirect(url_for('login_form'))
   
    try:
        connection = mysql.connector.connect(**db_config)
        if connection.is_connected():
            cursor = connection.cursor(dictionary=True)
            query = "SELECT * FROM users WHERE username = %s"
            cursor.execute(query, (username,))
            user = cursor.fetchone()
           
            if user and bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8')):  # Adjust index as needed


                mac_address = get_mac_address()

                # Check if the MAC address is already bound for this user
                if not user.get('mac_address'):  # If no MAC address is stored for this user
                    # Bind the MAC address on first login
                    cursor.execute("UPDATE users SET mac_address=%s WHERE id=%s", (mac_address, user['id']))
                    connection.commit()
                    print(f"✅ First login: Bound MAC {mac_address} for user {user['username']}")
                else:
                    # If MAC is already bound, just check if the current MAC matches
                    if user['mac_address'] != mac_address:
                        flash("Unauthorized device. Login denied.")
                        return redirect(url_for('login_form'))

                 # ✅ Validate payment_due_date before OTP
                payment_due_date = user.get('payment_due_date')

                if not payment_due_date:
                    flash("No subscription info found. Contact administrator.")
                    return redirect(url_for('login_form'))

                try:
                    if isinstance(payment_due_date, str):
                        payment_due_date = payment_due_date.strip("'")
                        payment_due_date = datetime.strptime(payment_due_date, "%Y-%m-%d").date()
                    elif isinstance(payment_due_date, datetime):
                        payment_due_date = payment_due_date.date()
                except Exception as e:
                    flash("Invalid subscription date format.")
                    print("❌ Error parsing payment_due_date:", e)
                    return redirect(url_for('login_form'))

                today = datetime.today().date()
                days_left = (payment_due_date - today).days
                print(f"📆 Subscription check: today = {today}, due = {payment_due_date}, days_left = {days_left}")

                if days_left < 0:
                    flash("Your License key has expired. Please contact your admin to proceed with your login.")
                    return redirect(url_for('login_form'))
                elif days_left <= 3:
                    flash("Your License Key is expiring soon. Please renew to use services seamlessly.")


                if action == 'resend':
                # Resend OTP
                    pre_2fa_user = session.get('pre_2fa_user')
                    if not pre_2fa_user:
                        flash("Session expired. Please login again.")
                        return redirect(url_for('login_form'))

                    otp = f"{randint(100000, 999999)}"
                    otp_expiry = datetime.now() + timedelta(minutes=5)
                    session['otp'] = otp
                    session['otp_expiry'] = otp_expiry.strftime('%Y-%m-%d %H:%M:%S')

                    send_email_otp(pre_2fa_user['email'], otp)
                    flash("A new OTP has been sent to your registered email.")
                    return render_template('otp_verify.html', username=username, password=password)

                if not otp_input:
                    # Generate OTP and expiry
                    otp = f"{randint(100000, 999999)}"
                    otp_expiry = datetime.now() + timedelta(minutes=5)

                    # Save OTP info & user details in session
                    session['otp'] = otp
                    session['otp_expiry'] = otp_expiry.strftime('%Y-%m-%d %H:%M:%S')
                    session['pre_2fa_user'] = {
                        'id': user['id'],
                        'username': user['username'],
                        'role': user['role'],
                        'admin_email': user.get('admin_email', None),
                        'email': user['email'],
                        'mac_address': user.get('mac_address'),
                        'payment_due_date': user.get('payment_due_date'),
                        # Add other user fields if needed for later use
                        'organization': user['organization'],
                    }

                    # Send OTP to user's email
                    send_email_otp(user['email'], otp)

                    flash("We've sent a one-time password (OTP) to your registered email")
                    # Render OTP entry form (new template)
                    return render_template('otp_verify.html', username=username, password=password)

                else:
                    # OTP submitted: validate it
                    start = time.time()
                    saved_otp = session.get('otp')
                    otp_expiry_str = session.get('otp_expiry')
                    pre_2fa_user = session.get('pre_2fa_user')
                    print(f"OTP fetch and session read took {time.time()-start} seconds")

                    if not saved_otp or not otp_expiry_str or not pre_2fa_user:
                        flash("Session expired or invalid. Please login again.")
                        return redirect(url_for('login_form'))

                    otp_expiry = datetime.strptime(otp_expiry_str, '%Y-%m-%d %H:%M:%S')
                    if datetime.now() > otp_expiry:
                        flash("OTP expired. Please login again.")
                        session.clear()
                        return redirect(url_for('login_form'))

                    if otp_input != saved_otp:
                        flash("Invalid OTP. Please try again.")
                        return render_template('otp_verify.html', username=username, password=password)

                    # OTP is valid — complete login using saved user info
                    user = pre_2fa_user

                    # mac_address = get_mac_address()  # Get MAC at login time

                    # # ✅ NEW: License key validation on first login (MAC not yet bound)
                    # if not user.get('mac_address'):
                    #     # Check if license key exists
                    #     cursor.execute("""
                    #         SELECT license_key, payment_due_date FROM users WHERE id = %s
                    #     """, (user['id'],))
                    #     license_info = cursor.fetchone()

                    #     if not license_info or not license_info.get('license_key'):
                    #         flash("No license key found. Contact administrator.")
                    #         return redirect(url_for('login_form'))

                    #     # Validate license period (reuse payment_due_date from earlier check)
                    #     if days_left < 0:
                    #         flash("License key has expired. Please renew.")
                    #         return redirect(url_for('login_form'))

                    #     # Bind MAC address on first login
                    #     cursor.execute("""
                    #         UPDATE users SET mac_address=%s WHERE id=%s
                    #     """, (mac_address, user['id']))
                    #     connection.commit()
                    #     print(f"✅ First login: Bound MAC {mac_address} for user {user['username']}")


                    # if user.get('mac_address') and user['mac_address'] != mac_address:
                    #     flash("Unauthorized device. Login denied.")
                    #     return redirect(url_for('login_form'))

                            
                    session['username'] = user['username']
                    session['role'] = user['role']
                    session['user_id'] = user['id']  # Store user ID in session for later use
                    session['organization'] = user['organization']
                    session['existing_table_id'] = user['id'] 
                    session['admin_email'] = user.get('admin_email', None)  # Fetch admin email, if exists
                    print(session)

                    location = str(get_location())
                    # device_type = str(get_device_type())
                    device_type = json.dumps(get_device_type())
                    
                    print(f"User Login - Location: {location}, Device: {device_type}")

                    # Update login status to 'logged_in'
                    start_db = time.time()
                    update_status_query = "UPDATE users SET login_status = 'logged_in' WHERE id = %s"
                    cursor.execute(update_status_query, (user['id'],))
                    connection.commit()
                    # Step 3: Insert or update login timestamp in the `device_report` table
                    
                    insert_query = """
                    INSERT INTO device_report (existing_table_id, recent_login_time, location, device)
                    VALUES (%s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE recent_login_time = VALUES(recent_login_time),
                    location = VALUES(location),
                    device = VALUES(device)
                    """
                    print("⏳ Attempting to insert into device_report...-------------------------------------------------------------")
                    cursor.execute(insert_query, (user['id'], datetime.now(), str(location), device_type))
                    connection.commit()

                    # ✅ Start a background thread to update hosts file
                    threading.Thread(target=update_hosts_file, args=(user['id'],), daemon=True).start()
                    print(f"🟢 Background thread started for User ID: {user['id']}")

                    # 🚀 **2️⃣ Start a background thread for continuous updates**
                    threading.Thread(target=update_hosts_continuously, args=(user['id'],), daemon=True).start()
                    print(f"🔄 Continuous hosts update started for User ID: {user['id']}")

                    # Fetch additional values from DB (like remaining_license_months & is_license_admin)
                    cursor.execute("""
                        SELECT is_license_admin, remaining_license_months 
                        FROM users WHERE id=%s
                    """, (user['id'],))
                    license_info = cursor.fetchone()
                    if license_info:
                        session['is_license_admin'] = license_info['is_license_admin']
                        session['remaining_license_months'] = license_info['remaining_license_months']
                    else:
                        session['is_license_admin'] = 0
                        session['remaining_license_months'] = 0

                    flash("Login successful!")
                
                    session.pop('otp', None)
                    session.pop('otp_expiry', None)
                    session.pop('pre_2fa_user', None)

                    # Fetch previous data (vulnerabilities_count, threats_count, usb_ports_count, software_count)
                    fetch_query = """
                    SELECT 
                        COALESCE(vulnerabilities_count, 0) AS total_vulnerabilities,
                        COALESCE(threats_count, 0) AS total_threats,
                        COALESCE(network_ports_count, 0) AS open_ports_count,
                        COALESCE(software_count, 0) AS total_software_installed
                    FROM device_report
                    WHERE existing_table_id = %s AND cpu_information IS NOT NULL
                    ORDER BY recent_login_time DESC
                    LIMIT 1
                    """
                    cursor.execute(fetch_query, (user['id'],))
                    previous_data = cursor.fetchone()
                    print(previous_data)

                    # If no data is found, fetch the data from the previous scan (next available data)
                    if not previous_data:
                        fetch_fallback_query = """
                        SELECT 
                            COALESCE(vulnerabilities_count, 0) AS total_vulnerabilities,
                            COALESCE(threats_count, 0) AS total_threats,
                            COALESCE(network_ports_count, 0) AS open_ports_count,
                            COALESCE(software_count, 0) AS total_software_installed
                            
                        FROM device_report
                        WHERE existing_table_id = %s AND cpu_information IS NOT NULL
                        ORDER BY recent_login_time ASC 
                        LIMIT 1
                        """
                        cursor.execute(fetch_fallback_query, (user['id'],))
                        previous_data = cursor.fetchone()
                        print(previous_data)

                    # Store the fetched data in session
                    if previous_data:
                        session['total_vulnerabilities'] = previous_data['total_vulnerabilities']
                        session['total_threats'] = previous_data['total_threats']
                        session['open_ports_count'] = previous_data['open_ports_count']
                        session['total_software_installed'] = previous_data['total_software_installed']
                    else:
                        # Handle the case where no data is available
                        session['total_vulnerabilities'] = 0
                        session['total_threats'] = 0
                        session['open_ports_count'] = 0
                        session['total_software_installed'] = 0

                    # Add query for the most recent scan time
                    scan_time_query = """
                    SELECT dr.last_check_timestamp 
                    FROM device_report dr
                    WHERE dr.existing_table_id = %s AND dr.cpu_information IS NOT NULL
                    ORDER BY dr.last_check_timestamp DESC
                    LIMIT 1
                    """
                    cursor.execute(scan_time_query, (user['id'],))
                    scan_time_result = cursor.fetchone()    
                    session['scan_time_result'] = (
                        scan_time_result['last_check_timestamp'] 
                        if scan_time_result else "No scan data available."
                    )
                    if scan_time_result:
                        print(f"Previous scan time fetched: {scan_time_result['last_check_timestamp']}")
                        session['previous_scan_time'] = scan_time_result['last_check_timestamp']
                    else:
                        print("No previous scan time found for user.")
                        session['previous_scan_time'] = None

                    print(previous_data)
                    print(f"User fetch took {time.time()-start_db} seconds")

                    user_id = session.get('user_id')  # Get the user ID from session
                    if user_id:
                        start_periodic_check(user_id)

                    return redirect(url_for('admin_dashboard' if user['role'] == 'admin' else 'welcome'))
            else:
                flash("Invalid username or password.")
    except Error as e:
        print(f"[ERROR] Login failed: {e}")  # 🔐 Log internally only
        flash("Something went wrong during login. Please try again.")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
    return redirect(url_for('login_form'))

def get_db_connection():
    connection = mysql.connector.connect(**db_config)
    return connection
 
@app.route('/welcome', methods=['GET'])
@login_required
def welcome():
    connection = get_db_connection()  # Your existing database connection function
    cursor = connection.cursor(dictionary=True)

    def convert_time_to_hours(time_str):
        """Convert 'hh:mm:ss' format to decimal hours."""
        h, m, s = map(int, time_str.split(':'))
        return h + m / 60 + s / 3600

    def format_decimal_hours_to_hms(decimal_hours):
        """Convert decimal hours to HH:MM:SS format."""
        total_seconds = int(decimal_hours * 3600)
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        seconds = total_seconds % 60
        return f"{hours:02}:{minutes:02}:{seconds:02}"

    try:
        user_id = session.get('user_id')  # Assuming user_id is stored in session
        cursor.execute("USE defaultdb;")  # Switch to your database

        scan_time_query = """
            SELECT dr.last_check_timestamp 
            FROM device_report dr
            WHERE dr.existing_table_id = %s AND dr.cpu_information IS NOT NULL
            ORDER BY dr.last_check_timestamp DESC
            LIMIT 1
        """
        cursor.execute(scan_time_query, (user_id,))
        scan_time_result = cursor.fetchone()
        session['scan_time_result'] = (
            scan_time_result['last_check_timestamp'] 
            if scan_time_result else "No scan data available."
        )
        if scan_time_result:
            print(f"Previous scan time fetched: {scan_time_result['last_check_timestamp']}")
            session['previous_scan_time'] = scan_time_result['last_check_timestamp']
        else:
            print("No previous scan time found for user.")
            session['previous_scan_time'] = None

        query = """
            SELECT 
                dr.recently_used_applications, 
                dr.recent_login_time
            FROM device_report dr
            INNER JOIN users u ON dr.existing_table_id = u.id
            WHERE u.id = %s AND recently_used_applications IS NOT NULL
            ORDER BY dr.recent_login_time DESC
            LIMIT 1;
        """
        cursor.execute(query, (user_id,))
        recent_apps_data = cursor.fetchone()

        # Debugging: Print the fetched data
        print("Fetched Recent Apps Data:", recent_apps_data)

        # Parse application data into a format suitable for graphing
        if recent_apps_data and recent_apps_data['recently_used_applications']:
            import json
            try:
                apps_data = json.loads(recent_apps_data['recently_used_applications'])
                print("Parsed Apps Data:", apps_data)  # Debugging print

                # Extract application names and time spent
                app_names = list(apps_data.keys())
                numeric_time_spent = [convert_time_to_hours(data['time_spent']) for data in apps_data.values()]
                formatted_time_spent = [format_decimal_hours_to_hms(hours) for hours in numeric_time_spent]

                # Debugging: Print the graph data
                print("App Names:", app_names)
                print("Numeric Time Spent (Hours):", numeric_time_spent)
                print("Formatted Time Spent (HH:MM:SS):", formatted_time_spent)
            except json.JSONDecodeError as e:
                print("JSON Parsing Error:", e)
                app_names, numeric_time_spent, formatted_time_spent = [], [], []
        else:
            app_names = []
            numeric_time_spent = []
            formatted_time_spent = []

    except Exception as e:
        print("Error:", e)
        app_names, numeric_time_spent, formatted_time_spent = [], [], []

    finally:
        cursor.close()
        connection.close()

    # Render the template with the graph data
    return render_template(
        'welcome.html',
        total_vulnerabilities=session.get('total_vulnerabilities', 0),
        total_threats=session.get('total_threats', 0),
        network_ports_count=session.get('open_ports_count', 0),
        total_software_installed=session.get('total_software_installed', 0),
        previous_scan_time=session.get('previous_scan_time'),
        app_names=app_names,
        numeric_time_spent=numeric_time_spent,
        formatted_time_spent=formatted_time_spent

    )

def get_db_connection():
    conn = mysql.connector.connect(**db_config)
    return conn

@app.route('/admin', methods=['GET', 'POST'])
@admin_required
def admin_dashboard():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # ✅ Get current admin's organization from session
    organization = session.get('organization')
    user_id = session.get('user_id')
    if not organization or not user_id:
        flash("Organization not found in session. Please log in again.")
        return redirect(url_for('login_form'))

    # Fetch users for GET request
    cursor.execute('''
        SELECT id, username, report, login_status, mac_address 
        FROM users 
        WHERE organization = %s
    ''', (organization,))
    users = cursor.fetchall()

    # Fetch current admin's license expiry date
    cursor.execute('SELECT payment_due_date FROM users WHERE id = %s', (user_id,))
    admin_license = cursor.fetchone()
    license_expiry_date = admin_license['payment_due_date'] if admin_license else None

    cursor.close()
    conn.close()

    return render_template('admin.html', users=users, license_expiry_date=license_expiry_date)

@app.route('/update_mac/<int:user_id>', methods=['POST'])
@admin_required
def update_mac(user_id):
    new_mac = request.form.get('mac_address')

    if not new_mac:
        flash("MAC address cannot be empty.")
        return redirect(url_for('admin_dashboard'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

     # ✅ Get current admin's organization from session
    admin_org = session.get('organization')
    if not admin_org:
        flash("Session expired. Please log in again.")
        return redirect(url_for('login_form'))

    # Fetch old MAC address
    cursor.execute("SELECT mac_address, organization  FROM users WHERE id = %s", (user_id,))
    result = cursor.fetchone()

    if not result:
        flash("User not found.")
        return redirect(url_for('admin_dashboard'))

    if result['organization'] != admin_org:
        flash("Unauthorized: You can only update users within your organization.")
        return redirect(url_for('admin_dashboard'))
    
    old_mac = result['mac_address'] if result else None

    # Update the user's MAC address
    cursor.execute("UPDATE users SET mac_address = %s WHERE id = %s", (new_mac, user_id))

    # Insert into log table
    cursor.execute(
        "INSERT INTO mac_update_logs (updated_by_admin_id, user_id, old_mac, new_mac, updated_by_admin_name) VALUES (%s, %s, %s, %s, %s)",
        (session.get('user_id'),  user_id, old_mac, new_mac, session.get('username'))
    )

    conn.commit()
    cursor.close()
    conn.close()

    flash(f"MAC address for User ID {user_id} updated and change logged.")
    return redirect(url_for('admin_dashboard'))


@app.route('/view-report/<int:user_id>')
@admin_required
def view_report(user_id):
    try:
        connection = mysql.connector.connect(**db_config)
        cursor = connection.cursor(dictionary=True)

        # ✅ Get current admin's organization from session
        organization = session.get('organization')
        if not organization:
            flash("Session expired or invalid. Please log in again.")
            return redirect(url_for('login_form'))

        # cursor.execute("SELECT username, report FROM users WHERE id = %s", (user_id,))
        cursor.execute("SELECT username, report, organization FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()

        if not user or not user["report"] or user["organization"] != organization:
            return "No report found for this user.", 403

        report_html = user["report"]

        # Inject download button at the top of the report
        download_button = f"""
        <div style='text-align:right; margin: 20px;'>
            <a href="{url_for('download_pdf', user_id=user_id)}" target="_blank" 
               style="padding: 10px 20px; background-color: #007BFF; color: white; 
                      text-decoration: none; border-radius: 5px;">
                ⬇️ Download as PDF
            </a>
        </div>
        """

        full_html = f"<!DOCTYPE html><html><head><title>View Report</title></head><body>{download_button}{report_html}</body></html>"
        return full_html

    except Exception as e:
        return f"Error retrieving report: {e}", 500
    finally:
        if 'connection' in locals() and connection.is_connected():
            cursor.close()
            connection.close()

#---------------------------------------------------------------audit-----------------------------------------------------------------------
 
# def fetch_audit_data(filter=None):
#     conn = mysql.connector.connect(**db_config)
#     cursor = conn.cursor()
 
#     # Add filter condition if provided
#     filter_condition = ""
#     if filter:
#         filter_condition = f"AND u.username LIKE '%{filter}%'"
 
#     query = f"""
#     WITH LatestLogins AS (
#         SELECT
#             e.existing_table_id,
#             MAX(e.recent_login_time) AS recent_login_time
#         FROM
#             device_report e
#         WHERE
#             e.existing_table_id IN (SELECT id FROM users)
#             AND e.recent_login_time >= CURDATE() - INTERVAL 30 DAY
#         GROUP BY
#             e.existing_table_id, DATE(e.recent_login_time)
#     )
#     SELECT
#         e.existing_table_id,
#         u.username,   -- Join to get the username
#         ll.recent_login_time,
#         e.location,
#         e.device
#     FROM
#         LatestLogins ll
#     JOIN
#         device_report e ON e.existing_table_id = ll.existing_table_id
#         AND e.recent_login_time = ll.recent_login_time
#     JOIN
#         users u ON e.existing_table_id = u.id  -- Join with the users table to get the username
#     WHERE 1=1
#     {filter_condition}
#     ORDER BY
#         ll.recent_login_time DESC;
#     """
 
#     cursor.execute(query)
#     audit_data = cursor.fetchall()
 
#     # Close the connection
#     cursor.close()
#     conn.close()
 
#     return audit_data

def fetch_audit_data(filter=None):
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor()

    organization = session.get('organization')
    if not organization:
        return []

    filter_condition = ""
    values = [organization]  # Organization is always the first filter

    if filter:
        filter_condition = "AND u.username LIKE %s"
        values.append(f"%{filter}%")

    query = f"""
    WITH LatestLogins AS (
        SELECT
            e.existing_table_id,
            MAX(e.recent_login_time) AS recent_login_time
        FROM
            device_report e
        WHERE
            e.existing_table_id IN (SELECT id FROM users)
            AND e.recent_login_time >= CURDATE() - INTERVAL 30 DAY
        GROUP BY
            e.existing_table_id, DATE(e.recent_login_time)
    )
    SELECT
        e.existing_table_id,
        u.username,
        ll.recent_login_time,
        e.location,
        e.device
    FROM
        LatestLogins ll
    JOIN
        device_report e ON e.existing_table_id = ll.existing_table_id
        AND e.recent_login_time = ll.recent_login_time
    JOIN
        users u ON e.existing_table_id = u.id
    WHERE
        u.organization = %s
        {filter_condition}
    ORDER BY
        ll.recent_login_time DESC;
    """

    cursor.execute(query, tuple(values))
    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    # Format result (your original code unchanged)
    formatted_rows = []
    for row in rows:
        user_id, username, login_time, location_json, device_json = row

        try:
            loc = json.loads(location_json.replace("'", '"'))
            location = f"""Public IP Address: {loc.get('public_ip')}
City: {loc.get('city')}
Region: {loc.get('region')}
Country: {loc.get('country')}
Latitude / Longitude: {loc.get('latitude_longitude')}
Internet Provider: {loc.get('isp')}"""
        except Exception:
            location = "N/A"

        try:
            dev = json.loads(device_json.replace("'", '"'))
            device = f"""Operating System: {dev.get('os_name')} {dev.get('os_version')}
Hostname: {dev.get('hostname')}
Processor: {dev.get('processor')}
Architecture: {dev.get('architecture')}
MAC Address: {dev.get('mac_address')}"""
        except Exception:
            device = "N/A"

        formatted_rows.append((user_id, username, login_time, location, device))

    return formatted_rows
 
@app.route('/audit', methods=['GET', 'POST'])
def audit():
    filter_value = request.args.get('filter', '')  # Get filter value from the URL
    audit_data = fetch_audit_data(filter_value)
    # Fetch current admin's license expiry date (same as admin_dashboard)
    organization = session.get('organization')
    user_id = session.get('user_id')
    license_expiry_date = None
    if organization and user_id:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT payment_due_date FROM users WHERE id = %s', (user_id,))
        admin_license = cursor.fetchone()
        license_expiry_date = admin_license['payment_due_date'] if admin_license else None
        cursor.close()
        conn.close()
    return render_template('audit.html', audit_data=audit_data, filter_value=filter_value, license_expiry_date=license_expiry_date)
 
import pandas as pd
import shutil
import os
from pathlib import Path
from datetime import datetime
from flask import send_file
 
 
@app.route('/download_audit_report', methods=['GET'])
@login_required
def download_audit_report():
    # Get the filter value from the URL (if any)

    organization = session.get('organization')
    if not organization:
        flash("Session expired. Please log in again.")
        return redirect(url_for('login_form'))
    
    filter_value = request.args.get('filter', '')  
    if filter_value == '':
        filter_value = None  # Treat empty filter as no filter
 
    # Fetch the filtered audit data based on the filter
    audit_data = fetch_audit_data(filter_value)
 
    # Convert to DataFrame
    df = pd.DataFrame(audit_data, columns=["ID", "Username", "Recent Login Time", "Location", "Device"])
 
    # Ensure the 'Recent Login Time' column is in datetime format
    df["Recent Login Time"] = pd.to_datetime(df["Recent Login Time"], errors='coerce')
 
    # Save to Excel with proper date formatting
    with pd.ExcelWriter('audit_report.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Audit Report")
        workbook = writer.book
        worksheet = writer.sheets['Audit Report']
 
        # Adjust column width for better visibility
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
 
        # Format the 'Recent Login Time' column
        for row in worksheet.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                if isinstance(cell.value, datetime):
                    cell.number_format = 'YYYY-MM-DD HH:MM:SS'
 
    # Get the user's Downloads folder dynamically (cross-platform)
    if os.name == 'posix':  # macOS/Linux
        downloads_folder = str(Path.home() / "Downloads")
    elif os.name == 'nt':  # Windows
        downloads_folder = str(Path(os.environ["USERPROFILE"]) / "Downloads")
    else:
        downloads_folder = "/tmp"  # Fall back to /tmp for non-supported OS
 
    # Set the full file path for saving
    file_path = os.path.join(downloads_folder, "audit_report.xlsx")
 
    # Save the file and return it
    shutil.move('audit_report.xlsx', file_path)
 
    # Send the file to the user as an attachment
    return send_file(file_path, as_attachment=True)
 
#--------------------------------------------------------------Blocked audit--------------------------------------------------------

 
@app.route('/blocked_audit')
@login_required
def blocked_audit():
    """
    Route to display the Blocked Audit page.
    Fetches data from the website_audit_log table where timestamp is within the past 7 days.
    Sends the data to the HTML template for rendering.
    """
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)  # Fetch results as dictionaries

        organization = session.get('organization')
        user_id = session.get('user_id')
        if not organization:
            flash("Session expired. Please log in again.")
            return redirect(url_for('login_form'))

        # Fetch current admin's license expiry date (same as admin_dashboard)
        cursor.execute('SELECT payment_due_date FROM users WHERE id = %s', (user_id,))
        admin_license = cursor.fetchone()
        license_expiry_date = admin_license['payment_due_date'] if admin_license else None

        seven_days_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')

        # Query to fetch data from the last 7 days
        query = """
            SELECT user_id, username, block_url, action, timestamp, location
            FROM website_audit_log
            JOIN users u ON user_id = u.id
            WHERE timestamp >= %s AND u.organization = %s
            ORDER BY timestamp DESC
        """
        cursor.execute(query, (seven_days_ago, organization))
        blocked_audit_data = cursor.fetchall()  # Returns a list of dictionaries

    except Exception as e:
        print("Database Error:", str(e))
        blocked_audit_data = []
        license_expiry_date = None

    finally:
        cursor.close()
        conn.close()

    return render_template('block_audit.html', blocked_audit_data=blocked_audit_data, license_expiry_date=license_expiry_date)
 
from flask import Response
@app.route('/download_blocked_audit_report')
@login_required
def download_blocked_audit_report():
    """
    Generates an Excel report for blocked audit logs from the past 7 days.
    """
    try:
        organization = session.get('organization')
        if not organization:
            flash("Session expired. Please log in again.")
            return redirect(url_for('login_form'))
        
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
 
        seven_days_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
 
        # Fetch data from the database
        query = """
            SELECT user_id, username, block_url, action, timestamp, location
            FROM website_audit_log
            JOIN users u ON user_id = u.id
            WHERE timestamp >= %s AND u.organization = %s
            ORDER BY timestamp DESC
        """
        cursor.execute(query, (seven_days_ago, organization))
        blocked_audit_data = cursor.fetchall()
 
        cursor.close()
        conn.close()
 
        if not blocked_audit_data:
            return "No data available for download", 404
 
        # Convert data into a pandas DataFrame
        df = pd.DataFrame(blocked_audit_data)
 
        # Fix Timestamp Formatting
        if 'timestamp' in df.columns:
            df['timestamp'] = pd.to_datetime(df['timestamp']).dt.strftime('%Y-%m-%d %H:%M:%S')
 
        # Create an in-memory Excel file
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name="Blocked Audit Logs")
 
            # Auto-adjust column width
            workbook = writer.book
            worksheet = writer.sheets["Blocked Audit Logs"]
 
            for i, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).apply(len).max(), len(col)) + 2
                worksheet.set_column(i, i, max_len)
 
        output.seek(0)
 
        # Send the Excel file as a response
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Blocked_Audit_Report.xlsx"}
        )
 
    except Exception as e:
        print("Error generating Excel report:", str(e))
        return "Error generating the report", 500
#-----------------------------------------------------------------------------------------------------------------

# def convert_html_to_pdf(html_content):
#     """Converts HTML content to a PDF using xhtml2pdf (pisa)."""

#     # Extract section headings and create TOC links
#     section_links = []
#     headings = re.findall(r'<h2 class="toggle-header">(.*?)</h2>', html_content)

#     for idx, heading in enumerate(headings):
#         section_id = f"section-{idx}"
#         section_links.append(
#             f'<li><a href="#{section_id}" style="text-decoration: none; color: #3498db;">{heading}</a></li>'
#         )
#         html_content = html_content.replace(
#             f'<h2 class="toggle-header">{heading}</h2>',
#             f'<h2 id="{section_id}" class="toggle-header">{heading}</h2>'
#         )

#     # Combine all HTML elements into a single string with inline CSS
#     final_html = f"""
#     <!DOCTYPE html>
#     <html>
#     <head>
#         <style>
#             @page {{
#                 size: A4;
#                 margin: 2cm;
#                 @frame footer {{
#                     -pdf-frame-content: footer;
#                     bottom: 1cm;
#                     margin-left: 1cm;
#                     margin-right: 1cm;
#                     height: 1cm;
#                 }}
#             }}
#             body {{
#                 font-family: Arial, sans-serif;
#                 font-size: 11px;
#                 line-height: 1.4;
#                 text-align: left;
#                 background-color: #ffffff;
#                 padding: 10px;
#             }}
#             h1 {{
#                 text-align: center;
#                 font-size: 16px;
#                 font-weight: bold;
#                 color: #2c3e50;
#             }}
#             h2 {{
#                 font-size: 14px;
#                 font-weight: bold;
#                 background-color: #3498db;
#                 color: #ffffff;
#                 padding: 6px;
#                 border-radius: 3px;
#                 margin-top: 10px;
#                 margin-bottom: 6px;
#             }}
#             ul {{
#                 list-style-type: none;
#                 padding-left: 0;
#                 font-size: 12px;
#             }}
#             li {{
#                 font-size: 11px;
#                 padding: 4px;
#             }}
#             table {{
#                 width: 100%;
#                 border-collapse: collapse;
#                 margin-bottom: 10px;
#                 background-color: #ffffff;
#             }}
#             th, td {{
#                 border: 1px solid #ddd;
#                 padding: 5px;
#                 text-align: left;
#                 font-size: 10px;
#                 word-wrap: break-word;
#                 white-space: pre-wrap;
#                 max-width: 100%;
#             }}
#             th {{
#                 background-color: #2c3e50;
#                 color: #ffffff;
#                 font-weight: bold;
#             }}
#             tr:nth-child(even) {{
#                 background-color: #f9f9f9;
#             }}
#             tr:hover {{
#                 background-color: #d5dbdb;
#             }}
#             .toc-heading {{
#                 text-align: center;
#                 background-color: #2c3e50;
#                 color: white;
#                 padding: 5px;
#                 border-radius: 5px;
#             }}
#             .header {{
#                 text-align: center;
#                 width: 100%;
#                 margin-bottom: 10px;
#             }}
#         </style>
#     </head>
#     <body>
#         <div class="header">
#             <img src="static/logo.png" width="100px" />
#             <h1>Device Report</h1>
#         </div>
        
#         <h2 class="toc-heading">Table of Contents</h2>
#         <ul>
#             {''.join(section_links)}
#         </ul>
        
#         <div style="page-break-after: always;"></div>
        
#         {html_content}
        
#         <div id="footer">
#             <pdf:pagenumber /> of <pdf:pagecount />
#         </div>
#     </body>
#     </html>
#     """

#     # Convert HTML to PDF
#     pdf_stream = BytesIO()
#     pisa_status = pisa.CreatePDF(final_html, dest=pdf_stream)

#     pdf_stream.seek(0)  # Reset stream position

#     if pisa_status.err:
#         return None  # Return None if PDF conversion fails

#     return pdf_stream
#---------------------------------------------------------------------------------------------------------------------------------
def get_latest_timestamp(user_id):
    try:
        # Establish a connection
        connection = mysql.connector.connect(**db_config)
        cursor = connection.cursor()

        # Define the SQL query
        query = """
            SELECT last_check_timestamp
            FROM device_report
            WHERE cpu_information IS NOT NULL 
            AND existing_table_id = %s
            ORDER BY last_check_timestamp DESC
            LIMIT 1;
        """

        # Debugging: Print query execution details
        print(f"Executing query: {query} with selected_user_id = {user_id}")

        # Execute the query with the selected user ID
        cursor.execute(query, (user_id,))
        
        # Fetch result
        result = cursor.fetchone()
        print(f"Query result: {result}")  # Debugging output

        # Close connections
        cursor.close()
        connection.close()
        
        return result[0] if result else "N/A"  # Return timestamp if found, otherwise "N/A"

    except mysql.connector.Error as e:
        print(f"Database error: {e}")
        return "N/A"

def fetch_recent_7_scans_download(user_id):
    if not user_id:
        print("❌ No user_id provided.")
        return []

    try:
        print(f"🟡 Attempting to fetch data for user_id: {user_id}")

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)

        # Debug: Print database being used
        cursor.execute("SELECT DATABASE();")
        current_db = cursor.fetchone()
        print(f"🔍 Connected to database: {current_db}")

        # Debug: Check if the table exists
        cursor.execute("SHOW TABLES LIKE 'device_report';")
        table_exists = cursor.fetchone()
        print(f"📂 device_report table exists: {table_exists is not None}")

        # Debug: Print column names
        cursor.execute("DESCRIBE device_report;")
        columns = cursor.fetchall()
        print("📝 Table Columns:", [col['Field'] for col in columns])

        # Debug: Check if data exists for the user_id
        check_query = "SELECT COUNT(*) FROM device_report WHERE existing_table_id = %s;"
        cursor.execute(check_query, (user_id,))
        count_result = cursor.fetchone()
        print(f"🟢 Total records found for user_id {user_id}: {count_result['COUNT(*)']}")

        query = """
        SELECT last_check_timestamp,
               software_count,
               threats_count,
               vulnerabilities_count,
               usb_ports_count
        FROM device_report
        WHERE existing_table_id = %s
          AND cpu_information IS NOT NULL
        ORDER BY last_check_timestamp DESC
        LIMIT 7;
        """

        print(f"🟠 Executing query:\n{query}")
        print(f"🔵 With parameters: {user_id}")

        cursor.execute(query, (user_id,))
        result = cursor.fetchall()

        print(f"✅ Query result: {result}")

        cursor.close()
        conn.close()

        return result

    except Exception as e:
        print(f"❌ Database Error: {e}")
        return []
    

def generate_graphs_download(user_id):
    data = fetch_recent_7_scans_download(user_id)
    
    if not data or len(data) == 0:
        print(f"No data found for recent 9 scans for user {user_id}. Data: {data}")
        return {}

    df = pd.DataFrame(data)
    print("DataFrame created successfully:", df)

    graphs = {}
    metrics = ["software_count", "threats_count", "vulnerabilities_count", "usb_ports_count"]
    
    df["last_check_timestamp"] = df["last_check_timestamp"].astype(str)

    for metric in metrics:
        fig, ax = plt.subplots(figsize=(7, 4))
        ax.bar(df["last_check_timestamp"], df[metric], color="royalblue", width=0.5)
        ax.set_xlabel("Last Check Timestamp")
        ax.set_ylabel(metric.replace("_", " ").title())
        ax.set_title(f"{metric.replace('_', ' ').title()} Over Recent 7 Scans")
        ax.tick_params(axis='x', rotation=45)
        ax.grid(axis="y", linestyle="--", alpha=0.7)

        plt.tight_layout()

        buffer = BytesIO()
        fig.savefig(buffer, format="png", bbox_inches="tight")
        buffer.seek(0)
        graphs[metric] = base64.b64encode(buffer.getvalue()).decode("utf-8")

        if not graphs[metric]:
            print(f"Graph generation failed for metric: {metric}")

        plt.close(fig)

    if not graphs:
        print("No graphs were generated. Check data and plotting logic.")

    return graphs


def convert_html_to_pdfs(html_content, user_id):
    latest_timestamp = get_latest_timestamp(user_id)
   
    cover_html = f"""<br><br><br>
    <div style="text-align: center; width: 100%; height: 100vh; display: flex; flex-direction: column; justify-content: center; align-items: center;">
        <h1 style="font-size: 80px; font-weight: bold; margin-bottom: 5px;">Device Report</h1>
        <h4 style="font-size: 25px; font-weight: normal; margin-bottom: 8px;">Generated on {latest_timestamp} </h4>
        <h3 style="font-size: 42px; font-weight: normal; margin-bottom: 13px;">For Vardaan Global</h3>
        <br>
        <br>
        <div style="position: absolute; bottom: 50px; width: 100%; text-align: center;">
            <p style="font-size: 24px; font-weight: bold;">Powered By:</p>
            <img src="static/logo.png" style="width: 100%; height: auto;" />
        </div>
    </div>
    <pdf:nextpage />
    """
 
    header_html = """
    <div style="text-align: center; width: 100%; margin-bottom: 10px;">
        <h1>Device Report</h1>
    </div>
    """
 
    # Extract section headings and create links for TOC
    section_links = []
    headings = re.findall(r'<h2 class="toggle-header">(.*?)</h2>', html_content)

    for idx, heading in enumerate(headings):
        section_id = f"section-{idx}"  # Unique section ID
        section_links.append(
            f'<li style="font-size: 12px; margin-bottom: 10px;">'  # Increased font size
            f'<pdf:link destination="{section_id}" style="text-decoration: none; color: #3498db; font-size: 20px;">'
            f'{heading}</pdf:link></li>'
        )
        html_content = html_content.replace(
            f'<h2 class="toggle-header">{heading}</h2>',
            f'<pdf:bookmark name="{heading}" /><pdf:destination name="{section_id}" /><h2 id="{section_id}">{heading}</h2>'
        )
 
    # Add Table of Contents (TOC) with clickable links and page numbers
    toc_html = f"""
    <h2 style="text-align: center; background-color: #2c3e50; color: white; padding: 5px; border-radius: 5px;">Table of Contents</h2>
    <ul style="padding: 15px; font-size: 12px;">
        {''.join(section_links)}
    </ul>
    <pdf:nextpage />
    """
 
    # Generate graphs and insert into PDF (All 4 Graphs on One Page)
    # Ensure graphs is generated before use
    graphss = generate_graphs_download(user_id)  # Ensure graphs are generated

    graphs_html = '<h2 style="text-align: center;">Recent 7 Scans Report</h2>'

    # Iterate through graphs and print 2 per page with increased spacing
    for idx, (metric, base64_img) in enumerate(graphss.items()):
        if idx % 2 == 0:  # Start a new page every 2 graphs
            graphs_html += '<div style="display: flex; flex-direction: column; align-items: center; gap: 90px; padding-top: 25px;">'

        graphs_html += f"""
            <div style="width: 90%; text-align: center;">
                <img src="data:image/png;base64,{base64_img}" style="width:600px; height:400px;" />
            </div>
        """

        if idx % 2 == 1 or idx == len(graphss) - 1:  # Close the container and add a page break after 2 graphs
            graphs_html += "</div><pdf:nextpage />"

 
    # Define a persistent footer for page numbers
    footer_html = """
    <pdf:staticcontent name="footer_content">
        <div style="width: 100%; text-align: center; font-size: 10px; color: #555;">
            
        </div>
    </pdf:staticcontent>
    """
 
    # Footer style
    footer_style = """
    <style>
        @page {
            size: A4;
            margin: 20mm;
            @frame footer_frame { -pdf-frame-content: footer_content; bottom: 10mm; height: 12mm; }
        }
    </style>
    """
 
    # CSS Styles
    enhanced_css = """
    <style>
        body {
            font-family: Arial, sans-serif;
            font-size: 11px;
            line-height: 1.4;
            text-align: left;
            background-color: #ffffff;
            padding: 10px;
        }
 
        h1 {
            text-align: center;
            font-size: 16px;
            font-weight: bold;
            color: #2c3e50;
        }
 
        h2 {
            font-size: 14px;
            font-weight: bold;
            background-color: #3498db;
            color: #ffffff;
            padding: 6px;
            border-radius: 3px;
            margin-top: 10px;
            margin-bottom: 6px;
        }
 
        h3 {
            font-size: 13px;
            font-weight: bold;
            color: #2c3e50;
            margin-bottom: 5px;
        }
 
        ul {
            list-style-type: none;
            padding-left: 0;
            font-size: 12px;
        }
 
        li {
            font-size: 11px;
            padding: 4px;
        }
 
        img {
            border: 1px solid #ddd;
            border-radius: 5px;
            padding: 5px;
            background: #f9f9f9;
        }
    </style>
    """
 
    # Final HTML for PDF conversion
    final_html = f"{footer_style}{enhanced_css}{cover_html}{header_html}{toc_html}{graphs_html}{html_content}{footer_html}"
 
    try:
        # Convert HTML to PDF
        pdf_stream = BytesIO()
        pisa_status = pisa.CreatePDF(final_html, dest=pdf_stream)
 
        if pisa_status.err:
            raise Exception("Error generating PDF")
 
        pdf_content = pdf_stream.getvalue()
        pdf_stream.close()
        return pdf_content
 
    except Exception as e:
        print(f"PDF Generation Failed: {e}")
        return None

#---------------------------------------------------------------------------------------------------------------------------------


@app.route('/download_report/<int:user_id>')
@login_required
def download_pdf(user_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        admin_org = session.get('organization')
        if not admin_org:
            flash("Session expired. Please log in again.")
            return redirect(url_for('login_form'))

        print(f"🟡 Fetching report for user_id: {user_id}")
        cursor.execute("SELECT username, report, organization FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()

        if not user or not user['report']:
            print(f"❌ No report found for user_id: {user_id}")
            return "Report not found for the selected user.", 404
        
        if user['organization'] != admin_org:
            print(f"⛔ Unauthorized: User {user_id} is from a different organization.")
            return "Access denied. You are not authorized to download this report.", 403
        
        username = user['username']
        html_content = user['report']
        print(f"✅ Report retrieved for user_id: {user_id}, generating graphs...")

        # Generate Graphs for the User
        graphss = generate_graphs_download(user_id)

        # Embed graphs into HTML content
        for metric, graph_data in graphss.items():
            print(f"🖼 Embedding graph for: {metric}")
            img_tag = f'<img src="data:image/png;base64,{graph_data}" width="600px">'
            html_content = html_content.replace(f"{{{{ {metric}_graph }}}}", img_tag)

        # Convert the modified HTML content to PDF
        pdf_content = convert_html_to_pdfs(html_content, user_id)
        pdf_with_page_numberss = add_page_numbers(pdf_content)

        if pdf_content:
            print(f"📄 PDF successfully generated for user_id: {user_id}")
            return send_file(BytesIO(pdf_with_page_numberss), as_attachment=True, 
                             download_name=f'{username}_{user_id}_report.pdf', mimetype='application/pdf')

        print(f"❌ Failed to generate PDF for user_id: {user_id}")
        return "Failed to generate PDF.", 500

    except Exception as e:
        print(f"❌ Error in download_pdf: {e}")
        return "Internal Server Error", 500

    finally:
        cursor.close()
        conn.close()


# @app.route('/download-excel', methods=['POST'])
# def download_excel():
#     # Get the selected user IDs from the form
#     user_ids = request.form.getlist('user_ids')  # List of selected user IDs

#     # Connect to the database and fetch user data
#     conn = get_db_connection()
#     cursor = conn.cursor(dictionary=True)
#     cursor.execute(f"SELECT id, username, report FROM users WHERE id IN ({','.join(user_ids)})")
#     users = cursor.fetchall()
#     cursor.close()
#     conn.close()

#     if not users:
#         return "No reports found for the selected users.", 404

#     # Create a memory buffer for the zip file
#     zip_buffer = BytesIO()

#     # Create a zip file in memory
#     with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
#         for user in users:
#             user_id = user['id']
#             username = user['username']  # Get the username
#             html_content = user['report']  # The HTML content of the report

#             # Convert the HTML content to PDF (as bytes)
#             pdf_bytes = convert_html_to_pdfs(html_content, user_id)
#             pdf_with_page_numberss_s = add_page_numbers(pdf_bytes)

#             if pdf_bytes:
#                 pdf_filename = f"{username}_report_{user_id}.pdf"  # Use username and user_id for the filename
#                 zip_file.writestr(pdf_filename, pdf_with_page_numberss_s)  # Write the PDF as bytes to the zip file

#     zip_buffer.seek(0)  # Reset buffer position before sending the file

#     return send_file(zip_buffer, as_attachment=True, download_name='reports.zip')

@app.route('/download-excel', methods=['POST'])
@login_required
def download_excel():
    user_ids = request.form.getlist('user_ids')  # List of selected user IDs

    # ✅ Prevent query crash if no user selected
    if not user_ids:
        return "No users selected.", 400

    # ✅ Get current organization from session
    organization = session.get('organization')
    if not organization:
        flash("Session expired. Please log in again.")
        return redirect(url_for('login_form'))

    placeholders = ','.join(['%s'] * len(user_ids))
    query = f"""
        SELECT id, username, report 
        FROM users 
        WHERE id IN ({placeholders}) 
        AND organization = %s
    """
    values = tuple(user_ids) + (organization,)

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute(query, values)
        users = cursor.fetchall()
    except mysql.connector.Error as e:
        print(f"❌ SQL Error: {e}")
        cursor.close()
        conn.close()
        return "Invalid query. Please select valid users.", 500

    cursor.close()
    conn.close()

    if not users:
        return "None of the selected users have reports available.", 404

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for user in users:
            html_content = user['report']
            if not html_content:
                continue  # ✅ Skip empty reports

            user_id = user['id']
            username = user['username']

            pdf_bytes = convert_html_to_pdfs(html_content, user_id)
            pdf_with_page_numberss_s = add_page_numbers(pdf_bytes)

            if pdf_bytes:
                filename = f"{username}_report_{user_id}.pdf"
                zip_file.writestr(filename, pdf_with_page_numberss_s)

        # ✅ Handle case where all users were skipped
        if len(zip_file.filelist) == 0:
            return "None of the selected users have reports available.", 404

    zip_buffer.seek(0)
    return send_file(zip_buffer, as_attachment=True, download_name='reports.zip')


@app.route('/get-user-insights', methods=['POST'])
def get_user_insights():
    try:
        if not request.is_json:
            print("❌ Request is not JSON")
            return jsonify({"error": "Invalid JSON request"}), 400

        data = request.get_json()
        user_id = data.get("user_id")

        if not user_id:
            print("❌ User ID is missing")
            return jsonify({"error": "User ID is required"}), 400

        print(f"✅ Fetching data for User ID: {user_id}")

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        query = """
        SELECT last_check_timestamp, detected_threats, vulnerabilities_count, threats_count, 
               recent_login_time, total_installed_software, software_count, usb_device_id
        FROM device_report
        WHERE existing_table_id = %s
        ORDER BY last_check_timestamp DESC
        LIMIT 10;
        """
        
        cursor.execute(query, (user_id,))
        records = cursor.fetchall()
        conn.close()

        if not records:
            return jsonify({"error": "No data found for this user"}), 404

        # Format timestamps and extract USB status counts
        for record in records:
            # ✅ **Fix for timestamp formatting**
            for field in ["last_check_timestamp", "recent_login_time"]:
                if record[field]:
                    try:
                        if isinstance(record[field], datetime):  
                            dt_object = record[field]  # Already a datetime object
                        else:
                            dt_object = datetime.strptime(record[field], "%Y-%m-%d %H:%M:%S")  # Convert to datetime

                        record[field] = dt_object.strftime("%m/%d/%Y, %I:%M:%S %p")  # Format correctly (AM/PM)
                    except (ValueError, TypeError):
                        record[field] = "Invalid Date"  # Handle errors gracefully

            # ✅ **Extract USB device count (opened and closed)**
            usb_data = record.get("usb_device_id", "[]")
            try:
                usb_list = json.loads(usb_data) if usb_data else []
                record["usb_device_open"] = sum(1 for device in usb_list if device.get("status") == "opened")
                record["usb_device_closed"] = sum(1 for device in usb_list if device.get("status") == "closed")
            except json.JSONDecodeError:
                record["usb_device_open"] = 0
                record["usb_device_closed"] = 0  # Default if JSON is invalid

        print(f"✅ Data Sent: {records}")  # Debugging output
        return jsonify(records)

    except Exception as e:
        print(f"❌ ERROR: {str(e)}")  # Debugging output
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500


@app.route('/graphs-dashboard')
def graphs_dashboard():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    cursor.execute("SELECT id, username FROM users")
    users = cursor.fetchall()
    conn.close()
    
    return render_template("admin.html", users=users)

@app.route('/download-reports-excel', methods=['POST'])
def download_reports_excel():
    
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    organization = session.get('organization')
    if not organization:
        flash("Session expired or invalid. Please log in again.")
        return redirect(url_for('login_form'))

    # Fetch the latest report for each user
    query = """
    SELECT 
        dr.id,
        u.username,
        dr.recent_login_time,
        dr.last_check_timestamp,
        dr.cpu_information,
        dr.memory_information,
        dr.disk_information,
        dr.network_information,
        dr.vpn_detections,
        dr.usb_activity_history,
        dr.vpn_usage_history,
        dr.recently_used_applications,
        dr.unused_applications,
        dr.vulnerability_detections,
        dr.detected_threats,
        dr.total_installed_software,
        dr.usb_ports_count,
        dr.vulnerabilities_count,
        dr.threats_count,
        dr.software_count,
        dr.network_ports_count,
        dr.usb_device_id,
        dr.usb_status
    FROM device_report dr
    INNER JOIN users u ON dr.existing_table_id = u.id
    WHERE u.organization = %s AND dr.id = (
        SELECT id FROM device_report 
        WHERE existing_table_id = dr.existing_table_id 
        ORDER BY last_check_timestamp DESC 
        LIMIT 1
    )
    ORDER BY u.username;
    """

    cursor.execute(query,(organization,))
    data = cursor.fetchall()
    conn.close()

    if not data:
        return "No valid data found.", 400

    # Creating a Pandas DataFrame
    df = pd.DataFrame(data)

    # Define column names
    column_mapping = {
        "id": "Report ID",
        "username": "Username",
        "recent_login_time": "Recent Login Time",
        "last_check_timestamp": "Last Check Timestamp",
        "cpu_information": "CPU Information",
        "memory_information": "Memory Information",
        "disk_information": "Disk Information",
        "network_information": "Network Information",
        "vpn_detections": "VPN Detections",
        "usb_activity_history": "USB Activity History",
        "vpn_usage_history": "VPN Usage History",
        "recently_used_applications": "Recently Used Applications",
        "unused_applications": "Unused Applications",
        "vulnerability_detections": "Vulnerability Detections",
        "detected_threats": "Detected Threats",
        "total_installed_software": "Total Installed Software",
        "usb_ports_count": "USB Ports Count",
        "vulnerabilities_count": "Vulnerabilities Count",
        "threats_count": "Threats Count",
        "software_count": "Software Count",
        "network_ports_count": "Network Ports Count",
        "usb_device_id": "USB Device ID",
        "usb_status": "USB Status"
    }
    
    df.rename(columns=column_mapping, inplace=True)

    # Convert DataFrame to Excel with styling
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Latest Reports", index=False)
        workbook = writer.book
        sheet = writer.sheets["Latest Reports"]

        # Define fill colors
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light Red
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Light Green

        # Get column indexes for vulnerabilities and threats
        # vulnerability_col_idx = None
        # threats_col_idx = None

        # for col_idx, cell in enumerate(sheet[1]):  # First row contains headers
        #     if cell.value == "Vulnerabilities Count":
        #         vulnerability_col_idx = col_idx
        #     if cell.value == "Threats Count":
        #         threats_col_idx = col_idx

        # if vulnerability_col_idx is not None and threats_col_idx is not None:
        #     for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        #         vulnerabilities_cell = row[vulnerability_col_idx]
        #         threats_cell = row[threats_col_idx]

        #         # Ensure the value is converted to an integer safely
        #         def safe_int(value):
        #             if isinstance(value, (int, float)):  # If value is int or float, convert directly
        #                 return int(value)
        #             elif isinstance(value, str) and value.strip().isdigit():  # If value is a digit string, convert
        #                 return int(value.strip())
        #             return 0  # Default to 0 if value is None or invalid

        #         vulnerabilities_value = safe_int(vulnerabilities_cell.value)
        #         threats_value = safe_int(threats_cell.value)

        #         if vulnerabilities_value > 0:
        #             vulnerabilities_cell.fill = red_fill
        #         else:
        #             vulnerabilities_cell.fill = green_fill

        #         if threats_value > 0:
        #             threats_cell.fill = red_fill
        #         else:
        #             threats_cell.fill = green_fill


                # Get column indexes for additional fields
        # Initialize column indexes
        vuln_detect_col_idx = None
        threats_detect_col_idx = None
        vuln_count_col_idx = None
        threats_count_col_idx = None

        # Get column indexes from the first row (headers)
        for col_idx, cell in enumerate(sheet[1]):
            if cell.value == "Vulnerability Detections":
                vuln_detect_col_idx = col_idx
            if cell.value == "Detected Threats":
                threats_detect_col_idx = col_idx
            if cell.value == "Vulnerabilities Count":
                vuln_count_col_idx = col_idx
            if cell.value == "Threats Count":
                threats_count_col_idx = col_idx

        # Apply inline logic row by row
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            # Handle "Vulnerability Detections"
            if vuln_detect_col_idx is not None:
                vuln_cell = row[vuln_detect_col_idx]
                vuln_val = str(vuln_cell.value or "").strip().lower()
                try:
                    parsed = json.loads(vuln_val)
                    if parsed == {} or parsed == [] or parsed == "":
                        vuln_cell.fill = green_fill
                    else:
                        vuln_cell.fill = red_fill
                except:
                    if "no" in vuln_val or vuln_val == "":
                        vuln_cell.fill = green_fill
                    else:
                        vuln_cell.fill = red_fill

            # Handle "Detected Threats"
            if threats_detect_col_idx is not None:
                threat_cell = row[threats_detect_col_idx]
                threat_val = str(threat_cell.value or "").strip().lower()
                try:
                    parsed = json.loads(threat_val)
                    if parsed == {} or parsed == [] or parsed == "":
                        threat_cell.fill = green_fill
                    else:
                        threat_cell.fill = red_fill
                except:
                    if "no" in threat_val or threat_val == "":
                        threat_cell.fill = green_fill
                    else:
                        threat_cell.fill = red_fill

            # Handle "Vulnerabilities Count"
            if vuln_count_col_idx is not None:
                vuln_count_cell = row[vuln_count_col_idx]
                vuln_count_val = vuln_count_cell.value
                if isinstance(vuln_count_val, (int, float)):
                    count = int(vuln_count_val)
                elif isinstance(vuln_count_val, str) and vuln_count_val.strip().isdigit():
                    count = int(vuln_count_val.strip())
                else:
                    count = 0
                vuln_count_cell.fill = red_fill if count > 0 else green_fill

            # Handle "Threats Count"
            if threats_count_col_idx is not None:
                threats_count_cell = row[threats_count_col_idx]
                threats_count_val = threats_count_cell.value
                if isinstance(threats_count_val, (int, float)):
                    count = int(threats_count_val)
                elif isinstance(threats_count_val, str) and threats_count_val.strip().isdigit():
                    count = int(threats_count_val.strip())
                else:
                    count = 0
                threats_count_cell.fill = red_fill if count > 0 else green_fill


    output.seek(0)

    # Return the Excel file as an attachment
    return send_file(
        output, 
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="latest_device_reports.xlsx"
    )

@app.route('/set_usb_expiry/<int:user_id>', methods=['POST'])
def set_usb_expiry(user_id):
    expiry_date = request.form.get('expiry_date')  # Get the selected date

    if not expiry_date:
        return jsonify({"error": "Expiry date is required"}), 400

    try:
        expiry_date_obj = datetime.strptime(expiry_date, "%Y-%m-%d").date()  # Convert to date
        today = datetime.today().date()

        # Ensure expiry date is within the allowed range (1-10 days from today)
        if not (today <= expiry_date_obj <= today + timedelta(days=10)):
            return jsonify({"error": "Expiry date must be within the next 10 days"}), 400

        # Update the database with the new expiry date
        connection = get_db_connection()
        cursor = connection.cursor()
        cursor.execute(
            "UPDATE users SET usb_access = 1, usb_access_expiry = %s WHERE id = %s",
            (expiry_date_obj, user_id)
        )
        connection.commit()
        cursor.close()
        connection.close()

        return redirect(url_for('usb_control'))  # Adjust the route name accordingly

    except ValueError:
        return jsonify({"error": "Invalid date format"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500
        
# ✅ Manual Revoke (Admin Action)
@app.route('/revoke_usb_access/<int:user_id>', methods=['POST'])
def revoke_usb_access(user_id):
    try:
        connection = get_db_connection()
        cursor = connection.cursor(dictionary=True)

        # Revoke USB access immediately (manual action)
        cursor.execute(
            "UPDATE users SET usb_access = 0, usb_access_expiry = NULL WHERE id = %s",
            (user_id,)
        )
        connection.commit()

        cursor.close()
        connection.close()
        return redirect(url_for('usb_control'))  # Adjust the route name accordingly

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/check_usb_expiry')
def check_usb_expiry():
    try:
        connection = get_db_connection()
        cursor = connection.cursor(dictionary=True)

        # Check if any users' expiry has passed and update them
        cursor.execute("""
            UPDATE users SET usb_access = 0, usb_access_expiry = NULL 
            WHERE usb_access = 1 AND usb_access_expiry <= NOW()
        """)
        connection.commit()

        # Get the updated user access data
        cursor.execute("SELECT id, usb_access FROM users")
        users = cursor.fetchall()

        cursor.close()
        connection.close()

        return jsonify(users)

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    

def check_local_usb_status():
    try:
        result = subprocess.run(
            ['powershell', 'Get-WmiObject', 'Win32_USBHub'],
            capture_output=True, text=True, startupinfo=startupinfo
        )
        
        if "Status" in result.stdout and "Error" not in result.stdout:
            print("USB devices found and functional.")
            return True
        else:
            print("USB devices found but have issues:", result.stdout)
            return False

    except Exception as e:
        print("Error checking USB status:", str(e))
        return False

# ✅ Flask Route to Get USB Status
@app.route('/get_usb_status')
def get_usb_status():
    try:
        connection = get_db_connection()  # Ensure the database connection function is correct
        cursor = connection.cursor(dictionary=True)

        # Get all users and their usb_access values from the database
        cursor.execute("SELECT id, usb_access FROM users")
        users = cursor.fetchall()

        local_usb_status = {}
        usb_enabled = check_local_usb_status()  # Check USB status
        
        for user in users:
            user_id = user['id']
            db_usb_access = user['usb_access']
            
            local_usb_status[user_id] = {
                "db_usb_access": db_usb_access,
                "local_usb_enabled": usb_enabled
            }

        cursor.close()
        connection.close()

        return jsonify(local_usb_status)

    except Exception as e:
        print("Error:", str(e))  # Debugging output
        return jsonify({"error": str(e)}), 500

# Database URI
app.config['SQLALCHEMY_DATABASE_URI'] = f"mysql+pymysql://{db_config['user']}:{db_config['password']}@{db_config['host']}:{db_config['port']}/{db_config['database']}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# Initialize SocketIO
socketio = SocketIO(app, async_mode='threading')

# User Model
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    usb_access = db.Column(db.Boolean, nullable=False)

# Function to enable USB controllers
# Function to enable USB controllers
def enable_usb_controllers():
    """Enable USB controllers based on OS."""
    os_name = platform.system()

    if os_name == "Windows":
        print("Enabling USB controllers on Windows...")
        command = """
        Get-WmiObject Win32_PnPEntity | Where-Object { $_.Caption -like "USB*" } | ForEach-Object {
            Enable-PnpDevice -InstanceId $_.PNPDeviceID -Confirm:$false
        }
        """
        try:
            result = subprocess.run(["powershell", "-Command", command], capture_output=True, text=True, startupinfo=startupinfo)
            if result.returncode == 0:
                print("USB controllers enabled successfully on Windows.")
            else:
                print("Error enabling USB controllers:", result.stderr)
        except Exception as e:
            print("Failed to enable USB on Windows:", str(e))

    # elif os_name == "Darwin":
    #     print("Enabling USB controllers on macOS...")
    #     try:
    #         subprocess.run(["sudo", "kextload", "/System/Library/Extensions/IOUSBMassStorageClass.kext"], check=True, startupinfo=startupinfo)
    #         print("USB enabled successfully on macOS.")
    #     except subprocess.CalledProcessError as e:
    #         print("Failed to enable USB on macOS:", str(e))

    elif os_name == "Darwin":
        print("Enabling USB controllers on macOS...")
        try:
            subprocess.run(["sudo", "pmset", "-a", "disablesleep", "0"], check=True, startupinfo=startupinfo)  # Allows USB wake-up
            subprocess.run(["sudo", "nvram", "boot-args=usb=on"], check=True, startupinfo=startupinfo)  # Enables USB in NVRAM
            print("USB enabled successfully on macOS.")
        except subprocess.CalledProcessError as e:
            print("Failed to enable USB on macOS:", str(e))

    elif os_name == "Linux":
        print("Enabling USB controllers on Linux...")
        try:
            subprocess.run(["sudo", "modprobe", "usb_storage"], check=True, startupinfo=startupinfo)
            print("USB enabled successfully on Linux.")
        except subprocess.CalledProcessError as e:
            print("Failed to enable USB on Linux:", str(e))

    else:
        print(f"Unsupported OS: {os_name}")

def disable_usb_controllers():
    """Disable USB controllers based on OS."""
    os_name = platform.system()

    if os_name == "Windows":
        print("Disabling USB controllers on Windows...")
        command = """
        Get-WmiObject Win32_PnPEntity | Where-Object { $_.Caption -like "USB*" } | ForEach-Object {
            Disable-PnpDevice -InstanceId $_.PNPDeviceID -Confirm:$false
        }
        """
        try:
            result = subprocess.run(["powershell", "-Command", command], capture_output=True, text=True, startupinfo=startupinfo)
            if result.returncode == 0:
                print("USB controllers disabled successfully on Windows.")
            else:
                print("Error disabling USB controllers:", result.stderr)
        except Exception as e:
            print("Failed to disable USB on Windows:", str(e))

    # elif os_name == "Darwin":
    #     print("Disabling USB controllers on macOS...")
    #     try:
    #         subprocess.run(["sudo", "kextunload", "/System/Library/Extensions/IOUSBMassStorageClass.kext"], check=True, startupinfo=startupinfo)
    #         print("USB disabled successfully on macOS.")
    #     except subprocess.CalledProcessError as e:
    #         print("Failed to disable USB on macOS:", str(e))

    elif os_name == "Darwin":
        print("Disabling USB controllers on macOS...")
        try:
            subprocess.run(["sudo", "pmset", "-a", "disablesleep", "1"], check=True, startupinfo=startupinfo)  # Prevent USB wake-up
            subprocess.run(["sudo", "nvram", "boot-args=usb=off"], check=True, startupinfo=startupinfo)  # Disable USB in NVRAM
            print("USB disabled successfully on macOS.")
        except subprocess.CalledProcessError as e:
            print("Failed to disable USB on macOS:", str(e))

    elif os_name == "Linux":
        print("Disabling USB controllers on Linux...")
        try:
            subprocess.run(["sudo", "modprobe", "-r", "usb_storage"], check=True, startupinfo=startupinfo)
            print("USB disabled successfully on Linux.")
        except subprocess.CalledProcessError as e:
            print("Failed to disable USB on Linux:", str(e))

    else:
        print(f"Unsupported OS: {os_name}")

# Function to fetch the user's usb_access value
def get_user_usb_access(user_id):
    user = User.query.filter_by(id=user_id).first()
    if user:
        return user.usb_access
    return False  # Default to False if no user found

# Periodic task to fetch usb_access and take action based on each user's usb_access value
# Function to check and execute USB access for the current logged-in user
def check_usb_access_for_current_user(user_id):
    with app.app_context():
        # Fetch the current logged-in user by user_id
        user = User.query.filter_by(id=user_id).first()
        if user:
            usb_access_value = user.usb_access  # Get the usb_access value for the user
            print(f"Current User ID: {user.id}, usb_access: {usb_access_value}")  # Debug print

            if usb_access_value:  # USB access is enabled
                print(f"Enabling USB for User ID: {user.id}")  # Debug print
                enable_usb_controllers()
            else:  # USB access is disabled
                print(f"Disabling USB for User ID: {user.id}")  # Debug print
                disable_usb_controllers()
        else:
            print(f"User with ID {user_id} not found.")  # Debug print

# Start the background scheduler for the current logged-in user
# Initialize the scheduler
scheduler = BackgroundScheduler()

def start_periodic_check(user_id):
    """Start the periodic check for the given user ID if the scheduler is not already running."""
    if not scheduler.running:  # Check if the scheduler is already running
        scheduler.start()  # Start the scheduler only if it's not running
    scheduler.add_job(func=lambda: check_usb_access_for_current_user(user_id), trigger="interval", seconds=10)


# Send email function
def send_alert_email(to_email, threats, timestamp):
    try:
        # Ensure the app context is active when sending email
        with current_app.app_context():
            subject = "⚠️ Security Alert: Threats Detected on Your Device!"
            body = f"""Dear User,

Our system detected the following threats on your device during the last check on {timestamp}:

{threats}

Please take immediate action to address these threats.

Stay Safe,
Your Security Team
            """.strip()  # Prevent extra spaces

            # Create the email message
            msg = Message(subject, recipients=[to_email])
            msg.body = body
            
            # Send the email
            mail.send(msg)
            print(f"✅ Alert email sent to {to_email}.")
    except Exception as e:
        print(f"❌ Failed to send email to {to_email}: {str(e)}")


# Function to check for threats and send emails
def check_and_send_alerts():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        print("🔍 Checking for threats in the database...")  # Debug Log

        query = """
            SELECT u.id AS user_id, u.email, 
            MAX(dr.last_check_timestamp) AS last_check_timestamp, 
            GROUP_CONCAT(DISTINCT dr.detected_threats ORDER BY dr.last_check_timestamp DESC) AS detected_threats
            FROM device_report dr
            INNER JOIN users u ON dr.existing_table_id = u.id
            LEFT JOIN email_alerts_log eal ON u.id = eal.user_id
            WHERE dr.detected_threats IS NOT NULL 
                AND (eal.last_sent IS NULL OR TIMESTAMPDIFF(HOUR, eal.last_sent, NOW()) >= 24) 
                AND dr.last_check_timestamp IS NOT NULL
            GROUP BY u.id, u.email;
        """

        cursor.execute(query)
        results = cursor.fetchall()

        if not results:
            print("✅ No threats found or emails already sent within 24 hours.")
        else:
            for record in results:
                print(f"⚠️ Sending alert email to {record['email']} - Threats: {record['detected_threats']}")
                send_alert_email(record['email'], record['detected_threats'], record['last_check_timestamp'])

                # Insert or update the log table
                cursor.execute("""
                    INSERT INTO email_alerts_log (user_id, email, last_sent) 
                    VALUES (%s, %s, NOW()) 
                    ON DUPLICATE KEY UPDATE last_sent = NOW()
                """, (record["user_id"], record["email"]))

        conn.commit()  # Save changes to the database

    except Exception as e:
        print("❌ Error while checking threats and sending alerts:", str(e))
    finally:
        cursor.close()
        conn.close()


# @app.route('/test-email')
# def test_email():
#     check_and_send_alerts()  # This will trigger sending emails to all users with threats
#     return "Test email sent (Check console for logs)."


def run_alert_in_thread():
    with app.app_context():  # Ensure app context is available
        check_and_send_alerts()
 
@app.route('/profile')
@login_required
def profile():
    start_time = time.time()
    user_id = session.get('user_id')
    if not user_id:
        flash("Please login to access your profile.")
        return redirect(url_for('login_form'))
    
    try:
        connection = mysql.connector.connect(**db_config)
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT username, phone, email, role, payment_due_date, usb_access, usb_access_expiry, bluetooth_access 
            FROM users 
            WHERE id = %s
        """, (user_id,))
        user_data = cursor.fetchone()
    except Exception as e:
        flash(f"Error fetching profile data: {e}")
        user_data = None
    finally:
        cursor.close()
        connection.close()

    if not user_data:
        flash("User data not found.")
        return redirect(url_for('login_form'))
    print(f"Profile route took {time.time() - start_time:.2f} seconds")
    return render_template('profile.html', user=user_data)

# @app.route('/profile_report_download')
# @login_required
# def profile_report_download():
#     start_time = time.time()
#     user_id = session.get('user_id')
#     if not user_id:
#         flash("Please login to download the report.")
#         return redirect(url_for('login_form'))

#     try:
#         connection = mysql.connector.connect(**db_config)
#         cursor = connection.cursor(dictionary=True)
#         cursor.execute("SELECT report FROM users WHERE id = %s", (user_id,))
#         user_data = cursor.fetchone()
#         if not user_data or not user_data.get('report'):
#             flash("No report available to download.")
#             return redirect(url_for('profile'))

#         source_html = user_data['report']

#         pdf_content = convert_html_to_pdf(source_html)
#         if not pdf_content:
#             flash("Failed to generate report PDF.")
#             return redirect(url_for('profile'))

#         pdf_with_page_numbers = add_page_numbers(pdf_content)

#         pdf_filename = f"Device_Report_{user_id}.pdf"
#         return send_file(
#             BytesIO(pdf_with_page_numbers),
#             mimetype='application/pdf',
#             download_name=pdf_filename,
#             as_attachment=True
#         )
#     except Exception as e:
#         flash(f"Error generating report: {e}")
#         print(f"PDF generation took {time.time() - start_time:.2f} seconds")
#         return redirect(url_for('profile'))
#     finally:
#         cursor.close()
#         connection.close()
 
@app.route('/settings')
@login_required
def settings():
    return render_template('settings.html')

@app.route('/settings/password-reset', methods=['POST'])
@login_required
def settings_password_reset_request():
    email = request.form['email']

    if not email_regex.match(email):
        flash("Invalid email address.")
        return redirect(url_for('settings'))  # Stay on settings page

    try:
        connection = mysql.connector.connect(**db_config)
        if connection.is_connected():
            cursor = connection.cursor(dictionary=True)
            cursor.execute("SELECT * FROM users WHERE email = %s", (email,))
            user = cursor.fetchone()

            if user:
                reset_token = str(uuid.uuid4())
                reset_token_expiry = datetime.utcnow() + timedelta(minutes=10)
                reset_otp = f"{random.randint(100000, 999999)}"  # 6-digit OTP
                reset_otp_expiry = datetime.utcnow() + timedelta(minutes=10)

                cursor.execute("""UPDATE users SET reset_token=%s, reset_token_expiry=%s, reset_otp=%s, reset_otp_expiry=%s WHERE email=%s""",
                               (reset_token, reset_token_expiry, reset_otp, reset_otp_expiry, email))
                connection.commit()

                reset_link = url_for('settings_verify_otp_form', token=reset_token, _external=True)

                msg = Message("Password Reset OTP", recipients=[email])
                msg.body = f"""Your OTP for password reset is: {reset_otp}

Click the following link to verify OTP and reset your password:
{reset_link}

This OTP is valid for 10 minutes."""
                mail.send(msg)
                flash("Password reset OTP sent to your email! Please check your inbox.")
                return redirect(url_for('settings_verify_otp_form', token=reset_token))
            else:
                flash("Email not found.")
    except Error as e:
        flash(f"Database error: {e}")
    except Exception as e:
        flash(f"Error sending email: {e}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

    return redirect(url_for('settings'))  # Stay on settings page

@app.route('/settings/verify-otp/<token>', methods=['GET'])
def settings_verify_otp_form(token):
    return render_template('settings_verify_otp.html', token=token)

@app.route('/settings/verify-otp/<token>', methods=['POST'])
def settings_verify_otp_process(token):
    otp_entered = request.form['otp']

    try:
        connection = mysql.connector.connect(**db_config)
        if connection.is_connected():
            cursor = connection.cursor(dictionary=True)
            cursor.execute("SELECT * FROM users WHERE reset_token = %s", (token,))
            user = cursor.fetchone()

            if not user:
                flash("Invalid or expired reset token.")
                return redirect(url_for('settings'))  # Stay on settings page

            if datetime.utcnow() > user['reset_otp_expiry']:
                flash("OTP expired. Please request password reset again.")
                return redirect(url_for('settings'))  # Stay on settings page

            if otp_entered != user['reset_otp']:
                flash("Incorrect OTP. Please try again.")
                return redirect(url_for('settings_verify_otp_form', token=token))

            # OTP verified successfully, redirect to password reset form
            return redirect(url_for('settings_password_reset_form', token=token))
    except Exception as e:
        flash(f"Error occurred: {e}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

    return redirect(url_for('settings'))  # Stay on settings page

@app.route('/settings/password-reset/<token>', methods=['GET'])
def settings_password_reset_form(token):
    # verify token is valid and not expired
    try:
        connection = mysql.connector.connect(**db_config)
        if connection.is_connected():
            cursor = connection.cursor(dictionary=True)
            cursor.execute("SELECT * FROM users WHERE reset_token = %s", (token,))
            user = cursor.fetchone()

            if not user or datetime.utcnow() > user['reset_token_expiry']:
                flash("Reset token expired or invalid.")
                return redirect(url_for('settings'))  # Stay on settings page
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

    return render_template('settings_password_reset.html', token=token)

@app.route('/settings/password-reset/<token>', methods=['POST'])
def settings_password_reset_process(token):
    new_password = request.form['new_password']
    confirm_password = request.form['confirm_password']

    if new_password != confirm_password:
        flash("Passwords do not match.")
        return redirect(url_for('settings_password_reset_form', token=token))
    
    if len(new_password) < 6:
        flash("Password must be at least 6 characters long.")
        return redirect(url_for('settings_password_reset_form', token=token))
    
    hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt())

    # Hash the password securely before storing it

    try:
        connection = mysql.connector.connect(**db_config)
        if connection.is_connected():
            cursor = connection.cursor()
            cursor.execute("SELECT * FROM users WHERE reset_token = %s", (token,))
            user = cursor.fetchone()

            if not user:
                flash("Invalid reset token.")
                return redirect(url_for('settings'))  # Stay on settings page

            cursor.execute("""
                UPDATE users SET password = %s, reset_token = NULL, reset_token_expiry = NULL, reset_otp = NULL, reset_otp_expiry = NULL
                WHERE reset_token = %s
            """, (hashed_password, token))
            connection.commit()
            flash("Password has been reset successfully! Please login.")
            return redirect(url_for('login'))  # Redirect to login after password reset
    except Exception as e:
        print(f"[ERROR] Password reset failed: {e}")
        flash("Something went wrong. Please try again.")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

    return redirect(url_for('settings'))  # Stay on settings page

 
@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/usb-control', methods=['GET', 'POST'])
@admin_required  # Only allow admin users to access this route
def usb_control():

    organization = session.get('organization')
    user_id = session.get('user_id')
    if not organization:
        flash("Session expired or invalid. Please log in again.")
        return redirect(url_for('login_form'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch current admin's license expiry date (same as admin_dashboard)
    cursor.execute('SELECT payment_due_date FROM users WHERE id = %s', (user_id,))
    admin_license = cursor.fetchone()
    license_expiry_date = admin_license['payment_due_date'] if admin_license else None

    # Handle POST request to update USB access and expiry date
    if request.method == 'POST':
        cursor.execute('SELECT id, usb_access, usb_access_expiry FROM users WHERE organization = %s', (organization,))
        users = cursor.fetchall()

        for user in users:
            usb_access = f'usb_access_{user["id"]}' in request.form
            expiry_date = request.form.get(f'expiry_date_{user["id"]}')
            cursor.execute('UPDATE users SET usb_access = %s, usb_access_expiry = %s WHERE id = %s', 
                           (usb_access, expiry_date, user['id']))
        conn.commit()

    cursor.execute('SELECT id, username, usb_access, usb_access_expiry FROM users WHERE organization = %s', (organization,))
    users = cursor.fetchall()

    cursor.close()
    conn.close()

    current_date = datetime.today().date()
    max_date = current_date + timedelta(days=10)

    return render_template('USB-Control.html', users=users, current_date=current_date, max_date=max_date, license_expiry_date=license_expiry_date)


@app.route('/logout', methods=['POST'])
@login_required
def logout():
    try:
        if 'user_id' in session:
            connection = mysql.connector.connect(**db_config)
            if connection.is_connected():
                cursor = connection.cursor()
                # Update login status to 'logged_out'
                update_status_query = "UPDATE users SET login_status = 'logged_out' WHERE id = %s"
                cursor.execute(update_status_query, (session['user_id'],))
                connection.commit()
                
    except Error as e:
        flash(f"Error occurred while logging out: {e}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

    # Clear session to avoid printing user info (e.g., usb_access)
    session.clear()
    
    flash("You have been logged out.")
    
    # Prevent caching of the page
    response = make_response(redirect(url_for('login_form')))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

 
@app.route('/forgot-password', methods=['GET'])
def forgot_password_form():
    return render_template('forgot_password.html')
@app.route('/forgot-password', methods=['POST'])
def request_password_reset():
    email = request.form['email']

    # Validate the email format
    if not email_regex.match(email):
        flash("Invalid email address.")
        return redirect(url_for('forgot_password_form'))
 
    try:
        connection = mysql.connector.connect(**db_config)
        if connection.is_connected():
            cursor = connection.cursor()
            cursor.execute("SELECT * FROM users WHERE email = %s", (email,))
            user = cursor.fetchone()

            if user:
                # Generate a unique reset token
                reset_token = str(uuid.uuid4())

                # Set the expiration time for the token (10 minutes)
                reset_token_expiry = datetime.utcnow() + timedelta(minutes=10)

                cursor.execute("UPDATE users SET reset_token = %s, reset_token_expiry = %s WHERE email = %s", 
                               (reset_token, reset_token_expiry, email))
                connection.commit()
               
                # Construct the reset link
                reset_link = url_for('reset_password', token=reset_token, _external=True)
 
                # Send the email
                msg = Message("Password Reset Request", recipients=[email])
                msg.body = f"""
Hi,

We received a request to reset your password.

Please click the link below to reset your password:
{reset_link}

Note: This link will expire in 10 minutes. If you did not request a password reset, you can safely ignore this email.

Thanks,
Security Team
                """.strip()
                mail.send(msg)
                flash("Password reset email sent!")
            else:
                flash("Email not found.")
    except Error as e:
        flash("Something went wrong. Please try again later.")
        print(f"[ERROR] DB error: {e}")
    except Exception as e:
        flash(f"Error occurred while sending email: {e}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
 
    return redirect(url_for('forgot_password_form'))
 
@app.route('/reset-password/<token>', methods=['GET'])
def reset_password(token):
    try:
        connection = mysql.connector.connect(**db_config)
        if connection.is_connected():
            cursor = connection.cursor(dictionary=True)  # Use dictionary cursor
            cursor.execute("SELECT * FROM users WHERE reset_token = %s", (token,))
            user = cursor.fetchone()

            if user:
                reset_token_expiry = user['reset_token_expiry']
                current_time = datetime.utcnow()

                # Check if the token has expired
                if reset_token_expiry < current_time:
                    flash("This password reset link has expired.")
                    return redirect(url_for('forgot_password_form'))
                return render_template('reset_password.html', token=token)
            else:
                flash("Invalid token.")
    except Error as e:
        flash("Something went wrong. Please try again.")
        print(f"[ERROR] Reset password token validation failed: {e}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

    return redirect(url_for('forgot_password_form'))
 
@app.route('/reset-password/<token>', methods=['POST'])
def update_password(token):
    new_password = request.form['new_password']
 
    try:
        connection = mysql.connector.connect(**db_config)
        if connection.is_connected():
            cursor = connection.cursor()
            cursor.execute("SELECT * FROM users WHERE reset_token = %s", (token,))
            user = cursor.fetchone()

            if user:
                hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                cursor.execute("UPDATE users SET password = %s, reset_token = NULL, reset_token_expiry = NULL WHERE reset_token = %s",
                               (hashed_password, token))
                connection.commit()
                flash("Your password has been reset! You can now log in.")
                return redirect(url_for('login_form'))
            else:
                flash("Invalid token.")
    except Error as e:
        flash(f"Error occurred: {e}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
 
    return redirect(url_for('forgot_password_form'))
    

@app.route('/update-admin-email', methods=['POST'])
@login_required
def update_admin_email():
    if 'user_id' not in session:
        flash("User is not logged in.")
        return redirect(url_for('login_form'))

    new_admin_email = request.form.get('admin_email')
    if not new_admin_email:
        flash("Admin email is required.")
        return redirect(url_for('settings'))
    
    # ✅ Email format validation
    EMAIL_REGEX = re.compile(r'^[^@]+@[^@]+\.[^@]+$')
    if not EMAIL_REGEX.match(new_admin_email):
        flash("Invalid email format.")
        return redirect(url_for('settings'))

    try:
        # Database connection
        connection = mysql.connector.connect(**db_config)
        cursor = connection.cursor()

        # Update the admin email for the logged-in user
        query = "UPDATE users SET admin_email = %s WHERE id = %s"
        cursor.execute(query, (new_admin_email, session['user_id']))
        connection.commit()

        # Update the session variable to reflect the new admin email
        session['admin_email'] = new_admin_email
        flash("Admin email updated successfully!")

        # Return the updated admin email to the settings page
        return redirect(url_for('settings'))  # Redirect back to settings page

    except mysql.connector.Error as e:
        print(f"[ERROR] Admin email update failed: {e}")
        flash("Something went wrong while updating admin email.")
        return redirect(url_for('settings'))

    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()
import string
def generate_unique_license_key():
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        while True:
            segments = []
            for _ in range(5):
                segment = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
                segments.append(segment)
            license_key = "LPS-" + "-".join(segments)

            # Check if exists
            cursor.execute("SELECT COUNT(*) AS count FROM users WHERE license_key=%s", (license_key,))
            count = cursor.fetchone()[0]

            if count == 0:
                return license_key
            # else, loop again to generate new key
    finally:
        cursor.close()
        conn.close()

import hashlib

from flask import request, redirect, url_for, flash, session, render_template
import hashlib
from datetime import datetime
from dateutil.relativedelta import relativedelta
import mysql.connector

import smtplib
def send_user_email(to_email, username, password, license_key, expiry_date):
    subject = "Your new ViCTAA credentials & license key"
    body = f"""
Hello {username},

✅ Your account has been created. Here are your details:

➡ Username: {username}
➡ Password: {password}
➡ Valid Until: {expiry_date}

Please login and activate your license.

Thank you,
ViCTAA Security Team
"""
    try:
        msg = Message(subject, recipients=[to_email])
        msg.body = body
        mail.send(msg)
        print(f"✅ Email sent to {to_email}")
    except Exception as e:
        print(f"❌ Failed to send email to {to_email}: {e}")


@app.route('/create_user_license', methods=['POST'])
@login_required
def create_user_license():
    username = request.form['username']
    phone = request.form['phone']
    email = request.form['email']
    password = request.form['password']
    license_id = request.form['license_id']
    license_key = request.form['license_key']
    organization = request.form['organization']
    role = request.form['role']

    # Hash the password before storing it in the database
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Validate license exists & is not used
    cursor.execute("SELECT * FROM licenses WHERE id = %s AND license_key = %s AND status = 'not_used'", (license_id, license_key))
    lic = cursor.fetchone()
    if not lic:
        flash("Invalid or already used license.")
        return redirect(url_for('Licensing'))

    today = datetime.now().date()
    expiry_date = today + relativedelta(months=lic['validity_months'])

    # Validate that the username, phone, or email are not already taken
    # Validate that the username, phone, or email are not already taken
    cursor.execute("SELECT COUNT(*) FROM users WHERE username = %s", (username,))
    result = cursor.fetchone()
    if result and result['COUNT(*)'] > 0:
        flash("Username already exists!")
        return redirect(url_for('Licensing'))

    cursor.execute("SELECT COUNT(*) FROM users WHERE phone = %s", (phone,))
    result = cursor.fetchone()
    if result and result['COUNT(*)'] > 0:
        flash("Phone number already taken!")
        return redirect(url_for('Licensing'))

    cursor.execute("SELECT COUNT(*) FROM users WHERE email = %s", (email,))
    result = cursor.fetchone()
    if result and result['COUNT(*)'] > 0:
        flash("Email already taken!")
        return redirect(url_for('Licensing'))
    try:
        # Insert user securely
        cursor.execute("""
            INSERT INTO users (username, phone, email, password, organization, license_key, payment_due_date, remaining_license_months, role)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            username, phone, email, hashed_password, 
            organization, license_key, expiry_date, 
            lic['validity_months'], role
        ))

        user_id = cursor.lastrowid

        # Mark license as used
        cursor.execute("""
            UPDATE licenses SET status = 'used', assigned_user_id = %s, assigned_at = NOW()
            WHERE id = %s
        """, (user_id, license_id))

        conn.commit()
        flash(f"✅ User created & license assigned! Sent credentials to {email}.")
        send_user_email(email, username, password, license_key, expiry_date)

    except Exception as e:
        flash(f"Error creating user: {str(e)}", "error")
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('Licensing'))

# @app.route('/create_user_license', methods=['POST'])
# @login_required
# def create_user_license():
#     username = request.form['username']
#     phone = request.form['phone']
#     email = request.form['email']
#     password = request.form['password']
#     license_id = request.form['license_id']
#     license_key = request.form['license_key']
#     organization = request.form['organization']
#     role = request.form['role']

#     hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

#     conn = get_db_connection()
#     cursor = conn.cursor(dictionary=True)

#     # ✅ Validate license exists & is not used
#     cursor.execute("SELECT * FROM licenses WHERE id=%s AND license_key=%s AND status='not_used'", (license_id, license_key))
#     lic = cursor.fetchone()
#     if not lic:
#         flash("Invalid or already used license.")
#         return redirect(url_for('Licensing'))

#     today = datetime.now().date()
#     expiry_date = today + relativedelta(months=lic['validity_months'])

#     # ✅ Insert user
#     cursor.execute("""
#         INSERT INTO users (username, phone, email, password, organization, license_key, payment_due_date, remaining_license_months, role)
#         VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
#     """, (
#         username, phone, email, hashed_password, 
#         organization, license_key, expiry_date, 
#         lic['validity_months'], role
#     ))

#     user_id = cursor.lastrowid

#     # ✅ Mark license as used
#     cursor.execute("""
#         UPDATE licenses SET status='used', assigned_user_id=%s, assigned_at=NOW()
#         WHERE id=%s
#     """, (user_id, license_id))

#     conn.commit()
#     flash(f"✅ User created & license assigned! Sent credentials to {email}.")
#     send_user_email(email, username, password, license_key, expiry_date)

#     cursor.close()
#     conn.close()
#     return redirect(url_for('Licensing'))

@app.route('/Licensing')
@login_required
def Licensing():
    user_id = session.get('user_id')
    if not user_id:
        flash("Session expired. Please log in again.")
        return redirect(url_for('login_form'))

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch admin organization
    cursor.execute("SELECT organization, payment_due_date FROM users WHERE id=%s", (user_id,))
    admin_info = cursor.fetchone()
    
    if not admin_info:
        flash("Admin details not found.")
        return redirect(url_for('login_form'))

    organization = admin_info['organization']
    license_expiry_date = admin_info['payment_due_date']

    # ✅ Calculate remaining licenses dynamically from licenses table
    cursor.execute("""
        SELECT COUNT(*) AS remaining_license_users
        FROM licenses 
        WHERE organization = %s AND status = 'not_used'
    """, (organization,))
    remaining_license_users = cursor.fetchone()['remaining_license_users']

    # ✅ Get licenses list for this organization
    cursor.execute("""
        SELECT id, license_key, type, validity_months, status, organization 
        FROM licenses
        WHERE organization = %s
    """, (organization,))
    licenses = cursor.fetchall()

    # ✅ Also get users for this organization
    cursor.execute("""
        SELECT id, username, license_key, payment_due_date 
        FROM users 
        WHERE organization = %s
    """, (organization,))
    users = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('Licensing.html', 
                           remaining_license_users=remaining_license_users,
                           licenses=licenses,
                           users=users,
                           license_expiry_date=license_expiry_date)

@app.route('/renew_license/<int:user_id>', methods=['POST'])
@login_required
def renew_license(user_id):
    validity = request.form['validity']

    # Check admin rights
    if not session.get('is_license_admin'):
        flash("❌ You are not authorized to renew licenses.")
        return redirect(url_for('Licensing'))

    months_required = {'1m': 1, '3m': 3, '12m': 12}[validity]

    if session.get('remaining_license_months', 0) < months_required:
        flash("❌ You do not have enough remaining license months to renew this license.")
        return redirect(url_for('Licensing'))

    # Renew the license by extending current expiry
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("SELECT payment_due_date FROM users WHERE id=%s", (user_id,))
        result = cursor.fetchone()
        if not result or result['payment_due_date'] is None:
            base_date = datetime.now()
        else:
            base_date = result['payment_due_date']

        new_expiry = base_date + timedelta(days=months_required * 30)

        cursor.execute("""
            UPDATE users SET payment_due_date=%s WHERE id=%s
        """, (new_expiry.strftime('%Y-%m-%d %H:%M:%S'), user_id))

        # Deduct from admin's remaining balance
        cursor.execute("""
            UPDATE users SET remaining_license_months = remaining_license_months - %s WHERE id=%s
        """, (months_required, session['user_id']))

        conn.commit()

        # Update session
        session['remaining_license_months'] -= months_required

        flash(f"✅ License renewed. New expiry: {new_expiry.strftime('%Y-%m-%d')}")
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('Licensing'))

def get_service_accounts():
    """Fetch list of service accounts running on Windows"""
    try:
        # PowerShell command to list all services and their accounts
        command = 'powershell.exe -Command "Get-WmiObject Win32_Service | Select-Object Name, StartName | ConvertTo-Json -Depth 5"'
        output = subprocess.check_output(command, shell=True, stderr=subprocess.PIPE)

        # Decode and parse the output from JSON format
        output_str = output.decode('utf-8')

        # Try to parse the output as JSON
        try:
            service_accounts = json.loads(output_str)
        except json.JSONDecodeError as json_error:
            print(f"[ERROR] JSON decode failed: {json_error}")
            print(f"[DEBUG] Raw output: {output_str}")
            return []

        # Filter relevant services (LocalSystem, NetworkService, NT AUTHORITY\LocalService)
        relevant_accounts = [
            {"service": service["Name"], "account": service["StartName"]}
            for service in service_accounts
            if service["StartName"] in ["LocalSystem", "NetworkService", "NT AUTHORITY\\LocalService"]
        ]

        return relevant_accounts

    except subprocess.CalledProcessError as e:
        print(f"[ERROR] PowerShell command failed: {e}")
        print(f"[DEBUG] PowerShell stderr: {e.stderr.decode()}")
        return []

from flask import session
# from tiny_llama_predictor import run_analysis_with_chunks

@app.route('/test-device', methods=['POST'])
@login_required
def test_device():
    pythoncom.CoInitialize()
    try:
        # Gather all necessary information
        cpu_info = get_cpu_info()
        memory_info = get_memory_info()
        disk_info = get_disk_info()
        network_info = get_network_info()
        installed_software = get_installed_software()
        usb_activity = get_usb_activity()
        logged_in_user = get_logged_in_user()
        usb_ports_count = get_usb_ports_count()
        vpn_status = check_vpn_status()
        usb_history = get_usb_history(usb_activity, days=10)
        sys_open_ports_data = sys_open_ports()
        usb_ports_status = get_usb_ports_status()
        patch_updates = get_patch_updates()
        battery_health = get_battery_health()
        service_accounts = get_service_accounts()
        threat_profile = get_device_type()

        # 🔍 Generate LLM threat prompt
        # 🔍 Generate LLM threat prompt
        # Limit the number of open ports and service accounts included in the prompt
        # llm_prompt = f"""
        # You are a cybersecurity expert. Given the following system configuration, please **list only the potential cybersecurity threats** based on this configuration.

        # **System Information:**
        # - OS: {threat_profile.get('os_name', 'Unknown')} {threat_profile.get('os_version', '')}
        # - Firewall: {"Enabled" if threat_profile.get('security_controls', {}).get('Firewall') else "Disabled"}
        # - BitLocker: {"Enabled" if threat_profile.get('security_controls', {}).get('BitLocker') else "Disabled"}
        # - Windows Defender: {"Running" if threat_profile.get('security_controls', {}).get('WindowsDefender') else "Not Running"}
        
        # **Service Accounts:**
        # """ + "\n".join([f"  - {svc['service']} under {svc['account']}" for svc in service_accounts]) + """
        
        # **Open Ports:**
        # """ + "\n".join([f"  - Port {p['port']} used by {p['process']}" for p in sys_open_ports_data]) + """

        # **Instructions:**
        # List only **distinct** cybersecurity threats, such as:
        # - Remote Code Execution
        # - Privilege Escalation
        # - Unauthorized Access
        # - Data Breach
        # - Denial of Service
        # - Elevation of Privileges
        # - SQL Injection

        # **Do not repeat any system information or threats**. 
        # Each threat should appear only once in the list.
        # """

        # # Chunk the prompt into 3 parts with a max token limit of 2000 tokens per chunk
        # chunks = chunk_prompt(llm_prompt, max_tokens=2000)
        
        # # List to store all distinct threats from all batches
        # llm_threats = []  # Updated name

        # # Process each chunk separately
        # for chunk in chunks:
        #     # Run the threat analysis for each batch
        #     llm_raw_output = run_threat_analysis(chunk)

        #     # Clean up the threat list (ensure we have distinct threats)
        #     lines = llm_raw_output.strip().splitlines()
        #     threats = []
        #     for line in lines:
        #         line = line.strip()
        #         if line.startswith("-"):
        #             threats.append(line)

        #     # Add threats from this batch to the final list
        #     llm_threats.extend(threats)

        # # Remove duplicates
        # llm_threats = list(dict.fromkeys(llm_threats))

        # # Store the predicted threats in session
        # session['predicted_threats'] = llm_threats


        # llm_prompt = f"""
        # You are a cybersecurity expert. Given the following system configuration, please **list only the potential cybersecurity threats** based on this configuration.

        # **System Information:**
        # - OS: {threat_profile.get('os_name', 'Unknown')} {threat_profile.get('os_version', '')}
        # - Firewall: {"Enabled" if threat_profile.get('security_controls', {}).get('Firewall') else "Disabled"}
        # - BitLocker: {"Enabled" if threat_profile.get('security_controls', {}).get('BitLocker') else "Disabled"}
        # - Windows Defender: {"Running" if threat_profile.get('security_controls', {}).get('WindowsDefender') else "Not Running"}
        
        # **Service Accounts:**
        # """ + "\n".join([f"  - {svc['service']} under {svc['account']}" for svc in service_accounts]) + """
        
        # **Open Ports:**
        # """ + "\n".join([f"  - Port {p['port']} used by {p['process']}" for p in sys_open_ports_data]) + """

        # **Instructions:**
        # List only **distinct** cybersecurity threats, such as:
        # - Remote Code Execution
        # - Privilege Escalation
        # - Unauthorized Access
        # - Data Breach
        # - Denial of Service
        # - Elevation of Privileges
        # - SQL Injection

        # **Do not repeat any system information or threats**. 
        # Each threat should appear only once in the list.
        # """

        # # Chunk the prompt into manageable batches (max 1024 tokens per batch)
        # chunks = chunk_prompt(llm_prompt, max_tokens=1024)

        # llm_threats = []  # List to store all predicted threats

        # # Process each chunk separately
        # for chunk in chunks:
        #     # Run the threat analysis for each batch
        #     llm_raw_output = run_threat_analysis(chunk)

        #     # Clean up the threat list (ensure we have distinct threats)
        #     lines = llm_raw_output.strip().splitlines()
        #     threats = []
        #     for line in lines:
        #         line = line.strip()
        #         if line.startswith("-"):
        #             threats.append(line)

        #     # Add threats from this batch to the final list
        #     llm_threats.extend(threats)

        # # Remove duplicates
        # llm_threats = list(dict.fromkeys(llm_threats))

        # # Store the predicted threats in session
        # session['predicted_threats'] = llm_threats

        # Generate threats based on system configuration
        # llm_threats = run_analysis_with_chunks(threat_profile, service_accounts, sys_open_ports_data)

        # # Store the predicted threats in session
        # session['predicted_threats'] = llm_threats
        
        print("Final Patch Updates Data Sent to HTML:", patch_updates)  # Debugging print
        print("Final Battery Health Data Sent to HTML:", battery_health)
        print(service_accounts)

#local_system_ports
        if sys_open_ports_data:
           print("Open ports detected:")
        for port_info in sys_open_ports_data:
            print(f"Port: {port_info['port']}, State: {port_info['state']}, Process: {port_info['process']}")

        usb_ports_status = get_usb_ports_status()
        for status in usb_ports_status:
            print(f"Port: {status['port']} - Status: {status['status']}")        
        
        # Get public IP
        public_ip = get_public_ip()
        print("Public IP:", public_ip)  # Debugging

# Scan open ports on the public IP
        if public_ip not in ["Unknown IP", "Timeout Error", "No Internet Connection", "Unknown Error"]:
            ports_to_scan = range(20, 1025)  # Define port range to scan
            open_ports = scan_open_ports(public_ip, ports_to_scan)
            open_ports_count = len(open_ports)
            print("Open Ports:", open_ports)  # Debugging
            print("Count_of_Ports:", open_ports_count)
        else:
            open_ports = ["Unable to scan ports due to public IP retrieval issue"]

        # Log VPN usage for history tracking
        log_vpn_usage()

        # Get VPN history for the last 24 hours
        vpn_history = get_vpn_history_last_24_hours()

        # Set directories to scan (multiple drives like C: and E:)
        directories_to_scan = ['C:\\', 'E:\\']  # Add or remove drives as needed
        cache_files = find_cache_files(directories_to_scan)  # Function to find browser cookies and cache
        restricted_files = find_restricted_files(directories_to_scan)

        # Handle recently used and unused applications
        recent_apps = check_recently_used_apps(cutoff_time_24_hours)
        print("Recently used apps:", recent_apps)  # Debug statement

        unused_apps = check_unused_apps(installed_apps, recently_used_apps)
        print("Unused apps:", unused_apps)  # Debug statement

        p_address = None
        if network_info:
            # Make sure network_info has a valid IP address key
            ip_address = network_info[0].get('IPAddress')
        
        if not ip_address:
            # Handle the case where IP address is not found
            raise ValueError("Unable to retrieve IP address")
        
        

        ip_address = network_info[0]['IPAddress']
        vulnerabilities = vulnerability_lookup_for_desktop(installed_software)  # Check for vulnerabilities on desktop
        threats = detect_all_threats(cpu_info, memory_info, installed_software, directories_to_scan) # Detect potential threats

        # Calculate the counts
        total_vulnerabilities = sum(len(v) for v in vulnerabilities.values())
        total_threats = len(threats)
        total_software_installed = len(installed_software)
        usb_ports_count = usb_ports_count

        # Store counts in session
        session['total_vulnerabilities'] = total_vulnerabilities
        session['total_threats'] = total_threats
        session['total_software_installed'] = total_software_installed
        session['usb_ports_count'] = usb_ports_count
        session['open_ports_count']=open_ports_count

        # Generate a complete report using the updated function
        report = generate_report(cpu_info, memory_info, disk_info, network_info, installed_software,
                                 vulnerabilities, threats, usb_activity, cache_files, restricted_files, logged_in_user,usb_ports_count,vpn_status,vpn_history,usb_history,recent_apps, unused_apps,open_ports,sys_open_ports_data,usb_ports_status,patch_updates, battery_health, threat_profile, service_accounts)
        user_id = session.get('user_id')
        save_report_to_db(user_id, report)
        # Call the function to save the data to the database
        save_device_data_to_db(user_id, cpu_info, memory_info, disk_info, network_info, usb_ports_count,
                               vulnerabilities, threats, usb_activity, vpn_status, usb_history, vpn_history,
                               recent_apps, unused_apps, report, installed_software,total_vulnerabilities,total_threats,total_software_installed,open_ports_count,usb_ports_status,service_accounts)
        admin_email = session.get('admin_email', 'default_admin@example.com')
        send_report_to_admin(admin_email, report)
        # Return the generated report (if needed)
        return report  # This will return the rendered report template directly from the generate_report function

    finally:
        pythoncom.CoUninitialize()

# Update generate_report function to accept restricted_files
def generate_report(cpu_info, memory_info, disk_info, network_info, software_list, vulnerabilities, threats, usb_activity,cache_files,restricted_files,logged_in_user,usb_ports_count,vpn_status,vpn_history,usb_history,recent_apps, unused_apps,open_ports,sys_open_ports_data,usb_ports_status,patch_updates, battery_health, threat_profile, service_accounts):
    try:
        # Render the report template with the provided information
        return render_template('report.html',
                               cpu_info=cpu_info,
                               memory_info=memory_info,
                               disk_info=disk_info,
                               network_info=network_info,
                               installed_software=software_list,
                               vulnerabilities=vulnerabilities,
                               threats=threats,
                               usb_activity=usb_activity,
                               cache_files=cache_files,
                               restricted_files=restricted_files,
                               logged_in_user=logged_in_user,
                               usb_ports_count=usb_ports_count,
                               vpn_status=vpn_status,
                               vpn_history=vpn_history,
                               usb_history=usb_history,
                               recent_apps=recent_apps, 
                               unused_apps=unused_apps,
                               open_ports=open_ports,
                               sys_open_ports=sys_open_ports_data,
                               usb_ports_status=usb_ports_status,
                               patch_updates=patch_updates, 
                               battery_health=battery_health,
                               threat_profile=threat_profile,
                               service_accounts=service_accounts) 
    except Exception as e:
        print(f"An error occurred while generating the report: {e}")
    finally:
        pythoncom.CoUninitialize()

# Function to save the device data to the database
def save_device_data_to_db(user_id, cpu_info, memory_info, disk_info, network_info, usb_ports_count,
                           vulnerabilities, threats, usb_activity, vpn_status, usb_history, vpn_history,
                           recent_apps, unused_apps, report, installed_software,total_vulnerabilities,total_threats,total_software_installed,open_ports_count,usb_ports_status,service_accounts):
    try:
        # Establish database connection
        connection = mysql.connector.connect(**db_config)
        if connection.is_connected():
            cursor = connection.cursor()

            # Prepare the data to be stored
            query = """
                INSERT INTO device_report (
                    existing_table_id, cpu_information, memory_information, disk_information,
                    network_information, vpn_detections, usb_activity_history, vpn_usage_history,
                    recently_used_applications, unused_applications, vulnerability_detections, detected_threats, total_installed_software,
                    usb_ports_count, vulnerabilities_count, threats_count, software_count,network_ports_count,last_check_timestamp,usb_device_id, usb_status,service_accounts

                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """

            # Format the data into strings (JSON or plain text) as needed for the database
            cursor.execute(query, (
                user_id,
                json.dumps(cpu_info),  # Convert lists or dicts into JSON format
                json.dumps(memory_info),
                json.dumps(disk_info),
                json.dumps(network_info),
                json.dumps(vpn_status),  # Assuming this is a list or dict
                json.dumps(usb_activity),
                json.dumps(vpn_history),
                json.dumps(recent_apps),
                json.dumps(unused_apps),
                json.dumps(vulnerabilities),
                json.dumps(threats),
                json.dumps(installed_software),
                usb_ports_count,
                total_vulnerabilities,
                total_threats,
                total_software_installed,
                json.dumps(open_ports_count), datetime.now(),
                json.dumps(usb_ports_status), "active",
                json.dumps(service_accounts)
            ))

            connection.commit()  # Commit the transaction
            print("Data saved to the database successfully.")

    except Error as e:
        print(f"Error saving data to the database: {e}")

    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()



def save_report_to_db(user_id, report_content):
    try:
        # Establish connection using db_config
        connection = mysql.connector.connect(**db_config)
        if connection.is_connected():
            cursor = connection.cursor()
            # SQL query to insert the report into the database
            query = "UPDATE users SET report = %s WHERE id = %s"
            # Execute the query with the report content and user_id
            cursor.execute(query, (report_content, user_id))
            connection.commit()  # Commit the changes
    except Error as e:
        print(f"Error saving report to database: {e}")
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

# def convert_html_to_pdf(source_html):
#     # Define the header with the logo and report title
#     header_html = """
#     <div style="text-align: center; width: 100%; margin-bottom: 10px;">
#         <img src="static/logo.png" width="100px" />
#         <h1>Device Report</h1>
#     </div>
#     """

#     # Extract section headings and create links for TOC
#     section_links = []
#     headings = re.findall(r'<h2 class="toggle-header">(.*?)</h2>', source_html)

#     for idx, heading in enumerate(headings):
#         section_id = f"section-{idx}"  # Unique section ID
        
#         # Modify TOC entry without page numbers
#         section_links.append(
#             f'<li><pdf:link destination="{section_id}" style="text-decoration: none; color: #3498db;">'
#             f'{heading}</pdf:link></li>'
#         )

#         # Modify source HTML to correctly reference each section
#         source_html = source_html.replace(
#             f'<h2 class="toggle-header">{heading}</h2>',
#             f'<pdf:bookmark name="{heading}" /><pdf:destination name="{section_id}" /><h2 id="{section_id}">{heading}</h2>'
#         )

#         # Add Table of Contents (TOC) without page numbers
#         toc_html = f"""
#         <h2 style="text-align: center; background-color: #2c3e50; color: white; padding: 5px; border-radius: 5px;">Table of Contents</h2>
#             <ul style="padding: 10px; font-size: 12px;">
#                 {''.join(section_links)}
#             </ul>
#         <pdf:nextpage />
#         """

#     # Define a persistent footer for page numbers
#     footer_html = """
#     <pdf:staticcontent name="footer" pages="all">
#         <div style="width: 100%; text-align: center; font-size: 10px; color: #555;">
#             Page <pdf:pagenumber /> of <pdf:pagecount />
#         </div>
#     </pdf:staticcontent>

#     """

#     # Ensure the footer appears on every page
#     footer_style = """
#     <style>
#         @page {
#             size: A4;
#             margin: 20mm;
            
#             /* Attach footer to EVERY page */
#             @frame footer_frame {
#                 -pdf-frame-content: footer;
#                 bottom: 10mm;
#                 height: 12mm;
#             }
#         }
#     </style>    

#     """
#     # Enhance the existing CSS for better formatting in PDF
#     enhanced_css = """
#     <style>
#         body {
#             font-family: Arial, sans-serif;
#             font-size: 11px;
#             line-height: 1.4;
#             text-align: left;
#             background-color: #ffffff;
#             padding: 10px;
#         }

#         h1 {
#             text-align: center;
#             font-size: 16px;
#             font-weight: bold;
#             color: #2c3e50;
#         }

#         h2 {
#             font-size: 14px;
#             font-weight: bold;
#             background-color: #3498db;
#             color: #ffffff;
#             padding: 6px;
#             border-radius: 3px;
#             margin-top: 10px;
#             margin-bottom: 6px;
#         }

#         ul {
#             list-style-type: none;
#             padding-left: 0;
#             font-size: 12px;
#         }

#         li {
#             font-size: 11px;
#             padding: 4px;
#         }

#         table {
#             width: 100%;
#             border-collapse: collapse;
#             margin-bottom: 10px;
#             background-color: #ffffff;
#         }

#         th, td {
#             border: 1px solid #ddd;
#             padding: 5px;
#             text-align: left;
#             font-size: 10px;
#             word-wrap: break-word;
#             white-space: pre-wrap;
#             max-width: 100%;
#         }

#         th {
#             background-color: #2c3e50;
#             color: #ffffff;
#             font-weight: bold;
#         }

#         tr:nth-child(even) {
#             background-color: #f9f9f9;
#         }

#         tr:hover {
#             background-color: #d5dbdb;
#         }
#     </style>
#     """
#     # Final HTML for PDF conversion
#     final_html = f"{enhanced_css}{header_html}{toc_html}{source_html}{footer_html}{footer_style}"

#     # Convert HTML to PDF
#     pdf_buffer = BytesIO()
#     pisa_status = pisa.CreatePDF(final_html, dest=pdf_buffer)

#     if pisa_status.err:
#         raise Exception("Error generating PDF")

#     pdf_content = pdf_buffer.getvalue()  # Convert BytesIO to bytes
#     pdf_buffer.close()

#     return pdf_content

#----------------------------------------------------------------------------------------------------------------------------------

import re
import base64
import matplotlib.pyplot as plt
import pandas as pd
import io
from xhtml2pdf import pisa
from flask import send_file
from io import BytesIO
import mysql.connector
from datetime import datetime, timedelta
 
 
def get_current_user_id():
    return session.get('user_id')  # Fetch user_id from the session where test_device is running
 
# Function to fetch the recent 7 scans for the current user based on existing_table_id
def fetch_recent_7_scans():
    user_id = get_current_user_id()
    if not user_id:
        print("No user_id found in session.")
        return []
 
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
 
        query = """
        SELECT last_check_timestamp,
               software_count,
               threats_count,
               vulnerabilities_count,
               usb_ports_count
        FROM device_report
        WHERE existing_table_id = %s
          AND cpu_information IS NOT NULL
        ORDER BY last_check_timestamp DESC
        LIMIT 7;
        """
        cursor.execute(query, (user_id,))
        result = cursor.fetchall()
        cursor.close()
        conn.close()
 
        return result
 
    except Exception as e:
        print(f"Database Error: {e}")
        return []
 
# Function to generate base64 images for graphs (Proper Bar Graph)
def generate_graphs():
    data = fetch_recent_7_scans()
    if not data:
        print("No data found for recent 7 scans.")
        return {}
 
    df = pd.DataFrame(data)
    graphs = {}
   
    metrics = ["software_count", "threats_count", "vulnerabilities_count", "usb_ports_count"]
   
    # Convert timestamps to string format to avoid histogram-like behavior
    df["last_check_timestamp"] = df["last_check_timestamp"].astype(str)
 
    for metric in metrics:
        plt.figure(figsize=(7, 4))
        plt.bar(df["last_check_timestamp"], df[metric], color="royalblue", width=0.5)
        plt.xlabel("Last Check Timestamp")
        plt.ylabel(metric.replace("_", " ").title())
        plt.title(f"{metric.replace('_', ' ').title()} Over Recent 7 Scans")
        plt.xticks(rotation=45, ha="right")  # Rotate X-axis labels for readability
        plt.grid(axis="y", linestyle="--", alpha=0.7)
 
        buffer = BytesIO()
        plt.savefig(buffer, format="png", bbox_inches="tight")
        buffer.seek(0)
        graphs[metric] = base64.b64encode(buffer.getvalue()).decode("utf-8")
        plt.close()
 
    return graphs
 
 
 
# Convert HTML to PDF function with Graphs Page
def convert_html_to_pdf(source_html):
   
    cover_html = """<br><br><br>
    <div style="text-align: center; width: 100%; height: 100vh; display: flex; flex-direction: column; justify-content: center; align-items: center;">
        <h1 style="font-size: 80px; font-weight: bold; margin-bottom: 10px;">Device Report</h1>
        <h3 style="font-size: 42px; font-weight: normal; margin-bottom: 40px;">For Vardaan Global</h3>
        <br>
        <br>
        <div style="position: absolute; bottom: 50px; width: 100%; text-align: center;">
            <p style="font-size: 24px; font-weight: bold;">Powered By:</p>
            <img src="static/logo.png" style="width: 80%; height: auto;" />
        </div>
    </div>
    <pdf:nextpage />
    """
 
    header_html = """
    <div style="text-align: center; width: 100%; margin-bottom: 10px;">
        <h1>Device Report</h1>
    </div>
    """
 
    # Extract section headings and create links for TOC
    section_links = []
    headings = re.findall(r'<h2 class="toggle-header">(.*?)</h2>', source_html)

    for idx, heading in enumerate(headings):
        section_id = f"section-{idx}"  # Unique section ID
        section_links.append(
            f'<li style="font-size: 12px; margin-bottom: 10px;">'  # Increased font size
            f'<pdf:link destination="{section_id}" style="text-decoration: none; color: #3498db; font-size: 20px;">'
            f'{heading}</pdf:link></li>'
        )
        source_html = source_html.replace(
            f'<h2 class="toggle-header">{heading}</h2>',
            f'<pdf:bookmark name="{heading}" /><pdf:destination name="{section_id}" /><h2 id="{section_id}">{heading}</h2>'
        )
 
    # Add Table of Contents (TOC) with clickable links and page numbers
    toc_html = f"""
    <h2 style="text-align: center; background-color: #2c3e50; color: white; padding: 5px; border-radius: 5px;">Table of Contents</h2>
    <ul style="padding: 15px; font-size: 12px;">
        {''.join(section_links)}
    </ul>
    <pdf:nextpage />
    """
 
    # Generate graphs and insert into PDF (All 4 Graphs on One Page)
    # Ensure graphs is generated before use
    graphs = generate_graphs()  # Ensure graphs are generated

    graphs_html = '<h2 style="text-align: center;">Recent 7 Scans Report</h2>'

    # Iterate through graphs and print 2 per page with increased spacing
    for idx, (metric, base64_img) in enumerate(graphs.items()):
        if idx % 2 == 0:  # Start a new page every 2 graphs
            graphs_html += '<div style="display: flex; flex-direction: column; align-items: center; gap: 180px; padding-top: 25px;">'

        graphs_html += f"""
            <div style="width: 90%; text-align: center;">
                <img src="data:image/png;base64,{base64_img}" style="width: 100%; height: 400px; object-fit: contain;" />
            </div>
        """

        if idx % 2 == 1 or idx == len(graphs) - 1:  # Close the container and add a page break after 2 graphs
            graphs_html += "</div><pdf:nextpage />"

 
    # Define a persistent footer for page numbers
    footer_html = """
    <pdf:staticcontent name="footer_content">
        <div style="width: 100%; text-align: center; font-size: 10px; color: #555;">
            
        </div>
    </pdf:staticcontent>
    """
 
    # Footer style
    footer_style = """
    <style>
        @page {
            size: A4;
            margin: 20mm;
            @frame footer_frame { -pdf-frame-content: footer_content; bottom: 10mm; height: 12mm; }
        }
    </style>
    """
 
    # CSS Styles
    enhanced_css = """
    <style>
        body {
            font-family: Arial, sans-serif;
            font-size: 11px;
            line-height: 1.4;
            text-align: left;
            background-color: #ffffff;
            padding: 10px;
        }
 
        h1 {
            text-align: center;
            font-size: 16px;
            font-weight: bold;
            color: #2c3e50;
        }
 
        h2 {
            font-size: 14px;
            font-weight: bold;
            background-color: #3498db;
            color: #ffffff;
            padding: 6px;
            border-radius: 3px;
            margin-top: 10px;
            margin-bottom: 6px;
        }
 
        h3 {
            font-size: 13px;
            font-weight: bold;
            color: #2c3e50;
            margin-bottom: 5px;
        }
 
        ul {
            list-style-type: none;
            padding-left: 0;
            font-size: 12px;
        }
 
        li {
            font-size: 11px;
            padding: 4px;
        }
 
        img {
            border: 1px solid #ddd;
            border-radius: 5px;
            padding: 5px;
            background: #f9f9f9;
        }
    </style>
    """
 
    # Final HTML for PDF conversion
    final_html = f"{footer_style}{enhanced_css}{cover_html}{header_html}{toc_html}{graphs_html}{source_html}{footer_html}"
 
    try:
        # Convert HTML to PDF
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(final_html, dest=pdf_buffer)
 
        if pisa_status.err:
            raise Exception("Error generating PDF")
 
        pdf_content = pdf_buffer.getvalue()
        pdf_buffer.close()
        return pdf_content
 
    except Exception as e:
        print(f"PDF Generation Failed: {e}")
        return None

#----------------------------------------------------------------------------------------------------------------------------------


def get_current_username():
    # Connect to the database
    connection = mysql.connector.connect(**db_config)
    cursor = connection.cursor()

    try:
        # Query to get the current logged-in user's username
        query = "SELECT username FROM users WHERE id = %s"  # Assuming 'id' is stored in the session
        user_id = session.get('user_id')  # Replace with dynamic user ID (e.g., session or authentication logic)
        cursor.execute(query, (user_id,))
        result = cursor.fetchone()

        if result:
            return result[0]  # Return the username
        else:
            return "Unknown User"  # Default in case no user is found
    except Exception as e:
        print(f"Error fetching username: {e}")
        return "Unknown User"  # Default in case of error
    finally:
        cursor.close()
        connection.close()

ftp_config = {
    'host': '202.53.78.150',
    'user': 'vardaan_ftp',
    'password': 'Vardaa@1234#',
    'upload_path': 'Vardaan Global/Device_Reports_VDS'  # Directory where reports will be uploaded
}


def add_page_numbers(pdf_content):
    # Create a buffer to write the PDF with added page numbers
    pdf_buffer = BytesIO()
 
    # Read the original PDF content into PdfReader
    original_pdf = PdfReader(BytesIO(pdf_content))
    pdf_writer = PdfWriter()
 
    # Iterate over all pages and add page numbers
    for page_num in range(len(original_pdf.pages)):
        page = original_pdf.pages[page_num]
 
        # Create a canvas to draw the page number
        packet = BytesIO()
        c = canvas.Canvas(packet, pagesize=letter)
        c.setFont("Helvetica", 10)
        c.drawString(500, 10, f"Page {page_num + 1}")
        c.save()
 
        # Move the packet to the beginning and append it to the page
        packet.seek(0)
        overlay_pdf = PdfReader(packet)
        overlay_page = overlay_pdf.pages[0]
 
        # Merge the page with the overlay (page number)
        page.merge_page(overlay_page)
 
        # Add the page with the page number to the PdfWriter
        pdf_writer.add_page(page)
 
    # Write the final PDF to the buffer
    pdf_writer.write(pdf_buffer)
    pdf_buffer.seek(0)
 
    # Return the modified PDF content
    return pdf_buffer.read()
    

def get_alternate_email_from_db():
    user_id = session.get('user_id')
    if not user_id:
        print("[ERROR] No user_id in session to fetch alternate email.")
        return None
    
    try:
        connection = mysql.connector.connect(**db_config)
        cursor = connection.cursor(dictionary=True)
        query = "SELECT email FROM users WHERE id = %s"
        cursor.execute(query, (user_id,))
        result = cursor.fetchone()
        cursor.close()
        connection.close()
        
        if result and result['email']:
            return result['email']
        else:
            print("[ERROR] Email not found for user id:", user_id)
            return None
    except Exception as e:
        print(f"[ERROR] DB error fetching email for user {user_id}: {e}")
        return None

def send_report_to_admin(admin_email, report_content):
    if not admin_email:
        print("[WARNING] Admin email is None, trying alternate email")
        admin_email = get_alternate_email_from_db()
    
    if not admin_email:
        print("[ERROR] No email found in admin_email or alternate email column. Continuing without sending email.")
        # Continue with rest of logic if needed, but skip sending email
        return

    pdf_content = convert_html_to_pdf(report_content)
    if not pdf_content:
        print("[ERROR] Failed to generate PDF from report")
        return

    pdf_with_page_numbers = add_page_numbers(pdf_content)
    if not pdf_with_page_numbers:
        print("[ERROR] Failed to add page numbers to PDF")
        return

    current_user = get_current_username() or "UnknownUser"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_filename = f"{current_user}_Report_{timestamp}.pdf"

    try:
        msg = Message("Device Report", sender="vdatasciences@gmail.com", recipients=[admin_email])
        msg.body = "Please find the attached report."
        msg.attach(pdf_filename, "application/pdf", pdf_with_page_numbers)
        mail.send(msg)
        print(f"[INFO] Report sent successfully to {admin_email}")
    except Exception as e:
        print(f"[ERROR] Failed to send email: {e}")


from flask import request, session, jsonify
from sqlalchemy import text

# @app.route('/get-detail')
# def get_detail():
#     type_map = {
#         'vulnerabilities': 'vulnerability_detections',
#         'threats': 'detected_threats',
#         'ports': 'network_ports_count',
#         'software': 'total_installed_software'
#     }

#     detail_type = request.args.get('type')
#     db_field = type_map.get(detail_type)

#     if not db_field:
#         return jsonify({"error": "Invalid type"}), 400

#     user_id = session.get('user_id')
#     if not user_id:
#         return jsonify({"error": "Unauthorized"}), 403

#     query = text(f"""
#         SELECT {db_field}
#         FROM device_report
#         WHERE existing_table_id = :uid
#         ORDER BY last_check_timestamp DESC
#         LIMIT 1
#     """)

#     result = db.session.execute(query, {'uid': user_id}).fetchone()

#     if result and result[0] is not None:
#         return jsonify({"data": result[0]})
#     else:
#         return jsonify({"data": "No data found"}), 200

@app.route('/get-detail')
@login_required
def get_detail():
    type_map = {
        'vulnerabilities': 'vulnerability_detections',
        'threats': 'detected_threats',
        'ports': 'network_information',  # assuming this holds port details list
        'software': 'total_installed_software'
    }

    detail_type = request.args.get('type')
    db_field = type_map.get(detail_type)

    if not db_field:
        return jsonify({"error": "Invalid type"}), 400

    user_id = session.get('user_id')
    if not user_id:
        return jsonify({"error": "Unauthorized"}), 403

    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch latest non-empty row for this user and field
        query = f"""
            SELECT {db_field}
            FROM device_report
            WHERE existing_table_id = %s AND {db_field} IS NOT NULL AND {db_field} != ''
            ORDER BY last_check_timestamp DESC
            LIMIT 1
        """
        cursor.execute(query, (user_id,))
        row = cursor.fetchone()

        if not row or not row[db_field]:
            return jsonify({"data": ["No data found"]})

        raw_data = row[db_field]

        # Try parsing JSON; if fail, return raw data split by commas
        try:
            import json
            parsed_data = json.loads(raw_data)
            if isinstance(parsed_data, dict):
                data_list = [f"{k}: {v}" for k, v in parsed_data.items()]
            elif isinstance(parsed_data, list):
                data_list = parsed_data
            else:
                data_list = [str(parsed_data)]
        except Exception:
            data_list = raw_data.split(',')

        return jsonify({"data": data_list})

    except Exception as e:
        print(f"Error in /get-detail: {e}")
        return jsonify({"data": ["Error fetching data"]}), 500

    finally:
        cursor.close()
        connection.close()


def upload_to_ftp(file_content, filename):
    try:
        # Debugging: Print the filename and file size
        print(f"Attempting to upload file: {filename}, size: {len(file_content)} bytes")

        # Connect to the FTP server
        with ftplib.FTP(ftp_config['host'], ftp_config['user'], ftp_config['password']) as ftp:
            ftp.cwd(ftp_config['upload_path'])  # Change to the desired directory
            
            # Debugging: Confirm directory change
            print(f"Current directory: {ftp.pwd()}")
            
            # Upload the file
            ftp.storbinary(f'STOR {filename}', BytesIO(file_content))
            
            print(f"Uploaded {filename} to FTP server successfully.")
    except ftplib.all_errors as e:
        print(f"FTP error: {e}")
    except Exception as general_error:
        print(f"General error: {general_error}")

# def schedule_startup():
#     print('Trying to Schedule a Task for LapSec')
#     system_name = platform.system()
#     script_path = os.path.abspath(__file__)  # Auto-detects script path

#     if system_name == "Windows":
#         task_name = "LapSec"
#         check_task = f'schtasks /query /tn "{task_name}"'
        
#         if subprocess.run(check_task, shell=True, capture_output=True, startupinfo=startupinfo).returncode == 0:
#             return  # Already scheduled
        
#         create_task = f'schtasks /create /tn "{task_name}" /tr "{script_path}" /sc onlogon /rl highest /f /IT /C'
#         subprocess.run(create_task, shell=True, startupinfo=startupinfo)

#     elif system_name == "Darwin":  # macOS
#         plist_path = os.path.expanduser("~/Library/LaunchAgents/com.myflaskapp.autostart.plist")
        
#         if os.path.exists(plist_path):
#             return  # Already scheduled

#         plist_content = f'''<?xml version="1.0" encoding="UTF-8"?>
#         <!DOCTYPE plist PUBLIC "-//Apple//DTD PLIST 1.0//EN" "http://www.apple.com/DTDs/PropertyList-1.0.dtd">
#         <plist version="1.0">
#         <dict>
#             <key>Label</key>
#             <string>com.myflaskapp.autostart</string>
#             <key>ProgramArguments</key>
#             <array>
#                 <string>{script_path}</string>
#             </array>
#             <key>RunAtLoad</key>
#             <true/>
#         </dict>
#         </plist>'''

#         with open(plist_path, "w") as f:
#             f.write(plist_content)

#         os.system(f"launchctl load {plist_path}")

#     elif system_name == "Linux":
#         cron_jobs = subprocess.run("crontab -l", shell=True, capture_output=True, text=True, startupinfo=startupinfo).stdout

#         if script_path in cron_jobs:
#             return  # Already scheduled

#         cron_command = f"@reboot {script_path}\n"
#         subprocess.run(f'(crontab -l; echo "{cron_command}") | crontab -', shell=True, startupinfo=startupinfo)

# schedule_startup()

def schedule_startup():
    print('Trying to create a task for ViCTAA')
    system_name = platform.system()
    script_path = os.path.abspath(__file__)  # Auto-detects script path

    if system_name == "Windows":
        task_name = "ViCTAA"
        # Check if the task already exists
        check_task = f'schtasks /query /tn "{task_name}"'
        
        # If the task exists, skip creating it
        if subprocess.run(check_task, shell=True, capture_output=True, startupinfo=startupinfo).returncode == 0:
            print('Skipping as the task is already exist.')
            return  # Task already exists, skip creating
        
        # Otherwise, create the task
        create_task = f'schtasks /create /tn "{task_name}" /tr "{script_path}" /sc onlogon /rl highest /f /IT'
        subprocess.run(create_task, shell=True, startupinfo=startupinfo)

    elif system_name == "Darwin":  # macOS
        plist_path = os.path.expanduser("~/Library/LaunchAgents/com.myflaskapp.autostart.plist")
        
        if os.path.exists(plist_path):
            return  # Already scheduled

        plist_content = f'''<?xml version="1.0" encoding="UTF-8"?>
        <!DOCTYPE plist PUBLIC "-//Apple//DTD PLIST 1.0//EN" "http://www.apple.com/DTDs/PropertyList-1.0.dtd">
        <plist version="1.0">
        <dict>
            <key>Label</key>
            <string>com.myflaskapp.autostart</string>
            <key>ProgramArguments</key>
            <array>
                <string>{script_path}</string>
            </array>
            <key>RunAtLoad</key>
            <true/>
        </dict>
        </plist>'''

        with open(plist_path, "w") as f:
            f.write(plist_content)

        os.system(f"launchctl load {plist_path}")

    elif system_name == "Linux":
        cron_jobs = subprocess.run("crontab -l", shell=True, capture_output=True, text=True, startupinfo=startupinfo).stdout

        if script_path in cron_jobs:
            return  # Already scheduled

        cron_command = f"@reboot {script_path}\n"
        subprocess.run(f'(crontab -l; echo "{cron_command}" startupinfo=startupinfo) | crontab -', shell=True)

schedule_startup()
 
if __name__ == '__main__':
    threading.Timer(1, open_browser).start()
    disable_usb_controllers()
    threading.Timer(1, run_alert_in_thread).start() 
    # Get and print USB ports status
    socketio.run(app, debug=False, port=5001, allow_unsafe_werkzeug=True)
    